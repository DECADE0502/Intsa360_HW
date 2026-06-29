from __future__ import annotations

import json
import re
import sys
from collections import OrderedDict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

ROOT = Path(__file__).resolve().parents[2]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from app.backend.capture_fields import CAPTURE_VISIBLE_PROPERTIES


REF_SPLIT = re.compile(r"[,;\s]+")

BASE_HEADERS = [
    "Item",
    "Quantity",
    "Reference",
    "Part Number",
    "Value",
    "规格型号",
    "器件描述（新整理）",
    "物料名称",
    "等级",
]

PROP_MAP = [
    (["Reference", "Part Reference", "Designator", "位号"], "Reference"),
    (["Part Number", "子项编码", "物料编码", "料号", "PN"], "Part Number"),
    (["Value", "值"], "Value"),
    (["规格型号", "型号", "Model", "MPN"], "规格型号"),
    (["器件描述（新整理）", "器件描述", "内容", "描述", "Description"], "器件描述（新整理）"),
    (["物料名称", "名称", "Part Type", "Name"], "物料名称"),
    (["等级", "物料优选等级", "优选等级"], "等级"),
]


def split_refs(value: object) -> list[str]:
    return [part for part in REF_SPLIT.split(str(value or "").strip()) if part]


def _value(part: dict[str, object], aliases: list[str]) -> str:
    for alias in aliases:
        if alias in part and str(part[alias]).strip():
            return str(part[alias]).strip()
    return ""


def map_props(part: dict[str, object]) -> OrderedDict[str, str]:
    out: OrderedDict[str, str] = OrderedDict()
    consumed: set[str] = set()

    for aliases, target in PROP_MAP:
        out[target] = _value(part, aliases)
        consumed.update(alias for alias in aliases if alias in part)

    for field in CAPTURE_VISIBLE_PROPERTIES:
        if field in BASE_HEADERS:
            continue
        if field not in out:
            out[field] = str(part.get(field, "") or "").strip()
        if field in part:
            consumed.add(field)

    for key, value in part.items():
        norm = str(key).strip()
        if norm and norm not in out and norm not in consumed and norm not in {"Item", "Quantity", "refs"}:
            out[norm] = str(value or "").strip()

    return out


def main() -> None:
    if len(sys.argv) < 3:
        print("usage: convert_cadence_bom.py <parts.json> <output.xlsx>", file=sys.stderr)
        sys.exit(1)

    parts = json.loads(Path(sys.argv[1]).read_text(encoding="utf-8-sig"))
    out = Path(sys.argv[2])

    mapped = [
        map_props(part)
        for part in parts
        if str(part.get("Reference") or part.get("Part Reference") or "").strip()
    ]

    data_headers: list[str] = []
    for header in BASE_HEADERS[3:]:
        if header not in data_headers:
            data_headers.append(header)
    for field in CAPTURE_VISIBLE_PROPERTIES:
        if field not in {"Reference"} and field not in data_headers:
            data_headers.append(field)
    for item in mapped:
        for key in item:
            if key not in {"Reference"} and key not in data_headers:
                data_headers.append(key)

    headers = ["Item", "Quantity", "Reference"] + data_headers

    wb = Workbook()
    ws = wb.active
    ws.title = "BOM"
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")

    for index, item in enumerate(mapped, start=1):
        refs = split_refs(item.get("Reference", ""))
        row = [index, len(refs) or 1, ",".join(refs)]
        for col in headers[3:]:
            row.append(item.get(col, ""))
        ws.append(row)

    for column in ws.columns:
        width = min(max(len(str(cell.value or "")) for cell in column) + 2, 60)
        ws.column_dimensions[column[0].column_letter].width = width
    wb.save(out)


if __name__ == "__main__":
    main()
