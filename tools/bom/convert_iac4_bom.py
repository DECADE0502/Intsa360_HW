from __future__ import annotations

import json
import re
import shutil
from copy import copy
from collections import OrderedDict
from pathlib import Path

from openpyxl import Workbook, load_workbook


ROOT = Path(__file__).resolve().parents[2]
SOURCE_BOM = ROOT / "data" / "raw" / "bom" / "IAC4_MB_POWER_V02_20260618A.xlsx"
PACKAGE_DIR = next(p for p in ROOT.iterdir() if p.is_dir() and p.name.startswith("BOM"))
TEMPLATE = next(PACKAGE_DIR.rglob("203010100819_ERP_BOM导入模板.xlsx"))
OUTPUT_DIR = ROOT / "data" / "outputs" / "bom"

PARENT_CODE = "203010100819"
PARENT_DESC = "IAC4_MB_POWER_V02_PCBA"
OUTPUT_BASENAME = "IAC4_MB_POWER_V02_PCBA_20260618"

HEADERS = [
    "父项编码",
    "描述",
    "子项编码",
    "名称",
    "型号",
    "描述",
    "单位",
    "数量",
    "位号",
    "备注",
    "物料优选等级",
    "物料优选等级备注",
    "替代组编码",
    "替代策略",
    "替代方式",
    "替代优先级",
    "发料方式",
    "是否参与MRP运算",
    "是否跳层",
]

VERSION_RULES = {
    "PVT_VER6": {
        "suffix": "CXDD5DBAM_YMEC",
        "u22": {
            "code": "311040400008",
            "model": "YMEC7C0TG1A2C3",
            "desc": "EMMC,YMEC7C0TG1A2C3,emmc5.1,64GB,3.3V/1.8V,HS400,-25~85℃,BGA153,11.5x13x1.0 mm",
        },
        "u2400": {
            "code": "311040100023",
            "model": "CXDD5DBAM-MC-M",
            "desc": "DRAM,CXDD5DBAM-MC-M,LPDDR5x,4GB,1.8V/1.05V/0.9V/0.5V,8533Mbps,-25~85℃,315-Ball,12.4x15x0.97mm",
        },
        "fit_resistors": {"R88", "R89"},
        "not_fit_resistors": {"R87"},
    },
    "PVT_VER7": {
        "suffix": "FL5P2004G_TEMCG",
        "u22": {
            "code": "311040400014",
            "model": "TEMCG101T30-A0S5",
            "desc": "EMMC,TEMCG101T30-A0S5,emmc5.1,64GB,3.3V/1.8V,HS400,-25~85℃,BGA153",
        },
        "u2400": {
            "code": "311040100013",
            "model": "FL5P2004G-60",
            "desc": "DRAM,FL5P2004G-60,LPDDR5x,4GB",
        },
        "fit_resistors": {"R87", "R88", "R89"},
        "not_fit_resistors": set(),
    },
}


def split_refs(value: object) -> list[str]:
    if value is None:
        return []
    parts = re.split(r"[,;\s]+", str(value).strip())
    return [p for p in parts if p]


def normalize_ref(ref: str) -> str:
    match = re.fullmatch(r"(U\d+)[A-Z]", ref)
    if match:
        return match.group(1)
    return ref


def natural_key(ref: str) -> tuple[str, int, str]:
    match = re.match(r"([A-Za-z_]+)(\d+)(.*)", ref)
    if not match:
        return (ref, -1, "")
    return (match.group(1), int(match.group(2)), match.group(3))


NC_HEADERS = [
    "原始行号",
    "位号",
    "子项编码",
    "物料名称",
    "型号",
    "描述",
    "Value",
    "过滤原因",
    "适用版本",
    "备注",
]


def exclusion_reason(row: dict[str, object], refs: list[str], version: str) -> str | None:
    child_code = str(row.get("Part Number") or "").strip()
    if not child_code:
        return "子项编码为空"

    value = str(row.get("Value") or "").strip()
    version_rule = VERSION_RULES[version]
    version_enabled_refs = version_rule["fit_resistors"]
    if (value == "NC" or value.startswith("NC/")) and not (set(refs) & version_enabled_refs):
        return "NC/未贴"

    forbidden_prefixes = ("JP", "TP", "Z_TP", "SH")
    if any(ref.upper().startswith(forbidden_prefixes) for ref in refs):
        matched = next(ref for ref in refs if ref.upper().startswith(forbidden_prefixes))
        if matched.upper().startswith("JP"):
            return "跳线 JP*"
        if matched.upper().startswith(("TP", "Z_TP")):
            return "测试点 TP*/Z_TP*"
        return "屏蔽/非贴装 SH*"

    text = " ".join(str(v or "") for v in row.values())
    for token in ("Test", "测试点", "跳线"):
        if token in text:
            return f"字段包含 {token}"
    return None


def is_forbidden(row: dict[str, object], refs: list[str], version: str) -> bool:
    return exclusion_reason(row, refs, version) is not None


def load_source_rows(version: str) -> tuple[list[dict[str, object]], list[list[object]]]:
    wb = load_workbook(SOURCE_BOM, data_only=True)
    ws = wb.active
    source_headers = [ws.cell(12, col).value for col in range(1, ws.max_column + 1)]
    rows: list[dict[str, object]] = []
    excluded: list[list[object]] = []
    for row_num in range(15, ws.max_row + 1):
        row = {
            header: ws.cell(row_num, col).value
            for col, header in enumerate(source_headers, start=1)
            if header
        }
        refs = [normalize_ref(ref) for ref in split_refs(row.get("Reference"))]
        reason = exclusion_reason(row, refs, version)
        if reason:
            excluded.append(
                [
                    row_num,
                    ",".join(refs),
                    row.get("Part Number"),
                    row.get("物料名称"),
                    row.get("规格型号"),
                    row.get("器件描述（新整理）"),
                    row.get("Value"),
                    reason,
                    version,
                    None,
                ]
            )
            continue
        rows.append(row)
    return rows, excluded


def apply_version(row: dict[str, object], version: str) -> dict[str, object] | None:
    rule = VERSION_RULES[version]
    refs = {normalize_ref(ref) for ref in split_refs(row.get("Reference"))}
    if refs & rule["not_fit_resistors"]:
        return None

    updated = dict(row)
    if "R87" in refs and "R87" in rule["fit_resistors"]:
        updated["Value"] = "100K/F"

    if "U22" in refs:
        updated["Part Number"] = rule["u22"]["code"]
        updated["规格型号"] = rule["u22"]["model"]
        updated["器件描述（新整理）"] = rule["u22"]["desc"]
    if "U2400" in refs:
        updated["Part Number"] = rule["u2400"]["code"]
        updated["规格型号"] = rule["u2400"]["model"]
        updated["器件描述（新整理）"] = rule["u2400"]["desc"]
    return updated


def build_records(version: str) -> list[list[object]]:
    groups: OrderedDict[tuple[object, ...], dict[str, object]] = OrderedDict()
    source_rows, _ = load_source_rows(version)
    for row in source_rows:
        updated = apply_version(row, version)
        if updated is None:
            continue

        refs = sorted(
            {normalize_ref(ref) for ref in split_refs(updated.get("Reference"))},
            key=natural_key,
        )
        key = (
            str(updated.get("Part Number") or "").strip(),
            str(updated.get("物料名称") or "").strip(),
            str(updated.get("规格型号") or "").strip(),
            str(updated.get("器件描述（新整理）") or "").strip(),
            "ea",
        )

        if key not in groups:
            groups[key] = {
                "child_code": key[0],
                "name": key[1],
                "model": key[2],
                "desc": key[3],
                "unit": key[4],
                "refs": set(),
            }
        groups[key]["refs"].update(refs)

    records = []
    for group in groups.values():
        refs = sorted(group["refs"], key=natural_key)
        records.append(
            [
                PARENT_CODE,
                PARENT_DESC,
                group["child_code"],
                group["name"],
                group["model"],
                group["desc"],
                group["unit"],
                len(refs),
                ",".join(refs),
                None,
                None,
                None,
                None,
                None,
                None,
                None,
                "直接领料",
                "是",
                "否",
            ]
        )
    return records


def clear_data_rows(ws) -> None:
    if ws.max_row >= 3:
        ws.delete_rows(3, ws.max_row - 2)


def write_version(version: str) -> Path:
    output = OUTPUT_DIR / f"{OUTPUT_BASENAME} BOM_{version}.xlsx"
    shutil.copy2(TEMPLATE, output)

    wb = load_workbook(output)
    ws = wb.worksheets[0]
    ws.title = version
    clear_data_rows(ws)

    records = build_records(version)
    for row in records:
        ws.append(row)

    wb.save(output)
    return output


def write_nc_summary(version: str) -> Path:
    _, excluded = load_source_rows(version)
    output = OUTPUT_DIR / f"{OUTPUT_BASENAME} NC未贴器件汇总_{version}.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = f"{version}_NC"
    ws.append(NC_HEADERS)
    for row in excluded:
        ws.append(row)
    for col in range(1, len(NC_HEADERS) + 1):
        font = copy(ws.cell(1, col).font)
        font.bold = True
        ws.cell(1, col).font = font
    wb.save(output)
    return output


def summarize(path: Path) -> dict[str, object]:
    wb = load_workbook(path, data_only=True)
    ws = wb.worksheets[0]
    data_rows = 0
    qty_total = 0
    blank_codes = 0
    refs: set[str] = set()
    for row_num in range(3, ws.max_row + 1):
        values = [ws.cell(row_num, col).value for col in range(1, 20)]
        if all(v is None or str(v).strip() == "" for v in values):
            continue
        data_rows += 1
        qty_total += int(ws.cell(row_num, 8).value or 0)
        if not str(ws.cell(row_num, 3).value or "").strip():
            blank_codes += 1
        refs.update(split_refs(ws.cell(row_num, 9).value))
    return {
        "file": str(path.relative_to(ROOT)),
        "sheet": ws.title,
        "max_column": ws.max_column,
        "data_rows": data_rows,
        "quantity_total": qty_total,
        "blank_child_codes": blank_codes,
        "has_instruction_sheet": len(wb.worksheets) >= 2,
        "merged_ranges": [str(rng) for rng in ws.merged_cells.ranges],
        "has_R87": "R87" in refs,
        "has_R88": "R88" in refs,
        "has_R89": "R89" in refs,
    }


def run_conversion() -> dict[str, object]:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    outputs = []
    nc_summaries = []
    for version in VERSION_RULES:
        outputs.append(write_version(version))
        nc_summaries.append(write_nc_summary(version))
    summary = [summarize(path) for path in outputs]
    summary_path = OUTPUT_DIR / "conversion_summary.json"
    summary_path.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    return {
        "outputs": {
            "main_boms": [str(path.relative_to(ROOT)) for path in outputs],
            "nc_summaries": [str(path.relative_to(ROOT)) for path in nc_summaries],
            "summary": str(summary_path.relative_to(ROOT)),
        },
        "summary": summary,
    }


def main() -> int:
    result = run_conversion()
    summary = result["summary"]
    print(json.dumps(summary, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
