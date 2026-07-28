from __future__ import annotations

import runpy
from pathlib import Path

from openpyxl import load_workbook

from app.backend.bom_semantics.diff import compare_board_boms
from app.backend.bom_semantics.normalization import normalize_workbook
from app.backend.bom_semantics.report_export import REPORT_SHEETS, export_compare_report
from app.backend.bom_semantics.substitutes import build_board_boms


BUILDER = Path(__file__).resolve().parent / "fixtures" / "bom_semantics" / "build_fixtures.py"


def test_layered_report_round_trip_matches_compare_counts(tmp_path: Path) -> None:
    paths = runpy.run_path(str(BUILDER))["build_all"](tmp_path)
    result = compare_board_boms(
        build_board_boms(normalize_workbook(paths["ordinary"])),
        build_board_boms(normalize_workbook(paths["substitutes"])),
    )
    output = export_compare_report(result, tmp_path / "report.xlsx")

    workbook = load_workbook(output, data_only=True)
    try:
        assert tuple(workbook.sheetnames) == REPORT_SHEETS
        assert workbook["业务事件"].max_row - 1 == len(result.events)
        assert workbook["实际贴装差异"].max_row - 1 == len(result.placement_groups)
        assert workbook["实际贴装差异"]["C2"].value
        assert workbook["实际贴装差异"]["D2"].value >= 1
        assert workbook["业务事件"]["B2"].number_format == "@"
        summary = {
            row[0].value: row[1].value
            for row in workbook["对比摘要"].iter_rows(min_row=2, max_col=2)
        }
        assert summary["全部事件数"] == result.summary.changed_event_count
        assert summary["需复核事件数"] == result.summary.review_event_count
        assert summary["非装配变更记录数"] == result.summary.metadata_change_count
        assert summary["非装配字段变化数"] == result.summary.metadata_field_count
    finally:
        workbook.close()
