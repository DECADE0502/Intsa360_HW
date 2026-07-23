from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import PatternFill


PLM_HEADERS = [
    "父项编码",
    "描述",
    "子项编码",
    "名称",
    "型号",
    "描述",
    "单位",
    "数量",
    "位号",
    "备注",
    "物料优选等级",
    "物料优选等级备注",
    "替代组编码",
    "替代策略",
    "替代方式",
    "替代优先级",
    "发料方式",
    "是否参与MRP运算",
    "是否跳层",
]


def _plm_book(path: Path, rows: list[list[object]], *, repeated_header_at: int | None = None) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "BOM导入模版"
    sheet.append(["父项物料", "", "子项物料属性", "", "", "", "", "", "BOM属性", "", "", "", "", "", "", "", "", "", ""])
    sheet.append(PLM_HEADERS)
    for index, row in enumerate(rows, start=3):
        if repeated_header_at is not None and index == repeated_header_at:
            sheet.append(PLM_HEADERS)
        sheet.append(row)
    guide = workbook.create_sheet("说明")
    guide["A1"] = "仅供合成测试"
    workbook.save(path)
    workbook.close()


def build_all(output_dir: Path) -> dict[str, Path]:
    output_dir.mkdir(parents=True, exist_ok=True)
    ordinary = output_dir / "ordinary.xlsx"
    substitutes = output_dir / "substitutes.xlsx"
    multi_parent = output_dir / "multi_parent.xlsx"
    styled = output_dir / "styled.xlsx"

    _plm_book(
        ordinary,
        [
            ["BOARD-A", "Board A", "MAT-R", "电阻", "10K", "电阻 10K", "ea", 2, "R1,R2", "", "优选", "", "", "", "", "", "直接发料", "是", "否"],
            ["BOARD-A", "Board A", "PCB-A", "PCB", "", "测试板", "ea", 1, "", "", "验证中", "", "", "", "", "", "直接领料", "是", "否"],
        ],
    )
    _plm_book(
        substitutes,
        [
            ["BOARD-A", "Board A", "MAT-A", "电容", "1uF", "主料", "ea", 4, "C1,C2,C3,C4", "", "优选", "", "MAT-A", "可替代", "替代", 0, "直接发料", "是", "否"],
            ["BOARD-A", "Board A", "MAT-B", "电容", "1uF", "替代一", "ea", 4, "", "", "正常", "", "MAT-A", "可替代", "替代", 1, "直接发料", "是", "否"],
            ["BOARD-A", "Board A", "MAT-C", "电容", "1uF", "替代二", "ea", 4, "60", "", "正常", "", "MAT-A", "可替代", "替代", 2, "直接发料", "是", "否"],
        ],
    )
    _plm_book(
        multi_parent,
        [
            ["BOARD-A", "Board A", "MAT-A", "电阻", "10K", "A", "ea", 1, "R1", "", "优选", "", "", "", "", "", "直接发料", "是", "否"],
            ["BOARD-B", "Board B", "MAT-B", "电阻", "10K", "B", "ea", 1, "R1", "", "优选", "", "", "", "", "", "直接发料", "是", "否"],
            ["BOARD-B", "Board B", "PCB-B", "PCB", "", "B PCB", "ea", 1, "72", "", "验证中", "", "", "", "", "", "直接领料", "是", "否"],
        ],
        repeated_header_at=4,
    )
    _plm_book(
        styled,
        [
            ["BOARD-A", "Board A", "MAT-R", "电阻", "10K", "电阻 10K", "ea", 2, "R1,R2", "", "优选", "", "", "", "", "", "直接发料", "是", "否"],
        ],
    )
    workbook = Workbook()
    workbook.close()
    styled_book = __import__("openpyxl").load_workbook(styled)
    styled_sheet = styled_book["BOM导入模版"]
    styled_sheet["C3"].fill = PatternFill(fill_type="solid", fgColor="FFF2CC")
    styled_book.save(styled)
    styled_book.close()
    return {
        "ordinary": ordinary,
        "substitutes": substitutes,
        "multi_parent": multi_parent,
        "styled": styled,
    }


if __name__ == "__main__":
    build_all(Path(__file__).resolve().parent / "generated")
