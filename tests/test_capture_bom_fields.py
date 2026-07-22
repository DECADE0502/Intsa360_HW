from __future__ import annotations

import json
import subprocess
import sys
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook

from app.backend.tools import bom_process


ROOT = Path(__file__).resolve().parents[1]
CONVERTER = ROOT / "tools" / "bom" / "convert_cadence_bom.py"
TCL_TEMPLATE = ROOT / "cadence" / "iac_bom_tool.tcl"
FRONTEND_WIZARD = ROOT / "frontend" / "src" / "tools" / "BomProcessWizard.tsx"

VISIBLE_CAPTURE_FIELDS = [
    "Color", "datasheet", "Designator", "Graphic", "ID", "Implementation", "Implementation Path", "Implementation Type",
    "Location X-Coordinate", "Location Y-Coordinate", "Name", "OriginalSymbolOrigin", "Part Number", "Part Reference", "Part Type",
    "PCB Footprint", "PCB封装", "Power Pins Visible", "Primitive", "Reference", "Source Library",
    "Source Package", "Source Part", "SPLIT_INST", "SWAP_INFO", "Value", "等级", "等级备注", "规格型号",
    "器件描述（旧）", "器件描述（新整理）", "物料名称", "制造商",
]

PLM_HEADERS = [
    "父项编码", "描述", "子项编码", "名称", "型号", "描述", "单位", "数量", "位号", "备注",
    "物料优选等级", "物料优选等级备注", "替代组编码", "替代策略", "替代方式", "替代优先级",
    "发料方式", "是否参与MRP运算", "是否跳层",
]


class CaptureBomFieldTests(unittest.TestCase):
    def test_capture_fields_constant_contains_all_visible_properties(self) -> None:
        from app.backend.capture_fields import CAPTURE_VISIBLE_PROPERTIES

        for field in VISIBLE_CAPTURE_FIELDS:
            self.assertIn(field, CAPTURE_VISIBLE_PROPERTIES)

    def test_tcl_prop_names_includes_visible_capture_properties(self) -> None:
        text = TCL_TEMPLATE.read_text(encoding="utf-8")
        start = text.index("variable PROP_NAMES")
        end = text.index("proc shouldProcess")
        prop_block = text[start:end]
        for field in VISIBLE_CAPTURE_FIELDS:
            if field.isascii():
                candidates = [field, f"{{{field}}}", f'"{field}"']
            else:
                escaped = field.encode("unicode_escape").decode("ascii")
                candidates = [f"{{{field}}}", f'"{escaped}"']
            self.assertTrue(
                any(candidate in prop_block for candidate in candidates),
                f"{field!r} missing from Tcl PROP_NAMES block",
            )
        self.assertIn("NewEffectivePropsIter", text)

    def test_tcl_display_dsn_is_a_basename_without_its_extension(self) -> None:
        text = TCL_TEMPLATE.read_text(encoding="utf-8")
        start = text.index("proc DisplayDsnName")
        end = text.index("proc CleanDesignName", start)
        display_proc = text[start:end]
        get_start = text.index("proc GetDsnName")
        get_end = text.index("proc JsonEscape", get_start)
        get_proc = text[get_start:get_end]

        self.assertIn("file tail", display_proc)
        self.assertIn("file rootname", display_proc)
        self.assertIn("return [::IAC::DisplayDsnName $ret]", get_proc)

    def test_frontend_capture_config_contains_practical_fields(self) -> None:
        text = FRONTEND_WIZARD.read_text(encoding="utf-8")
        expected = (
            "{Item}\\\\t{Quantity}\\\\t{Reference}\\\\t{Part Number}\\\\t{Value}\\\\t{规格型号}"
            "\\\\t{器件描述（新整理）}\\\\t{器件描述（旧）}\\\\t{物料名称}\\\\t{等级}\\\\t{等级备注}"
            "\\\\t{制造商}\\\\t{datasheet}\\\\t{PCB Footprint}\\\\t{PCB封装}\\\\t{Part Type}"
            "\\\\t{Part Reference}\\\\t{Name}\\\\t{Designator}\\\\t{Color}\\\\t{Source Library}\\\\t{Source Package}"
            "\\\\t{Source Part}\\\\t{Implementation}\\\\t{Implementation Path}\\\\t{Implementation Type}"
            "\\\\t{Primitive}\\\\t{Graphic}\\\\t{ID}\\\\t{OriginalSymbolOrigin}\\\\t{Power Pins Visible}"
            "\\\\t{Location X-Coordinate}\\\\t{Location Y-Coordinate}\\\\t{SPLIT_INST}\\\\t{SWAP_INFO}"
        )
        self.assertIn(expected, text)

    def test_parser_retains_capture_trace_fields_and_uses_old_description_as_fallback(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            source = Path(tmp) / "capture.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.append([
                "Reference", "Part Number", "器件描述（新整理）", "器件描述（旧）", "制造商",
                "datasheet", "Color", "Source Library", "Implementation Path", "OriginalSymbolOrigin",
            ])
            ws.append([
                "R1", "PN-001", "", "旧版电阻描述", "Vendor-A", "resistor.pdf",
                "Default", "LIB-A", "RES_NP/Normal", "390,985",
            ])
            wb.save(source)

            parsed = bom_process.parse_source(source)

            self.assertEqual(len(parsed.raw_rows), 1)
            row = parsed.raw_rows[0]
            self.assertEqual(row["desc"], "旧版电阻描述")
            self.assertEqual(row["old_desc"], "旧版电阻描述")
            self.assertEqual(row["manufacturer"], "Vendor-A")
            self.assertEqual(row["datasheet"], "resistor.pdf")
            self.assertEqual(row["color"], "Default")
            self.assertEqual(row["source_library"], "LIB-A")
            self.assertEqual(row["implementation_path"], "RES_NP/Normal")
            self.assertEqual(row["original_symbol_origin"], "390,985")

    def test_converter_preserves_visible_capture_properties_in_raw_xlsx(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            src = tmp_path / "parts.json"
            out = tmp_path / "bom.xlsx"
            part = {field: f"{field}-value" for field in VISIBLE_CAPTURE_FIELDS}
            part.update({
                "Reference": "U1",
                "Part Number": "PN-001",
                "Value": "IC",
                "规格型号": "MODEL-A",
                "器件描述（新整理）": "main chip",
                "物料名称": "SoC",
                "PCB封装": "BGA",
                "Source Package": "PKG-A",
                "SPLIT_INST": "1",
            })
            src.write_text(json.dumps([part], ensure_ascii=False), encoding="utf-8")
            subprocess.run([sys.executable, str(CONVERTER), str(src), str(out)], check=True)

            wb = load_workbook(out, read_only=True, data_only=True)
            ws = wb.active
            headers = [ws.cell(1, col).value for col in range(1, ws.max_column + 1)]
            row = [ws.cell(2, col).value for col in range(1, ws.max_column + 1)]
            wb.close()

            for field in VISIBLE_CAPTURE_FIELDS:
                if field != "Reference":
                    self.assertIn(field, headers)
            self.assertEqual(row[headers.index("PCB封装")], "BGA")
            self.assertEqual(row[headers.index("Source Package")], "PKG-A")
            self.assertEqual(row[headers.index("SPLIT_INST")], "1")

    def test_converter_keeps_each_capture_occurrence_in_raw_xlsx(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            src = tmp_path / "parts.json"
            out = tmp_path / "bom.xlsx"
            parts = [
                {
                    "Reference": "R1",
                    "Part Reference": "R1",
                    "Part Number": "PN-001",
                    "Value": "10K",
                    "规格型号": "MODEL-A",
                    "器件描述（新整理）": "resistor",
                    "物料名称": "电阻",
                    "等级": "优选",
                },
                {
                    "Reference": "R2",
                    "Part Reference": "R2",
                    "Part Number": "PN-001",
                    "Value": "10K",
                    "规格型号": "MODEL-A",
                    "器件描述（新整理）": "resistor",
                    "物料名称": "电阻",
                    "等级": "优选",
                },
            ]
            src.write_text(json.dumps(parts, ensure_ascii=False), encoding="utf-8")
            subprocess.run([sys.executable, str(CONVERTER), str(src), str(out)], check=True)

            wb = load_workbook(out, read_only=True, data_only=True)
            ws = wb.active
            headers = [ws.cell(1, col).value for col in range(1, ws.max_column + 1)]
            self.assertEqual(ws.max_row, 3)
            self.assertEqual(ws.cell(2, headers.index("Reference") + 1).value, "R1")
            self.assertEqual(ws.cell(3, headers.index("Reference") + 1).value, "R2")
            self.assertEqual(ws.cell(2, headers.index("Quantity") + 1).value, 1)
            self.assertEqual(ws.cell(3, headers.index("Quantity") + 1).value, 1)
            wb.close()

    def test_bom_process_outputs_19_plm_columns_with_fallbacks_and_preserved_options(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            source = tmp_path / "source.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.append([
                "Item", "Quantity", "Reference", "Part Number", "Value", "规格型号", "器件描述（新整理）",
                "物料名称", "等级", "单位", "备注", "物料优选等级备注", "替代组编码", "替代策略",
                "替代方式", "替代优先级", "发料方式", "是否参与MRP运算", "是否跳层",
                "PCB Footprint", "PCB封装", "Part Type", "Source Package", "SPLIT_INST",
            ])
            ws.append([
                1, 2, "R1 R2", "PN-001", "10K", "", "", "", "优选", "pcs", "keep remark",
                "grade note", "ALT-1", "可替代", "手工替代", "2", "寄售发料", "否", "是",
                "R0402", "0402", "Resistor", "CAPTURE-PKG", "split-a",
            ])
            ws.append([2, 1, "C1", "PN-002", "1uF", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "0603", "Capacitor"])
            wb.save(source)

            parsed = bom_process.parse_source(source)
            result = bom_process.process(parsed, ["plm"], "PARENT", "Parent desc", "TEST", [], tmp_path, "STAMP")
            wb = load_workbook(result["outputs"][0], read_only=True, data_only=True)
            ws = wb.active
            headers = [ws.cell(2, col).value for col in range(1, ws.max_column + 1)]
            first = [ws.cell(3, col).value for col in range(1, ws.max_column + 1)]
            second = [ws.cell(4, col).value for col in range(1, ws.max_column + 1)]
            wb.close()

            self.assertEqual(headers, PLM_HEADERS)
            self.assertEqual(first, ["PARENT", "Parent desc", "PN-001", "Resistor", "10K", "10K", "pcs", 2, "R1,R2", "keep remark", "优选", "grade note", "ALT-1", "可替代", "手工替代", "2", "寄售发料", "否", "是"])
            self.assertEqual(second[3:9], ["Capacitor", "1uF", "1uF", "ea", 1, "C1"])
            self.assertEqual(second[16:19], ["直接发料", "是", "否"])

    def test_trace_fields_do_not_create_part_conflicts(self) -> None:
        rows = [
            {"reference": "R1", "part_number": "PN-001", "name": "Resistor", "model": "10K", "desc": "10K", "grade": "优选", "source_package": "PKG-A"},
            {"reference": "R2", "part_number": "PN-001", "name": "Resistor", "model": "10K", "desc": "10K", "grade": "优选", "source_package": "PKG-B"},
        ]
        self.assertEqual(bom_process.detect_part_conflicts(rows), [])

    def test_bom_process_sanitizes_output_filename_components(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            source = tmp_path / "source.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.append(["Item", "Quantity", "Reference", "Part Number", "Value"])
            ws.append([1, 1, "R1", "PN-001", "10K"])
            wb.save(source)

            parsed = bom_process.parse_source(source)
            result = bom_process.process(parsed, ["plm"], "PARENT", "Parent desc", "BAD?NAME", [], tmp_path, "STAMP?")

            self.assertTrue(Path(result["outputs"][0]).exists())
            self.assertNotIn("?", Path(result["outputs"][0]).name)


if __name__ == "__main__":
    unittest.main()
