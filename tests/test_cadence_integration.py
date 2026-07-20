from __future__ import annotations

import json
import os
import subprocess
import sys
import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

from openpyxl import load_workbook

from app.backend.api import cadence as cadence_api


ROOT = Path(__file__).resolve().parents[1]
CONVERTER = ROOT / "tools" / "bom" / "convert_cadence_bom.py"
TCL_TEMPLATE = ROOT / "cadence" / "iac_bom_tool.tcl"
CADENCE_EXPORT = ROOT / "cadence" / "cadence_export.ps1"
CADENCE_DISCOVERY = ROOT / "scripts" / "lib" / "CadenceDiscovery.ps1"
CADENCE_LIBRARY = ROOT / "scripts" / "lib" / "Cadence.ps1"


class CadenceIntegrationTests(unittest.TestCase):
    def test_redeploy_timeout_becomes_chinese_runtime_error_without_console(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            script = root / "scripts" / "redeploy_cadence_loader.ps1"
            script.parent.mkdir(parents=True)
            script.write_text("# test\n", encoding="ascii")
            powershell = r"C:\Windows\System32\WindowsPowerShell\v1.0\powershell.exe"
            timeout = subprocess.TimeoutExpired([powershell], 30)
            with patch("app.backend.api.cadence.system_powershell", return_value=powershell, create=True), patch.object(
                cadence_api.subprocess,
                "run",
                side_effect=timeout,
            ) as runner:
                with self.assertRaisesRegex(RuntimeError, "超时"):
                    cadence_api.redeploy_cadence_loader(root)

        command = runner.call_args.args[0]
        self.assertEqual(command[0], powershell)
        self.assertEqual(
            runner.call_args.kwargs["creationflags"],
            getattr(subprocess, "CREATE_NO_WINDOW", 0),
        )

    def test_discovery_returns_empty_without_cadence(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            base = Path(tmp)
            home = base / "home"
            profile = base / "profile"
            appdata = base / "appdata"
            localappdata = base / "localappdata"
            home.mkdir()
            profile.mkdir()
            (profile / "AppData" / "Roaming").mkdir(parents=True)
            appdata.mkdir()
            localappdata.mkdir()
            before = sorted(path.relative_to(base).as_posix() for path in base.rglob("*"))
            env = {
                **os.environ,
                "HOME": str(home),
                "USERPROFILE": str(profile),
                "APPDATA": str(appdata),
                "LOCALAPPDATA": str(localappdata),
                "HOMEDRIVE": "",
                "HOMEPATH": "",
                "SPB_DATA": "",
                "CDS_DATA": "",
                "CDSROOT": "",
                "CDS_ROOT": "",
                "CADENCE_ROOT": "",
            }
            script = str(CADENCE_DISCOVERY).replace("'", "''")
            completed = subprocess.run(
                [
                    "powershell.exe",
                    "-NoProfile",
                    "-ExecutionPolicy",
                    "Bypass",
                    "-Command",
                    f". '{script}'; Get-HwAgentCadenceDiscovery | ConvertTo-Json -Depth 6 -Compress",
                ],
                capture_output=True,
                text=True,
                encoding="utf-8",
                errors="replace",
                timeout=30,
                env=env,
            )
            after = sorted(path.relative_to(base).as_posix() for path in base.rglob("*"))

        self.assertEqual(completed.returncode, 0, completed.stdout + completed.stderr)
        payload = json.loads(completed.stdout.strip())
        self.assertEqual(payload["user_autoload_dirs"], [])
        self.assertEqual(payload["vendor_installations"], [])
        self.assertEqual(after, before)

    def test_no_hardcoded_cadence_drive_paths(self) -> None:
        production_sources = [
            *sorted((ROOT / "scripts").rglob("*.ps1")),
            *sorted((ROOT / "app").rglob("*.py")),
        ]

        for source in production_sources:
            text = source.read_text(encoding="utf-8-sig").casefold()
            self.assertNotIn(r"d:\cadence", text, source)
            self.assertNotIn(r"c:\cadence", text, source)

    def test_empty_cadence_deployment_emits_none_marker(self) -> None:
        library = str(CADENCE_LIBRARY).replace("'", "''")
        root = str(ROOT).replace("'", "''")
        python = str(Path(sys.executable)).replace("'", "''")
        completed = subprocess.run(
            [
                "powershell.exe",
                "-NoProfile",
                "-ExecutionPolicy",
                "Bypass",
                "-Command",
                f". '{library}'; Install-CadenceLoader -ToolRoot '{root}' -PythonPath '{python}' -AutoLoadDirs @()",
            ],
            capture_output=True,
            text=True,
            encoding="utf-8",
            errors="replace",
            timeout=30,
        )

        self.assertEqual(completed.returncode, 0, completed.stdout + completed.stderr)
        self.assertIn("__HWAGENT_CADENCE_NONE__", completed.stdout)
        self.assertEqual(cadence_api.parse_cadence_loader_paths(completed.stdout), [])
        self.assertIn("未检测到 Cadence", cadence_api.cadence_redeploy_message(completed.stdout))

    def test_cadence_converter_preserves_all_extra_properties(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            src = tmp_path / "parts.json"
            out = tmp_path / "bom.xlsx"
            src.write_text(
                json.dumps(
                    [
                        {
                            "Reference": "R1",
                            "Part Number": "R.001",
                            "Value": "10K",
                            "PCB Footprint": "R0402",
                            "Part Type": "Resistor",
                            "Manufacturer": "VendorA",
                            "Custom中文属性": "保留我",
                        }
                    ],
                    ensure_ascii=False,
                ),
                encoding="utf-8",
            )

            subprocess.run([sys.executable, str(CONVERTER), str(src), str(out)], check=True)

            wb = load_workbook(out, read_only=True, data_only=True)
            ws = wb.active
            headers = [ws.cell(1, col).value for col in range(1, ws.max_column + 1)]
            row = [ws.cell(2, col).value for col in range(1, ws.max_column + 1)]
            wb.close()

            self.assertIn("Manufacturer", headers)
            self.assertIn("Custom中文属性", headers)
            self.assertEqual(row[headers.index("Manufacturer")], "VendorA")
            self.assertEqual(row[headers.index("Custom中文属性")], "保留我")

    def test_cadence_converter_accepts_utf8_bom_json(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            src = tmp_path / "parts.json"
            out = tmp_path / "bom.xlsx"
            src.write_text(
                json.dumps([{"Reference": "R1", "Part Number": "R.001"}], ensure_ascii=False),
                encoding="utf-8-sig",
            )

            subprocess.run([sys.executable, str(CONVERTER), str(src), str(out)], check=True)

            self.assertTrue(out.exists())

    def test_cadence_converter_uses_union_of_group_columns(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            src = tmp_path / "parts.json"
            out = tmp_path / "bom.xlsx"
            src.write_text(
                json.dumps(
                    [
                        {"Reference": "R1", "Part Number": "R.001", "Part Type": "Resistor", "Description": "10K"},
                        {
                            "Reference": "C1",
                            "Part Number": "C.001",
                            "Part Type": "Capacitor",
                            "Description": "0.1uF",
                            "SecondOnly": "kept",
                        },
                    ],
                    ensure_ascii=False,
                ),
                encoding="utf-8",
            )

            subprocess.run([sys.executable, str(CONVERTER), str(src), str(out)], check=True)

            wb = load_workbook(out, read_only=True, data_only=True)
            ws = wb.active
            headers = [ws.cell(1, col).value for col in range(1, ws.max_column + 1)]
            second = [ws.cell(3, col).value for col in range(1, ws.max_column + 1)]
            wb.close()

            self.assertIn("SecondOnly", headers)
            self.assertEqual(second[headers.index("SecondOnly")], "kept")

    def test_cadence_converter_keeps_capture_bom_header_names(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            src = tmp_path / "parts.json"
            out = tmp_path / "bom.xlsx"
            src.write_text(
                json.dumps(
                    [
                        {
                            "Reference": "C1",
                            "Part Number": "C.001",
                            "Value": "0.1uF",
                            "规格型号": "0402X104K160",
                            "器件描述（新整理）": "陶瓷电容,0.1uF",
                            "物料名称": "电容",
                            "等级": "优选",
                        }
                    ],
                    ensure_ascii=False,
                ),
                encoding="utf-8",
            )

            subprocess.run([sys.executable, str(CONVERTER), str(src), str(out)], check=True)

            wb = load_workbook(out, read_only=True, data_only=True)
            ws = wb.active
            headers = [ws.cell(1, col).value for col in range(1, ws.max_column + 1)]
            row = [ws.cell(2, col).value for col in range(1, ws.max_column + 1)]
            wb.close()

            self.assertEqual(
                headers[:9],
                ["Item", "Quantity", "Reference", "Part Number", "Value", "规格型号", "器件描述（新整理）", "物料名称", "等级"],
            )
            self.assertEqual(row[headers.index("规格型号")], "0402X104K160")
            self.assertEqual(row[headers.index("器件描述（新整理）")], "陶瓷电容,0.1uF")
            self.assertEqual(row[headers.index("物料名称")], "电容")
            self.assertEqual(row[headers.index("等级")], "优选")

    def test_tcl_template_serializes_all_row_properties_without_tcl85_dict(self) -> None:
        text = TCL_TEMPLATE.read_text(encoding="utf-8")

        self.assertIn("proc RowSet", text)
        self.assertIn("proc RowKeys", text)
        self.assertIn("foreach key [::IAC::RowKeys $row]", text)
        self.assertIn("[::IAC::RowGet $row $key]", text)
        self.assertNotIn("dict keys", text)
        self.assertNotIn("dict set", text)
        self.assertNotIn("foreach key {Reference {Part Number} Value {PCB Footprint} {Part Type}}", text)

    def test_tcl_template_gets_design_name_from_effective_props(self) -> None:
        text = TCL_TEMPLATE.read_text(encoding="utf-8")
        start = text.index("proc GetDsnName")
        end = text.index("proc PartsToJson")
        proc = text[start:end]

        self.assertIn("NewEffectivePropsIter", proc)
        self.assertIn('if {$propname eq "Name"}', proc)
        self.assertNotIn('if {$ret ne ""} { return $ret }', proc)
        self.assertIn("return [::IAC::DisplayDsnName $ret]", proc)

    def test_tcl_template_sanitizes_full_dsn_path_for_the_job_directory(self) -> None:
        text = TCL_TEMPLATE.read_text(encoding="utf-8")
        display_start = text.index("proc DisplayDsnName")
        display_end = text.index("proc CleanDesignName")
        display_proc = text[display_start:display_end]
        clean_start = text.index("proc CleanDesignName")
        clean_end = text.index("proc GetDsnName")
        clean_proc = text[clean_start:clean_end]
        export_start = text.index("proc ExportAndProcess")
        export_end = text.index("# ----", export_start)
        export_proc = text[export_start:export_end]

        self.assertIn("file tail", display_proc)
        self.assertIn("file rootname", display_proc)
        self.assertIn('regsub -all {[\\\\/:*?"<>|]}', clean_proc)
        self.assertIn("set dsn [::IAC::GetDsnName]", export_proc)
        self.assertIn("CreateExportJob", export_proc)
        self.assertNotIn('${dsn}.xlsx', export_proc)

    def test_tcl_template_allocates_a_unique_job_directory_and_json_for_each_export(self) -> None:
        text = TCL_TEMPLATE.read_text(encoding="utf-8")
        start = text.index("proc CreateExportJob")
        end = text.index("proc ReportExportFailure", start)
        job_proc = text[start:end]

        self.assertIn('set jobRoot [file normalize "$::IAC_STATE_ROOT/data/jobs"]', job_proc)
        self.assertIn("set processId [pid]", job_proc)
        self.assertIn("set sequence [incr EXPORT_SEQUENCE]", job_proc)
        self.assertIn('set jsonPath [file join $jobDir "parts.json"]', job_proc)
        self.assertIn('set xlsxPath [file join $jobDir "bom.xlsx"]', job_proc)
        self.assertIn("return [list $jobDir $jsonPath $xlsxPath]", job_proc)

    def test_cadence_export_fails_closed_without_reusing_inbox_or_user_workbooks(self) -> None:
        text = TCL_TEMPLATE.read_text(encoding="utf-8")
        export_start = text.index("proc ExportAndProcess")
        export_end = text.index("# ----", export_start)
        export_proc = text[export_start:export_end]
        ps1 = CADENCE_EXPORT.read_text(encoding="utf-8")

        self.assertIn("proc ReportExportFailure", text)
        self.assertIn("tk_messageBox -icon error", text)
        self.assertNotIn("data/inbox", export_proc)
        self.assertNotIn("glob -nocomplain", export_proc)
        self.assertNotIn("Get-ChildItem", ps1)
        self.assertNotIn("Copy-Item", ps1)
        self.assertIn("COM BOM export failed", ps1)

    def test_tcl_template_uses_process_and_sequence_in_each_job_path_for_parallel_exports(self) -> None:
        text = TCL_TEMPLATE.read_text(encoding="utf-8")
        start = text.index("proc CreateExportJob")
        end = text.index("proc ReportExportFailure", start)
        job_proc = text[start:end]

        self.assertIn("variable EXPORT_SEQUENCE 0", text)
        self.assertIn("set processId [pid]", job_proc)
        self.assertIn("set sequence [incr EXPORT_SEQUENCE]", job_proc)
        self.assertIn('set jobName "${safeDsn}-${stamp}-${processId}-${sequence}"', job_proc)

    def test_tcl_template_uses_capture_sample_property_accessors(self) -> None:
        text = TCL_TEMPLATE.read_text(encoding="utf-8")
        start = text.index("proc GetReference")
        end = text.index("proc CleanDesignName")
        proc = text[start:end]

        self.assertIn("GetReference", proc)
        self.assertIn("GetEffectivePropStringValue", proc)
        self.assertIn("GetVariantProp", proc)
        self.assertIn("IsPrimitive", text)

    def test_tcl_template_uses_hidden_launcher_and_ascii_default_menu(self) -> None:
        text = TCL_TEMPLATE.read_text(encoding="utf-8")

        self.assertIn("launch_tool_suite_hidden.vbs", text)
        self.assertIn("wscript.exe", text)
        self.assertIn('"insta360_HW"', text)
        self.assertIn('"action" "Open Platform"', text)
        self.assertIn('"action" "Export and Process BOM"', text)
        self.assertNotIn('AddAccessoryMenu "insta360_HW" "Open Platform"', text)
        self.assertNotIn('AddAccessoryMenu "insta360_HW" "Export and Process BOM"', text)
        self.assertNotIn('AddAccessoryMenu "insta360_HW" "进入平台"', text)
        self.assertNotIn('AddAccessoryMenu "insta360_HW" "导出并处理BOM"', text)

    def test_tcl_template_exposes_capture_command_window_diagnostics(self) -> None:
        text = TCL_TEMPLATE.read_text(encoding="utf-8")

        self.assertIn("proc Diagnose", text)
        self.assertIn("proc iacdiag", text)
        self.assertIn("IAC: diagnostics", text)
        self.assertIn("info commands RegisterAction", text)
        self.assertIn("launch_tool_suite_hidden.vbs", text)
        self.assertIn("convert_cadence_bom.py", text)
        self.assertIn("IAC_ROOT", text)

    def test_tcl_template_writes_capture_loader_probe_log(self) -> None:
        text = TCL_TEMPLATE.read_text(encoding="utf-8")

        self.assertIn("proc Probe", text)
        self.assertIn("cadence_loader_probe.log", text)
        self.assertIn("IAC: loader probe", text)
        self.assertIn("RegisterAction=", text)
        self.assertIn("InsertXMLMenu=", text)
        self.assertIn("AddAccessoryMenu=", text)
        self.assertIn("::IAC::Probe", text)

    def test_tcl_template_does_not_load_custom_enhanced_tools_until_platform_registry_exists(self) -> None:
        text = TCL_TEMPLATE.read_text(encoding="utf-8")

        self.assertNotIn("LoadEnhancedTools", text)
        self.assertNotIn("orcad_enhanced_tools.tcl", text)
        self.assertNotIn("rename RegisterAction", text)
        self.assertNotIn("proc ::RegisterAction", text)
        self.assertNotIn('AddAccessoryMenu "insta360_HW" "选中器件切换NC"', text)
        self.assertNotIn('AddAccessoryMenu "insta360_HW" "其他脚本"', text)


if __name__ == "__main__":
    unittest.main()

