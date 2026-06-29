from __future__ import annotations

import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook

from app.backend.tools import analysis_tools, netlist_tools


ROOT = Path(__file__).resolve().parents[1]


def write_netlist(folder: Path, nets: str = "", parts: str = "") -> None:
    folder.mkdir(parents=True, exist_ok=True)
    (folder / "pstxnet.dat").write_text(nets, encoding="utf-8")
    (folder / "pstxprt.dat").write_text(parts, encoding="utf-8")


class NetlistAnalysisTests(unittest.TestCase):
    def test_parse_real_pstxnet_keeps_ref_pin_nodes(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            folder = Path(tmp)
            write_netlist(
                folder,
                nets="""FILE_TYPE = EXPANDEDNETLIST;
NET_NAME
'USB_DP'
 '@DSN(SCH_1):USB_DP':
 C_SIGNAL='usb_dp';
NODE_NAME\tR80 2
 '@DSN(SCH_1):INS1@LIB.RES.NORMAL(CHIPS)':
 '2':;
NODE_NAME\tU400 U24
 '@DSN(SCH_1):INS2@LIB.IC.NORMAL(CHIPS)':
 'GPIO3_16':;
""",
            )

            nets = analysis_tools._parse_net_file(folder)

            self.assertEqual(nets["USB_DP"]["refs"], ["R80", "U400"])
            self.assertEqual(nets["USB_DP"]["nodes"], ["R80.2", "U400.U24"])
            self.assertEqual(nets["USB_DP"]["pins"], ["2", "U24"])

    def test_parse_real_pstxprt_reads_ref_and_package(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            folder = Path(tmp)
            write_netlist(
                folder,
                parts="""FILE_TYPE = EXPANDEDPARTLIST;
PART_NAME
 C1 'CAP_NP_C0201-0P4-B_1UF/6.3V':;

SECTION_NUMBER 1
 '@DSN(SCH_1):INS1@LIB.CAP.NORMAL(CHIPS)':
 C_PATH='x';
PART_NAME
 U400 'A380H_BGA356':;
""",
            )

            parts = analysis_tools._parse_part_file(folder)

            self.assertEqual(parts["C1"], "CAP_NP_C0201-0P4-B_1UF/6.3V")
            self.assertEqual(parts["U400"], "A380H_BGA356")

    def test_netlist_compare_reports_pin_level_difference(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            left = root / "left"
            right = root / "right"
            write_netlist(
                left,
                nets="""NET_NAME
'USB_DP'
NODE_NAME R80 1
NODE_NAME U400 U24
""",
                parts="PART_NAME\n R80 'RES_NP_R0201_10K':;\n",
            )
            write_netlist(
                right,
                nets="""NET_NAME
'USB_DP'
NODE_NAME R80 2
NODE_NAME U400 U24
""",
                parts="PART_NAME\n R80 'RES_NP_R0201_10K':;\n",
            )

            result = analysis_tools.run_netlist_compare(ROOT, {"netlist1": str(left), "netlist2": str(right), "output_dir": str(root)})

            self.assertEqual(result["status"], "ok")
            self.assertGreaterEqual(result["summary"]["node_diffs"], 1)
            headers = result["table"]["headers"]
            self.assertIn("网表1节点", headers)
            self.assertIn("网表2节点", headers)
            row = next(row for row in result["table"]["rows"] if row[0] == "USB_DP")
            self.assertIn("R80.1", row[2])
            self.assertIn("R80.2", row[3])
            self.assertEqual(row[-1], "网络节点差异")

    def test_single_network_report_includes_nodes_and_pin_count(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            folder = root / "net"
            write_netlist(
                folder,
                nets="""NET_NAME
'NC_1'
NODE_NAME R1 1
NET_NAME
'ONE_REF_TWO_PINS'
NODE_NAME U1 A1
NODE_NAME U1 B2
NET_NAME
'NORMAL'
NODE_NAME R1 2
NODE_NAME C1 1
""",
            )

            result = analysis_tools.run_single_network_check(ROOT, {"netlist": str(folder), "output_dir": str(root)})

            self.assertEqual(result["status"], "ok")
            headers = result["table"]["headers"]
            self.assertEqual(headers, ["网络", "类型", "位号", "节点/Pin", "位号数", "节点数"])
            rows = result["table"]["rows"]
            self.assertTrue(any(row[0] == "NC_1" and row[3] == "R1.1" for row in rows))
            self.assertTrue(any(row[0] == "ONE_REF_TWO_PINS" and row[3] == "U1.A1,U1.B2" for row in rows))

    def test_smt_package_check_uses_normalized_package_tokens(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            folder = root / "net"
            write_netlist(folder, parts="PART_NAME\n C1 'CAP_NP_C0201-0P4-B_1UF/6.3V':;\n")

            bom = root / "bom.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.append(["位号", "编号", "描述", "数量", "名称", "封装名"])
            ws.append(["C1", "C.001", "陶瓷电容,1UF/6.3V", 1, "电容", "C0201-0P4-B"])
            wb.save(bom)

            result = analysis_tools.run_smt_package_check(ROOT, {"netlist": str(folder), "bom": str(bom), "output_dir": str(root)})

            self.assertEqual(result["status"], "ok")
            self.assertEqual(result["summary"]["passed_count"], 1)
            self.assertEqual(result["table"]["rows"][0][-2], "机器初筛通过")

    def test_netlist_tools_module_exposes_compatible_runner(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            left = root / "left"
            right = root / "right"
            write_netlist(left, nets="NET_NAME\n'N1'\nNODE_NAME R1 1\n", parts="PART_NAME\n R1 'R0201':;\n")
            write_netlist(right, nets="NET_NAME\n'N1'\nNODE_NAME R1 2\n", parts="PART_NAME\n R1 'R0201':;\n")

            result = netlist_tools.run_netlist_compare(
                ROOT,
                {"netlist1": str(left), "netlist2": str(right), "output_dir": str(root)},
            )

            self.assertEqual(result["status"], "ok")
            self.assertEqual(result["tool"], "netlist_compare")
            self.assertGreaterEqual(result["summary"]["diff_count"], 1)


    def test_smt_package_check_builds_review_for_missing_extra_multi_package_and_high_risk(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            folder = root / "allegro"
            write_netlist(
                folder,
                parts="""PART_NAME
 C1 'CAP_NP_C0201-0P4-B_1UF/6.3V':;
PART_NAME
 C2 'CAP_NP_C0402-0P4-B_1UF/6.3V':;
PART_NAME
 U1 'EMMC_BGA153':;
PART_NAME
 R9 'RES_NP_R0201_10K':;
""",
            )

            bom = root / "bom.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.append(["位号", "编号", "描述", "数量", "名称", "封装名", "规格型号", "等级"])
            ws.append(["C1,C2", "C.001", "陶瓷电容,1UF/6.3V", 2, "电容", "C0201-0P4-B", "1UF", "A"])
            ws.append(["U1", "IC.001", "eMMC 存储器 BGA153", 1, "存储器", "BGA153", "EMMC", "A"])
            ws.append(["L5", "L.001", "电感 2.2UH", 1, "电感", "L0402", "2.2UH", "A"])
            wb.save(bom)

            result = analysis_tools.run_smt_package_check(ROOT, {"netlist": str(folder), "bom": str(bom), "output_dir": str(root)})

            self.assertEqual(result["status"], "ok")
            review = result["smt_package_review"]
            counts = review["status_counts"]
            self.assertEqual(counts["missing_bom"], 1)
            self.assertEqual(counts["extra_bom"], 1)
            self.assertEqual(counts["multi_package"], 1)
            self.assertGreaterEqual(counts["high_risk"], 1)
            self.assertGreaterEqual(result["summary"]["manual_count"], 1)
            self.assertGreaterEqual(result["summary"]["high_risk"], 1)
            self.assertTrue(any(item["status"] == "BOM 缺位号" and item["ref"] == "R9" for item in review["items"]))
            self.assertTrue(any(item["status"] == "BOM 多余位号" and item["ref"] == "L5" for item in review["items"]))
            self.assertTrue(any(item["status"] == "同料多封装" and item["part_number"] == "C.001" for item in review["focus_items"]))
            self.assertTrue(any(item["status"] == "高风险封装" and item["ref"] == "U1" for item in review["items"]))

    def test_smt_package_check_accepts_capture_bom_headers_with_braces(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            folder = root / "allegro"
            write_netlist(folder, parts="PART_NAME\n C1 'CAP_NP_C0201-0P4-B_1UF/6.3V':;\n")

            bom = root / "capture.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.append(["Bill Of Materials"])
            ws.append([""])
            ws.append(["{Item}", "{Quantity}", "{Reference}", "{Part Number}", "{Value}", "{规格型号}", "{器件描述（新整理）}", "{物料名称}", "{等级}", "{PCB封装}", "{Designator}"])
            ws.append([1, 1, "C1", "C.001", "1UF", "C0201-0P4-B", "陶瓷电容,1UF/6.3V", "电容", "A", "C0201-0P4-B", ""])
            wb.save(bom)

            result = analysis_tools.run_smt_package_check(ROOT, {"netlist": str(folder), "bom": str(bom), "output_dir": str(root)})

            self.assertEqual(result["status"], "ok")
            self.assertEqual(result["summary"]["passed_count"], 1)

    def test_smt_package_check_treats_missing_nc_parts_as_skipped_not_bom_missing(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            folder = root / "allegro"
            write_netlist(
                folder,
                parts="""PART_NAME
 C1 'CAP_NP_C0201-0P4-B_1UF/6.3V':;
PART_NAME
 C2 'CAP_NP_C0201-0P4-B_NC/0.1UF/10V':;
""",
            )

            bom = root / "pcba.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.append(["父项编码", "描述", "子项编码", "名称", "型号", "描述", "单位", "数量", "位号"])
            ws.append(["PCBA", "demo", "C.001", "", "C0201-0P4-B", "陶瓷电容,1UF/6.3V", "ea", 1, "C1"])
            wb.save(bom)

            result = analysis_tools.run_smt_package_check(ROOT, {"netlist": str(folder), "bom": str(bom), "output_dir": str(root)})

            self.assertEqual(result["status"], "ok")
            review = result["smt_package_review"]
            self.assertEqual(review["status_counts"]["missing_bom"], 0)
            self.assertEqual(review["status_counts"]["nc_skipped"], 1)
            self.assertEqual(result["summary"]["manual_count"], 0)
            self.assertFalse(any(item["ref"] == "C2" for item in review["focus_items"]))

    def test_smt_package_check_accepts_processed_oa_bom_headers(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            folder = root / "allegro"
            write_netlist(folder, parts="PART_NAME\n C1 'CAP_NP_C0201-0P4-B_1UF/6.3V':;\n")

            bom = root / "oa.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.append([
                "序号",
                "编码（父）*##bmf",
                "描述（父）##msf",
                "编码（子）*##bm",
                "描述（子）##ms",
                "数量*##sl",
                "单位*##dw",
                "位号##wh",
                "备注##bz",
                "物料优选等级##tdyxj",
            ])
            ws.append(["", "PCBA", "demo", "C.001", "陶瓷电容,1UF/6.3V,C0201", 1, "ea", "C1", "", "优选"])
            wb.save(bom)

            result = analysis_tools.run_smt_package_check(ROOT, {"netlist": str(folder), "bom": str(bom), "output_dir": str(root)})

            self.assertEqual(result["status"], "ok")
            self.assertEqual(result["summary"]["passed_count"], 1)
            self.assertEqual(result["summary"]["manual_count"], 0)

    def test_smt_package_check_treats_missing_non_smt_parts_as_skipped_not_bom_missing(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            folder = root / "allegro"
            write_netlist(
                folder,
                parts="""PART_NAME
 C1 'CAP_NP_C0201-0P4-B_1UF/6.3V':;
PART_NAME
 JP1 'SHORT_L3_SP2-L1_SHORT_L1':;
PART_NAME
 TP1 'TP_NP_TP0P4_TP0P4':;
PART_NAME
 H1 '1516A_2_HOLE_PC2P6DC1P6_302020400107':;
""",
            )

            bom = root / "pcba.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.append(["父项编码", "描述", "子项编码", "名称", "型号", "描述", "单位", "数量", "位号"])
            ws.append(["PCBA", "demo", "C.001", "", "C0201-0P4-B", "陶瓷电容,1UF/6.3V", "ea", 1, "C1"])
            wb.save(bom)

            result = analysis_tools.run_smt_package_check(ROOT, {"netlist": str(folder), "bom": str(bom), "output_dir": str(root)})

            self.assertEqual(result["status"], "ok")
            review = result["smt_package_review"]
            self.assertEqual(review["status_counts"]["missing_bom"], 0)
            self.assertEqual(review["status_counts"]["non_smt_skipped"], 3)
            self.assertEqual(result["summary"]["manual_count"], 0)
            self.assertEqual(review["focus_items"], [])

    def test_smt_package_check_does_not_put_matched_high_risk_parts_in_default_focus(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            folder = root / "allegro"
            write_netlist(folder, parts="PART_NAME\n U1 'EMMC_BGA153':;\n")

            bom = root / "pcba.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.append(["父项编码", "描述", "子项编码", "名称", "型号", "描述", "单位", "数量", "位号"])
            ws.append(["PCBA", "demo", "IC.001", "", "EMMC BGA153", "eMMC 存储器 BGA153", "ea", 1, "U1"])
            wb.save(bom)

            result = analysis_tools.run_smt_package_check(ROOT, {"netlist": str(folder), "bom": str(bom), "output_dir": str(root)})

            self.assertEqual(result["status"], "ok")
            review = result["smt_package_review"]
            self.assertEqual(review["status_counts"]["high_risk"], 1)
            self.assertEqual(result["summary"]["manual_count"], 0)
            self.assertEqual(review["focus_items"], [])


    def test_smt_package_check_matches_common_metric_size_codes_inside_supplier_models(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            folder = root / "net"
            write_netlist(
                folder,
                parts="""PART_NAME
 C1 'CAP_NP_C0201-0P4-B_1UF/6.3V':;
PART_NAME
 R1 'RES_NP_R0201-0P26-A_2.2K/F':;
PART_NAME
 C2 'CAP_NP_C0402-0P7-D_4.7UF/25V':;
""",
            )

            bom = root / "pcba.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.append(["父项编码", "描述", "子项编码", "名称", "型号", "描述", "单位", "数量", "位号"])
            ws.append(["PCBA", "demo", "C.001", "", "CL03A105MQ3CSNH,0201X105M6R3NT", "", "ea", 1, "C1"])
            ws.append(["PCBA", "demo", "R.001", "", "RC0201FR-072K2L,0201WMF2201TEE", "贴片电阻", "ea", 1, "R1"])
            ws.append(["PCBA", "demo", "C.002", "", "TDK105CBJ475MV-F,GRM155R61E475ME1#,CL05A475MA5NUNC", "", "ea", 1, "C2"])
            wb.save(bom)

            result = analysis_tools.run_smt_package_check(ROOT, {"netlist": str(folder), "bom": str(bom), "output_dir": str(root)})

            self.assertEqual(result["status"], "ok")
            self.assertEqual(result["summary"]["passed_count"], 3)
            self.assertEqual(result["summary"]["manual_count"], 0)


if __name__ == "__main__":
    unittest.main()
