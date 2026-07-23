from __future__ import annotations

import runpy
from dataclasses import replace
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

from app.backend.bom_semantics.models import FindingSeverity, ValidationFinding
from app.backend.bom_semantics.normalization import normalize_workbook
from app.backend.bom_semantics.plm_export import (
    PLMExportBlockedError,
    collect_plm_export_blockers,
    export_plm_template,
    verify_plm_export,
)
from app.backend.bom_semantics.substitutes import build_board_boms


BUILDER = Path(__file__).resolve().parent / "fixtures" / "bom_semantics" / "build_fixtures.py"


def _fixtures(tmp_path: Path) -> dict[str, Path]:
    return runpy.run_path(str(BUILDER))["build_all"](tmp_path)


def _boards(path: Path):
    return build_board_boms(normalize_workbook(path))


def _reordered_template(path: Path) -> tuple[list[str], set[str]]:
    headers = [
        "数量",
        "子项编码",
        "父项编码",
        "父项描述",
        "位号",
        "描述",
        "名称",
        "型号",
        "单位",
        "备注",
        "物料优选等级",
        "物料优选等级备注",
        "替代组编码",
        "替代策略",
        "替代方式",
        "替代优先级",
        "发料方式",
        "是否参与MRP运算",
        "是否跳层",
    ]
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Reordered PLM"
    sheet.merge_cells("A1:D1")
    sheet["A1"] = "Template title"
    sheet["A1"].font = Font(bold=True)
    for column, header in enumerate(headers, start=1):
        sheet.cell(2, column).value = header
        sheet.cell(2, column).font = Font(bold=True)
    for column in range(1, len(headers) + 1):
        cell = sheet.cell(3, column)
        cell.fill = PatternFill(fill_type="solid", fgColor="FFF2CC")
        cell.number_format = "General"
    sheet["B3"] = "OLD-MATERIAL"
    sheet["C3"] = "OLD-PARENT"
    sheet["A3"] = 1
    sheet.merge_cells("A5:B5")
    sheet["A5"] = "KEEP TAIL INSTRUCTION"
    sheet.freeze_panes = "A3"
    guide = workbook.create_sheet("Guide")
    guide["A1"] = "Keep this helper sheet unchanged"
    workbook.save(path)
    workbook.close()
    return headers, {"A1:D1", "A7:B7"}


def test_export_preserves_reordered_template_and_round_trips_semantics(tmp_path: Path) -> None:
    boards = _boards(_fixtures(tmp_path)["substitutes"])
    template = tmp_path / "reordered_template.xlsx"
    output = tmp_path / "semantic_export.xlsx"
    headers, merged_ranges = _reordered_template(template)

    result = export_plm_template(boards, template, output)

    assert result.output_path == output.resolve()
    assert result.rows_written == 3
    assert result.parent_codes == ("BOARD-A",)
    workbook = load_workbook(output, data_only=False)
    try:
        sheet = workbook["Reordered PLM"]
        assert workbook.sheetnames == ["Reordered PLM", "Guide"]
        assert [sheet.cell(2, index).value for index in range(1, len(headers) + 1)] == headers
        assert {str(cell_range) for cell_range in sheet.merged_cells.ranges} == merged_ranges
        assert sheet.freeze_panes == "A3"
        assert sheet["A5"].fill.fgColor.rgb == "00FFF2CC"
        assert sheet["A7"].value == "KEEP TAIL INSTRUCTION"
        assert sheet["C3"].value == "BOARD-A"
        assert sheet["C3"].data_type == "s"
        assert sheet["C3"].number_format == "@"
        assert sheet["B3"].value == "MAT-A"
        assert sheet["B3"].data_type == "s"
        assert sheet["B4"].value == "MAT-B"
        assert sheet["B5"].value == "MAT-C"
        assert sheet["E3"].value == "C1,C2,C3,C4"
        assert sheet["E4"].value is None
        assert sheet["E5"].value is None
        assert workbook["Guide"]["A1"].value == "Keep this helper sheet unchanged"
    finally:
        workbook.close()

    round_tripped = verify_plm_export(output, boards)
    group = round_tripped[0].substitute_groups[0]
    assert group.main_item is not None
    assert group.main_item.material_code == "MAT-A"
    assert [item.material_code for item in group.alternative_items] == ["MAT-B", "MAT-C"]
    assert group.physical_references == ("C1", "C2", "C3", "C4")


def test_export_rejects_board_blockers_before_creating_a_file(tmp_path: Path) -> None:
    board = _boards(_fixtures(tmp_path)["ordinary"])[0]
    blocked = replace(
        board,
        findings=(
            ValidationFinding(
                code="test_blocker",
                severity=FindingSeverity.BLOCKER,
                message="test",
                parent_code=board.parent_code,
            ),
        ),
    )
    output = tmp_path / "must_not_exist.xlsx"

    with pytest.raises(PLMExportBlockedError) as error:
        export_plm_template(blocked, _fixtures(tmp_path)["ordinary"], output)

    assert error.value.findings[0].code == "test_blocker"
    assert not output.exists()


def test_export_rechecks_main_reference_structure_at_output_boundary(tmp_path: Path) -> None:
    board = _boards(_fixtures(tmp_path)["substitutes"])[0]
    group = board.substitute_groups[0]
    invalid_group = replace(group, physical_references=("C1",))
    invalid_board = replace(board, substitute_groups=(invalid_group,))

    blockers = collect_plm_export_blockers(invalid_board)

    assert {finding.code for finding in blockers} >= {"substitute_group_physical_references_mismatch"}
