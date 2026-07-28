from __future__ import annotations

import os
from pathlib import Path

import pytest
from openpyxl import load_workbook

from app.backend.tools.bom_risk import run_bom_risk_check
from app.backend.tools.bom_risk_model import build_risk_model
from app.backend.tools.bom_rules import (
    evaluate_bom_risk_report,
    risk_rule_config,
)
from tests.fixtures.bom_semantics.build_fixtures import build_all


def _finding(report: dict[str, object], code: str) -> dict[str, object]:
    findings = report["findings"]
    assert isinstance(findings, list)
    return next(item for item in findings if item["code"] == code)


def test_risk_model_excludes_substitute_alternatives_from_placement_count(
    tmp_path: Path,
) -> None:
    source = build_all(tmp_path)["substitutes"]

    model = build_risk_model(source)

    assert len(model.substitute_groups) == 1
    assert len(model.actual_references) == 4
    assert {row.part_number for row in model.placement_rows} == {"MAT-A"}
    assert sum(row.is_substitute_alternative for row in model.rows) == 2


def test_risk_report_uses_configured_property_keywords_without_part_whitelists(
    tmp_path: Path,
) -> None:
    source = build_all(tmp_path)["ordinary"]
    model = build_risk_model(source)
    config = risk_rule_config({"mechanical_keywords": ["10K"]})

    report = evaluate_bom_risk_report(model, config=config)

    mechanical = _finding(report, "mechanical_items")
    assert mechanical["detail_count"] == 1
    assert mechanical["level"] == "info"


def test_risk_check_is_standalone_and_writes_a_readable_multisheet_report(
    tmp_path: Path,
) -> None:
    source = build_all(tmp_path)["substitutes"]

    result = run_bom_risk_check(tmp_path, {"bom": str(source)})

    assert result["status"] == "ok"
    assert result["risk_report"]["profile"] == "plm_single_board"
    assert result["risk_report"]["substitute_groups"]
    assert result["risk_report"]["decision_manifest"] == ""
    output = Path(result["outputs"][0])
    workbook = load_workbook(output, read_only=True, data_only=True)
    try:
        assert len(workbook.sheetnames) == 7
        assert all(workbook[name].max_row >= 1 for name in workbook.sheetnames)
    finally:
        workbook.close()


GOLDEN_EXPECTATIONS = {
    "203010100836_IAC4_MB_POWER_V03_PCBA_BOM.xlsx": (122, 893, 0),
    "203010100819_IAC4_MB_POWER_V02_PCBA_BOM.xlsx": (122, 892, 0),
    "203010100900_IAC4_MB_V10_PCBA_系统导出旧BOM_20260728.xlsx": (143, 800, 26),
}


@pytest.mark.skipif(
    not os.environ.get("HW_BOM_RISK_GOLDEN_DIR"),
    reason="Set HW_BOM_RISK_GOLDEN_DIR to run read-only production BOM checks.",
)
@pytest.mark.parametrize(
    ("filename", "expected"),
    GOLDEN_EXPECTATIONS.items(),
)
def test_readonly_golden_bom_risk_contract(
    tmp_path: Path,
    filename: str,
    expected: tuple[int, int, int],
) -> None:
    root = Path(os.environ["HW_BOM_RISK_GOLDEN_DIR"])
    matches = list(root.rglob(filename))
    assert len(matches) == 1, f"Expected one golden file named {filename}, found {matches}"

    model = build_risk_model(matches[0])
    report = evaluate_bom_risk_report(model)
    expected_rows, expected_refs, expected_groups = expected

    assert len(model.rows) == expected_rows
    assert len(model.actual_references) == expected_refs
    assert len(model.substitute_groups) == expected_groups
    assert report["profile"] == "plm_single_board"
    assert not any(
        item["level"] == "blocker"
        for item in report["findings"]
    )
