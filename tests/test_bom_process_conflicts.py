from __future__ import annotations

import json
import tempfile
import unittest
from pathlib import Path
from unittest import mock

from openpyxl import Workbook, load_workbook

from app.backend.tools import bom_process
from app.backend.tools.analysis_tools import run_bom_process
from fixture_builders import build_capture_bom


FIXTURES = Path(__file__).parent / "fixtures" / "bom"


def record_signature(record: dict[str, object]) -> dict[str, str]:
    return {
        "model": str(record.get("model") or ""),
        "description": str(record.get("desc") or ""),
        "name": str(record.get("name") or ""),
        "grade": str(record.get("grade") or ""),
        "unit": str(record.get("unit") or ""),
    }


def make_source(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.append(["Item", "Quantity", "Reference", "Part Number", "Value", "规格型号", "器件描述（新整理）", "物料名称", "等级"])
    ws.append([1, 1, "R1", "P1", "10K", "M1", "D1", "电阻", "正常"])
    ws.append([2, 1, "R2", "P1", "10K", "M2", "D2", "电阻", "优选"])
    ws.append([3, 1, "C1", "P2", "1uF", "CM1", "CD1", "电容", "正常"])
    wb.save(path)


class BomProcessConflictTests(unittest.TestCase):
    def test_missing_part_number_materials_require_review_and_can_be_restored(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = root / "source.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.append(["Reference", "Part Number", "Value", "PCB Footprint", "Source Package"])
            refs = ("ANCHOR1", "ANCHOR2", "ANCHOR3", "ANCHOR4")
            recovered_code = "123456789012"
            for ref in refs:
                ws.append([ref, "", recovered_code, "MECH_PAD", "MECH_LIB"])
            wb.save(source)

            params: dict[str, object] = {
                "source_bom": str(source),
                "formats": ["plm"],
                "parent_code": "203010100819",
                "name": "HARDWARE_PARTS",
            }
            review = run_bom_process(root, params)

            self.assertEqual(review["status"], "needs_confirmation")
            self.assertEqual(review["reason"], "placement_review")
            self.assertEqual(len(review["groups"]), 1)
            candidate = review["groups"][0]
            self.assertEqual(candidate["suggested_code"], recovered_code)
            self.assertEqual(candidate["recommended_action"], "keep")
            self.assertEqual(candidate["refs"], list(refs))

            params["placement_resolutions"] = {
                candidate["key"]: {
                    "action": "keep",
                    "part_number": recovered_code,
                    "field_patch": {
                        "name": "焊接结构件",
                        "model": "MECH-A",
                        "desc": "焊接结构件",
                    },
                }
            }
            result = run_bom_process(root, params)

            self.assertEqual(result["status"], "ok")
            self.assertEqual(result["summary"]["placement_review"]["kept_groups"], 1)
            plm_path = next(Path(path) for path in result["outputs"] if path.endswith("_PLM_BOM.xlsx"))
            nc_path = next(Path(path) for path in result["outputs"] if path.endswith("_NC未贴汇总.xlsx"))
            plm = load_workbook(plm_path, data_only=True)
            nc = load_workbook(nc_path, data_only=True)
            try:
                plm_rows = list(plm.active.iter_rows(min_row=3, values_only=True))
                nc_rows = list(nc.active.iter_rows(min_row=2, values_only=True))
            finally:
                plm.close()
                nc.close()

            restored = next(row for row in plm_rows if row[2] == recovered_code)
            self.assertEqual(restored[3], "焊接结构件")
            self.assertEqual(restored[4], "MECH-A")
            self.assertEqual(restored[5], "焊接结构件")
            self.assertEqual(restored[7], 4)
            self.assertEqual(restored[8], ",".join(refs))
            self.assertFalse(any(refs[0] in str(row[1] or "") for row in nc_rows))

    def test_missing_part_number_material_can_be_explicitly_confirmed_not_installed(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = root / "source.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.append(["Reference", "Part Number", "Value", "PCB Footprint", "Source Package"])
            ws.append(["LINK1", "", "Short_L1", "sp2-L1", "Short_L3"])
            wb.save(source)

            params: dict[str, object] = {"source_bom": str(source), "formats": ["plm"], "name": "NO_STUFF"}
            review = run_bom_process(root, params)
            self.assertEqual(review["reason"], "placement_review")
            candidate = review["groups"][0]
            self.assertEqual(candidate["state"], "suspected_process")
            self.assertEqual(candidate["recommended_action"], "exclude")

            params["placement_resolutions"] = {
                candidate["key"]: {
                    "action": "exclude",
                    "part_number": "",
                    "field_patch": {},
                }
            }
            result = run_bom_process(root, params)

            self.assertEqual(result["status"], "ok")
            nc_path = next(Path(path) for path in result["outputs"] if path.endswith("_NC未贴汇总.xlsx"))
            nc = load_workbook(nc_path, data_only=True)
            try:
                rows = list(nc.active.iter_rows(min_row=2, values_only=True))
            finally:
                nc.close()
            self.assertEqual(rows[0][1], "LINK1")
            self.assertEqual(rows[0][7], "用户确认不装（疑似工艺件）")
            self.assertEqual(rows[0][8], "process_default")

    def test_parsed_source_can_filter_shields_without_reopening_workbook(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            source = Path(tmp) / "source.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.append(["Reference", "Part Number", "Value", "物料名称"])
            ws.append(["R1", "P1", "10K", "电阻"])
            ws.append(["SH1", "SH-PN", "SHIELD", "屏蔽支架"])
            wb.save(source)

            parsed = bom_process.parse_source(source)
            with mock.patch.object(
                bom_process,
                "open_bom_workbook",
                wraps=bom_process.open_bom_workbook,
            ) as opener:
                without_shields, excluded = bom_process.filter_rows(parsed, include_shields=False)
                with_shields, _ = bom_process.filter_rows(parsed, include_shields=True)

            self.assertEqual(opener.call_count, 0)
            self.assertEqual([row["part_number"] for row in without_shields], ["P1"])
            self.assertEqual([row["part_number"] for row in with_shields], ["P1", "SH-PN"])
            self.assertEqual(excluded[0][1], "SH1")

    def test_sh_row_keeps_capture_original_fields_in_nc_summary(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            source = Path(tmp) / "source.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.append(["Reference", "Part Number", "Value", "Description", "物料名称"])
            ws.append(["SH1", "SH-PN", "SHIELD-5S", None, "屏蔽罩 5S-A"])
            wb.save(source)

            parsed = bom_process.parse_source(source)
            _, excluded = bom_process.filter_rows(parsed, include_shields=False)

        self.assertEqual(excluded[0][3], "屏蔽罩 5S-A")
        self.assertEqual(excluded[0][5], "")
        self.assertEqual(excluded[0][6], "SHIELD-5S")

    def test_complete_bom_process_opens_source_workbook_once(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = root / "source.xlsx"
            make_source(source)

            with mock.patch.object(
                bom_process,
                "open_bom_workbook",
                wraps=bom_process.open_bom_workbook,
            ) as opener:
                result = run_bom_process(
                    root,
                    {
                        "source_bom": str(source),
                        "formats": ["plm", "oa"],
                        "parent_code": "203010100819",
                        "name": "TEST",
                        "confirm_shields": False,
                        "merge_conflicts": True,
                        "conflict_choices": {"P1": 1},
                    },
                )

            self.assertEqual(result["status"], "ok")
            self.assertEqual(opener.call_count, 1)

    def test_jumper_resistor_description_is_not_excluded(self) -> None:
        row = {
            "part_number": "R.001",
            "value": "0R",
            "name": "电阻",
            "desc": "跳线电阻 0402",
        }

        self.assertIsNone(bom_process.exclusion_reason(row, ["R100"]))

    def test_test_socket_description_is_not_excluded(self) -> None:
        row = {
            "part_number": "J.001",
            "value": "SOCKET",
            "name": "连接器",
            "desc": "Test socket 治具配套",
        }

        self.assertIsNone(bom_process.exclusion_reason(row, ["J1"]))

    def test_jumper_name_on_non_r_ref_is_flagged_for_review(self) -> None:
        row = {
            "part_number": "J.002",
            "value": "",
            "name": "跳线",
            "desc": "",
        }

        reason = bom_process.exclusion_reason(row, ["J5"])

        self.assertIsNotNone(reason)
        self.assertIn("工艺件", reason)

    def test_prefix_alone_no_longer_excludes_material(self) -> None:
        row = {
            "part_number": "P001",
            "value": "",
            "name": "常规器件",
            "desc": "普通贴片物料",
            "model": "MODEL-A",
        }

        for ref in ("TP1", "Z_TP2", "JP1", "H1", "MH1", "MTG3", "R100", "C55", "U9"):
            with self.subTest(ref=ref):
                self.assertIsNone(bom_process.exclusion_reason(row, [ref]))
        self.assertEqual(bom_process.exclusion_reason(row, ["SH1"]), "屏蔽支架 SH*")

    def test_sh_still_triggers_confirmation_regardless_of_partnumber(self) -> None:
        row = {"part_number": "P001", "value": "", "desc": "0402 电容", "name": "电容"}

        self.assertEqual(bom_process.exclusion_reason(row, ["SH1"]), "屏蔽支架 SH*")
        self.assertIsNone(bom_process.exclusion_reason(row, ["SH1"], include_shields=True))

    def test_material_description_drives_process_candidate(self) -> None:
        row = {
            "part_number": "P001",
            "value": "",
            "name": "探针",
            "desc": "测试点 探针",
            "model": "PROBE-A",
        }
        key = bom_process._process_candidate_key("P001", ["R100"])

        self.assertEqual(bom_process.exclusion_reason(row, ["R100"]), "工艺件（描述含 测试点）")
        self.assertIsNone(
            bom_process.exclusion_reason(row, ["R100"], process_material_keeps={key})
        )

    def test_process_material_regex_matches_only_independent_words(self) -> None:
        row = {
            "part_number": "P001",
            "value": "10K",
            "name": "电阻",
            "desc": "10K 电阻 测试用",
        }

        self.assertIsNone(bom_process.exclusion_reason(row, ["R100"]))

    def test_value_nc_overrides_process_keyword(self) -> None:
        row = {"part_number": "P001", "value": "NC", "name": "探针", "desc": "测试点 探针"}

        self.assertEqual(bom_process.exclusion_reason(row, ["R100"]), "NC/未贴")

    def test_sh_position_overrides_process_keyword(self) -> None:
        row = {"part_number": "P001", "value": "", "name": "探针", "desc": "测试点"}

        self.assertEqual(bom_process.exclusion_reason(row, ["SH1"]), "屏蔽支架 SH*")

    def test_detect_process_material_candidates_excludes_sh_nc_and_no_partnumber(self) -> None:
        rows = [
            {"reference": "R100", "part_number": "P001", "value": "", "name": "探针", "desc": "测试点 探针"},
            {"reference": "SH1", "part_number": "P002", "value": "", "name": "探针", "desc": "测试点"},
            {"reference": "TP2", "part_number": "P003", "value": "NC", "name": "探针", "desc": "测试点"},
            {"reference": "TP3", "part_number": "", "value": "", "name": "探针", "desc": "测试点"},
            {"reference": "R101", "part_number": "P004", "value": "10K", "name": "电阻", "desc": "普通电阻"},
        ]

        candidates = bom_process.detect_process_material_candidates(rows)

        self.assertEqual(len(candidates), 1)
        self.assertEqual(candidates[0]["key"], "P001|R100")
        self.assertEqual(candidates[0]["matched_keyword"], "测试点")

    def test_adapter_returns_shield_and_process_materials_in_one_review(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = root / "source.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.append(["Reference", "Part Number", "Value", "Description", "Name"])
            ws.append(["SH1", "SH-PN", "", "屏蔽支架", "屏蔽支架"])
            ws.append(["TP5", "TP-PN", "", "测试点 探针", "探针"])
            wb.save(source)
            params = {"source_bom": str(source), "formats": ["plm"], "name": "FLOW"}

            review = run_bom_process(root, params)
            groups = {group["category"]: group for group in review["groups"]}
            params["placement_resolutions"] = {
                groups["shield"]["key"]: {
                    "action": "keep",
                    "part_number": "SH-PN",
                    "field_patch": {},
                },
                groups["suspected_process"]["key"]: {
                    "action": "exclude",
                    "part_number": "TP-PN",
                    "field_patch": {},
                },
            }
            completed = run_bom_process(root, params)

        self.assertEqual(review["reason"], "placement_review")
        self.assertEqual(set(groups), {"shield", "suspected_process"})
        self.assertEqual(completed["status"], "ok")

    def test_letter_notation_numeric_pairs_require_manual_choice(self) -> None:
        def variants(first: str, second: str) -> list[dict[str, object]]:
            return [
                {
                    "name": "Resistor",
                    "model": model,
                    "desc": "Chip resistor",
                    "grade": "Preferred",
                    "unit": "ea",
                    "count": 1,
                }
                for model in (first, second)
            ]

        pairs = (
            ("1u", "1u5"),
            ("4n7", "4n70"),
            ("5V", "5V0"),
            ("1n", "1n5"),
            ("3V3", "3V30"),
            ("4R7", "4R75"),
            ("2m2", "2m25"),
            ("100", "1000"),
            ("1K", "1K2"),
        )
        for first, second in pairs:
            with self.subTest(first=first, second=second):
                recommendation = bom_process._conflict_recommendation(variants(first, second))

                self.assertEqual(recommendation["reason"], "conflicting_candidate_values")
                self.assertFalse(recommendation["high_confidence"])
                self.assertTrue(recommendation["manual_choice_required"])

    def test_digit_extension_prefix_is_never_truncation(self) -> None:
        for first, second in (("ABC-1", "ABC-12"), ("X 5", "X 52")):
            with self.subTest(first=first, second=second):
                variants = [
                    {
                        "name": "Component",
                        "model": model,
                        "desc": "Description",
                        "grade": "Preferred",
                        "unit": "ea",
                        "count": 1,
                    }
                    for model in (first, second)
                ]

                recommendation = bom_process._conflict_recommendation(variants)

                self.assertEqual(recommendation["reason"], "conflicting_candidate_values")
                self.assertFalse(recommendation["high_confidence"])
                self.assertTrue(recommendation["manual_choice_required"])

    def test_looks_numeric_classification_table(self) -> None:
        numeric_values = (
            "100",
            "1000",
            "1K",
            "1K2",
            "4R7",
            "1u",
            "1u5",
            "4n7",
            "3V3",
            "5V",
            "5V0",
            "4.7uF",
            "10nF",
            "100 Ohm",
            "3.3",
            "0402",
        )
        text_values = ("10K resistor 0402 5%", "GRM155R71C104KA88D", "", "Chip resistor")

        for value in numeric_values:
            with self.subTest(value=value, expected=True):
                self.assertTrue(bom_process._looks_numeric(value))
        for value in text_values:
            with self.subTest(value=value, expected=False):
                self.assertFalse(bom_process._looks_numeric(value))

    def test_text_prefix_completion_remains_high_confidence(self) -> None:
        variants = [
            {
                "name": "Resistor",
                "model": "R0402",
                "desc": description,
                "grade": "Preferred",
                "unit": "ea",
                "count": 1,
            }
            for description in ("10K resistor 0402 5%", "10K resistor 0402 5% RoHS compliant")
        ]

        recommendation = bom_process._conflict_recommendation(variants)

        self.assertEqual(recommendation["reason"], "truncation_prefix_completion")
        self.assertTrue(recommendation["high_confidence"])
        self.assertFalse(recommendation["manual_choice_required"])

    def test_load_source_prefers_new_description_over_content_column(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            source = Path(tmp) / "source.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.append(["Item", "Quantity", "Reference", "Part Number", "内容", "器件描述（新整理）"])
            ws.append([1, 1, "R1", "P1", "", "new desc"])
            wb.save(source)

            rows, _ = bom_process.load_source(source)

            self.assertEqual(rows[0]["desc"], "new desc")

    def test_run_bom_process_requires_confirmation_for_same_code_conflicts(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = root / "source.xlsx"
            make_source(source)

            result = run_bom_process(
                root,
                {
                    "source_bom": str(source),
                    "formats": ["plm"],
                    "parent_code": "203010100819",
                    "name": "TEST",
                },
            )

            self.assertEqual(result["status"], "needs_confirmation")
            self.assertEqual(result["reason"], "part_property_conflicts")
            self.assertEqual(result["conflict_count"], 1)
            self.assertEqual(result["summary"]["conflicts"], 1)
            self.assertEqual(result["conflicts"][0]["code"], "P1")
            self.assertEqual(result["conflicts"][0]["total_refs"], 2)

    def test_recommended_merge_leaves_low_confidence_conflicts_split(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            source = tmp_path / "source.xlsx"
            make_source(source)
            parsed = bom_process.parse_source(source)

            result = bom_process.process(
                parsed,
                ["plm"],
                "203010100819",
                "",
                "TEST",
                [],
                tmp_path,
                "STAMP",
                None,
                merge_conflicts=True,
            )

            records = result["records"]
            p1_records = [record for record in records if record["code"] == "P1"]
            self.assertEqual(len(p1_records), 2)
            self.assertEqual([record["refs"] for record in p1_records], [["R1"], ["R2"]])
            self.assertEqual(
                {(record["model"], record["desc"], record["grade"]) for record in p1_records},
                {("M1", "D1", "正常"), ("M2", "D2", "优选")},
            )
            self.assertEqual(result["summary"]["unresolved_conflicts"], 1)

            wb = load_workbook(result["outputs"][0], read_only=True, data_only=True)
            ws = wb.active
            rows = [[ws.cell(r, c).value for c in range(1, 12)] for r in range(3, ws.max_row + 1)]
            wb.close()
            p1_rows = [row for row in rows if row[2] == "P1"]
            self.assertEqual(len(p1_rows), 2)

    def test_golden_conflict_recommendations_select_only_existing_signatures(self) -> None:
        expectations = json.loads((FIXTURES / "expected_recommendations.json").read_text(encoding="utf-8"))
        with tempfile.TemporaryDirectory() as tmp:
            source = build_capture_bom(Path(tmp) / "conflicts.xlsx", "known_conflicts")
            rows, _ = bom_process.load_source(source)
            conflicts = {item["code"]: item for item in bom_process.detect_part_conflicts(rows)}

        self.assertEqual(set(conflicts), set(expectations["conflict_codes"]))
        for code, expected in expectations["recommendations"].items():
            conflict = conflicts[code]
            actual_signatures = {
                json.dumps(
                    {
                        "model": variant["model"],
                        "description": variant["desc"],
                        "name": variant["name"],
                        "grade": variant["grade"],
                        "unit": variant["unit"],
                    },
                    ensure_ascii=False,
                    sort_keys=True,
                )
                for variant in conflict["variants"]
            }
            expected_signatures = {
                json.dumps(signature, ensure_ascii=False, sort_keys=True)
                for signature in expected["allowed_signatures"]
            }
            self.assertEqual(actual_signatures, expected_signatures, code)
            self.assertEqual(conflict["confidence"], expected["expected_confidence"], code)
            self.assertEqual(conflict["reason"], expected["expected_reason"], code)
            self.assertEqual(conflict["high_confidence"], expected["high_confidence"], code)
            self.assertEqual(conflict["manual_choice_required"], expected["manual_choice_required"], code)
            selected = conflict["recommended_signature"]
            selected_payload = {
                "model": selected["model"],
                "description": selected["desc"],
                "name": selected["name"],
                "grade": selected["grade"],
                "unit": selected["unit"],
            }
            self.assertIn(conflict["recommended_index"], range(len(conflict["variants"])))
            self.assertIn(
                json.dumps(selected_payload, ensure_ascii=False, sort_keys=True),
                actual_signatures,
                code,
            )
            if expected["high_confidence"]:
                self.assertEqual(
                    selected_payload,
                    expected["selected_signature"],
                    code,
                )

    def test_bulk_recommendation_merges_only_high_confidence_golden_groups(self) -> None:
        expectations = json.loads((FIXTURES / "expected_recommendations.json").read_text(encoding="utf-8"))
        with tempfile.TemporaryDirectory() as tmp:
            source = build_capture_bom(Path(tmp) / "conflicts.xlsx", "known_conflicts")
            parsed = bom_process.parse_source(source)
            records = bom_process.build_records(parsed, merge_conflicts=True)

        by_code: dict[str, list[dict[str, object]]] = {}
        for record in records:
            by_code.setdefault(str(record["code"]), []).append(record)
        for code, expected in expectations["recommendations"].items():
            allowed = expected["allowed_signatures"]
            actual = by_code[code]
            if expected["high_confidence"]:
                self.assertEqual(len(actual), 1, code)
                self.assertEqual(record_signature(actual[0]), expected["selected_signature"], code)
            else:
                self.assertEqual(len(actual), len(allowed), code)
                self.assertEqual(
                    {json.dumps(record_signature(record), ensure_ascii=False, sort_keys=True) for record in actual},
                    {json.dumps(signature, ensure_ascii=False, sort_keys=True) for signature in allowed},
                    code,
                )

    def test_complementary_blanks_never_synthesize_a_new_signature(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            source = Path(tmp) / "complementary.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.append(["Reference", "Part Number", "规格型号", "器件描述（新整理）", "物料名称", "等级", "单位"])
            ws.append(["R1", "P1", "M1", "", "RESISTOR", "优选", "ea"])
            ws.append(["R2", "P1", "", "D1", "RESISTOR", "优选", "ea"])
            wb.save(source)

            parsed = bom_process.parse_source(source)
            rows, _ = bom_process.filter_rows(parsed)
            conflict = bom_process.detect_part_conflicts(rows)[0]
            records = bom_process.build_records(parsed, merge_conflicts=True)

        self.assertEqual(conflict["confidence"], "low")
        self.assertEqual(conflict["reason"], "complementary_incomplete_candidates")
        self.assertEqual(len(records), 2)
        self.assertNotIn(
            {"model": "M1", "description": "D1", "name": "RESISTOR", "grade": "优选", "unit": "ea"},
            [record_signature(record) for record in records],
        )

    def test_blank_completion_is_high_confidence_and_uses_complete_candidate(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            source = Path(tmp) / "blank-completion.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.append(["Reference", "Part Number", "规格型号", "器件描述（新整理）", "物料名称", "等级", "单位"])
            ws.append(["R1", "P1", "M1", "D1", "", "优选", "ea"])
            ws.append(["R2", "P1", "M1", "D1", "RESISTOR", "优选", "ea"])
            wb.save(source)

            parsed = bom_process.parse_source(source)
            rows, _ = bom_process.filter_rows(parsed)
            conflict = bom_process.detect_part_conflicts(rows)[0]
            records = bom_process.build_records(parsed, merge_conflicts=True)

        self.assertEqual(conflict["confidence"], "high")
        self.assertEqual(conflict["reason"], "blank_completion")
        self.assertEqual(len(records), 1)
        self.assertEqual(
            record_signature(records[0]),
            {"model": "M1", "description": "D1", "name": "RESISTOR", "grade": "优选", "unit": "ea"},
        )

    def test_recommended_api_merge_reprompts_for_low_confidence_groups(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = root / "source.xlsx"
            make_source(source)

            result = run_bom_process(
                root,
                {
                    "source_bom": str(source),
                    "formats": ["plm"],
                    "parent_code": "203010100819",
                    "name": "TEST",
                    "merge_conflicts": True,
                    "conflict_choices": {},
                },
            )
            recommended_index = result["conflicts"][0]["recommended_index"]
            resolved = run_bom_process(
                root,
                {
                    "source_bom": str(source),
                    "formats": ["plm"],
                    "parent_code": "203010100819",
                    "name": "TEST",
                    "merge_conflicts": True,
                    "conflict_choices": {"P1": recommended_index},
                },
            )

        self.assertEqual(result["status"], "needs_confirmation")
        self.assertEqual(result["reason"], "part_property_conflicts")
        self.assertEqual(result["conflict_count"], 1)
        self.assertEqual(result["conflicts"][0]["confidence"], "low")
        self.assertEqual(resolved["status"], "ok")
        p1_rows = [row for row in resolved["preview"]["rows"] if row[0] == "P1"]
        self.assertEqual(len(p1_rows), 1)
        self.assertEqual(p1_rows[0][1:3], ["M2", "D2"])

    def test_low_confidence_conflict_still_exposes_a_deterministic_recommendation(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            source = Path(tmp) / "source.xlsx"
            make_source(source)

            rows, _ = bom_process.load_source(source)
            conflict = bom_process.detect_part_conflicts(rows)[0]

        self.assertEqual(conflict["confidence"], "low")
        self.assertTrue(conflict["manual_choice_required"])
        self.assertEqual(conflict["recommended_index"], 1)
        self.assertEqual(
            conflict["recommended_signature"],
            {
                "name": "电阻",
                "model": "M2",
                "desc": "D2",
                "grade": "优选",
                "unit": "",
            },
        )

    def test_recommended_api_merge_finishes_a_grade_only_conflict(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = root / "grade-only.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.append(["Reference", "Part Number", "规格型号", "器件描述（新整理）", "物料名称", "等级", "单位"])
            ws.append(["R1", "P1", "M1", "D1", "RESISTOR", "正常", "ea"])
            ws.append(["R2", "P1", "M1", "D1", "RESISTOR", "优选", "ea"])
            wb.save(source)

            result = run_bom_process(
                root,
                {
                    "source_bom": str(source),
                    "formats": ["plm"],
                    "parent_code": "203010100819",
                    "name": "TEST",
                    "merge_conflicts": True,
                    "conflict_choices": {},
                },
            )

        self.assertEqual(result["status"], "ok")
        self.assertEqual(result["summary"]["records"], 1)
        self.assertEqual(result["preview"]["rows"][0][5], "优选")

    def test_manual_choice_preserves_the_selected_unit(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            source = Path(tmp) / "unit-conflict.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.append(["Reference", "Part Number", "规格型号", "器件描述（新整理）", "物料名称", "等级", "单位"])
            ws.append(["R1", "P1", "M1", "D1", "RESISTOR", "优选", "ea"])
            ws.append(["R2", "P1", "M1", "D1", "RESISTOR", "优选", "pcs"])
            wb.save(source)
            parsed = bom_process.parse_source(source)

            records = bom_process.build_records(
                parsed,
                merge_conflicts=True,
                conflict_choices={"P1": 1},
            )

        self.assertEqual(len(records), 1)
        self.assertEqual(records[0]["unit"], "pcs")

    def test_process_uses_user_selected_conflict_variant_when_merging(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            source = tmp_path / "source.xlsx"
            make_source(source)
            parsed = bom_process.parse_source(source)

            result = bom_process.process(
                parsed,
                ["plm"],
                "203010100819",
                "",
                "TEST",
                [],
                tmp_path,
                "STAMP",
                None,
                merge_conflicts=True,
                conflict_choices={"P1": 1},
            )

            p1 = next(record for record in result["records"] if record["code"] == "P1")
            self.assertEqual(p1["name"], "电阻")
            self.assertEqual(p1["model"], "M2")
            self.assertEqual(p1["desc"], "D2")
            self.assertEqual(p1["grade"], "优选")

    def test_process_keeps_conflicts_split_when_not_confirmed_to_merge(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            source = tmp_path / "source.xlsx"
            make_source(source)
            parsed = bom_process.parse_source(source)

            result = bom_process.process(
                parsed,
                ["plm"],
                "203010100819",
                "",
                "TEST",
                [],
                tmp_path,
                "STAMP",
                None,
                merge_conflicts=False,
            )

            p1_records = [record for record in result["records"] if record["code"] == "P1"]
            self.assertEqual(len(p1_records), 2)
            self.assertEqual([record["qty"] for record in p1_records], [1, 1])

    def test_run_bom_process_requires_confirmation_for_sh_shield_brackets(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = root / "source.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.append(["Item", "Quantity", "Reference", "Part Number", "Value", "规格型号", "器件描述（新整理）", "物料名称", "等级"])
            ws.append([1, 1, "SH1", "SH-PN", "SHIELD", "shield bracket", "屏蔽支架", "屏蔽支架", "正常"])
            wb.save(source)

            result = run_bom_process(
                root,
                {
                    "source_bom": str(source),
                    "formats": ["plm"],
                    "parent_code": "203010100819",
                    "name": "TEST",
                },
            )

            self.assertEqual(result["status"], "needs_confirmation")
            self.assertEqual(result["reason"], "placement_review")
            self.assertEqual(len(result["groups"]), 1)
            self.assertEqual(result["groups"][0]["category"], "shield")
            self.assertEqual(result["groups"][0]["refs"], ["SH1"])
            self.assertEqual(result["groups"][0]["original_fields"]["part_number"], "SH-PN")

    def test_shield_confirmation_precedes_conflict_prompt_for_same_code(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = root / "source.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.append(["Reference", "Part Number", "Value", "Model", "Description", "Name"])
            ws.append(["SH1", "P-SHARED", "SHIELD", "BRACKET-A", "Shield bracket", "屏蔽支架"])
            ws.append(["R1", "P-SHARED", "10K", "R0402", "Chip resistor", "电阻"])
            wb.save(source)
            params: dict[str, object] = {
                "source_bom": str(source),
                "formats": ["plm"],
                "parent_code": "203010100819",
                "name": "TEST",
            }

            shield_review = run_bom_process(root, params)
            shield_group = shield_review["groups"][0]
            params["placement_resolutions"] = {
                shield_group["key"]: {
                    "action": "keep",
                    "part_number": "P-SHARED",
                    "field_patch": {},
                }
            }
            conflict_review = run_bom_process(root, params)

        self.assertEqual(shield_review["status"], "needs_confirmation")
        self.assertEqual(shield_review["reason"], "placement_review")
        self.assertNotIn("conflicts", shield_review)
        self.assertEqual(conflict_review["status"], "needs_confirmation")
        self.assertEqual(conflict_review["reason"], "part_property_conflicts")
        self.assertEqual(conflict_review["conflicts"][0]["code"], "P-SHARED")

    def test_confirmed_sh_shield_brackets_enter_final_bom_not_nc_summary(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            source = tmp_path / "source.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.append(["Item", "Quantity", "Reference", "Part Number", "Value", "规格型号", "器件描述（新整理）", "物料名称", "等级"])
            ws.append([1, 1, "SH1", "SH-PN", "SHIELD", "shield bracket", "屏蔽支架", "屏蔽支架", "正常"])
            wb.save(source)
            parsed = bom_process.parse_source(source)

            result = bom_process.process(
                parsed,
                ["plm"],
                "203010100819",
                "",
                "TEST",
                [],
                tmp_path,
                "STAMP",
                None,
                confirm_shields=True,
            )

            shield = next(record for record in result["records"] if record["code"] == "SH-PN")
            self.assertEqual(shield["refs"], ["SH1"])
            self.assertEqual(shield["qty"], 1)
            self.assertEqual(shield["name"], "屏蔽支架")
            self.assertEqual(result["summary"]["shield_candidates"], 1)
            self.assertEqual(result["summary"]["excluded"], 0)

    def test_run_bom_process_requires_confirmation_for_sh_even_when_value_is_nc(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = root / "source.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.append(["Item", "Quantity", "Reference", "Part Number", "Value", "Model", "Description", "Name"])
            ws.append([1, 1, "SH1", "SH-PN", "NC", "shield bracket", "shield bracket", "shield bracket"])
            wb.save(source)

            result = run_bom_process(
                root,
                {
                    "source_bom": str(source),
                    "formats": ["plm"],
                    "parent_code": "203010100819",
                    "name": "TEST",
                },
            )

            self.assertEqual(result["status"], "needs_confirmation")
            self.assertEqual(result["reason"], "placement_review")
            self.assertEqual(result["groups"][0]["state"], "conflicting")
            self.assertEqual(result["groups"][0]["refs"], ["SH1"])

    def test_confirmed_sh_with_nc_value_enters_final_bom_not_nc_summary(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            source = tmp_path / "source.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.append(["Item", "Quantity", "Reference", "Part Number", "Value", "Model", "Description", "Name"])
            ws.append([1, 1, "SH1", "SH-PN", "NC", "shield bracket", "shield bracket", "shield bracket"])
            wb.save(source)
            parsed = bom_process.parse_source(source)

            result = bom_process.process(
                parsed,
                ["plm"],
                "203010100819",
                "",
                "TEST",
                [],
                tmp_path,
                "STAMP",
                None,
                confirm_shields=True,
            )

            shield = next(record for record in result["records"] if record["code"] == "SH-PN")
            self.assertEqual(shield["refs"], ["SH1"])
            self.assertEqual(shield["qty"], 1)
            self.assertEqual(result["summary"]["excluded"], 0)

            nc_wb = load_workbook(result["nc_summary"], data_only=True)
            nc_rows = list(nc_wb.active.iter_rows(values_only=True))
            self.assertEqual(len(nc_rows), 1)

    def test_nc_prefix_is_case_insensitive_but_not_nc_dash(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            source = tmp_path / "source.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.append(["Item", "Quantity", "Reference", "Part Number", "Value", "Model", "Description", "Name"])
            ws.append([1, 1, "R1", "P-NC-LOWER", "nc/not-mounted", "M1", "lower nc", "Resistor"])
            ws.append([2, 1, "R2", "P-NC-MIXED", "Nc/not-mounted", "M2", "mixed nc", "Resistor"])
            ws.append([3, 1, "R3", "P-NC-DASH", "NC-keep", "M3", "dash should stay", "Resistor"])
            ws.append([4, 1, "R4", "P-NC-EXACT", "NC", "M4", "exact nc", "Resistor"])
            wb.save(source)
            parsed = bom_process.parse_source(source)

            result = bom_process.process(
                parsed,
                ["plm"],
                "203010100819",
                "",
                "TEST",
                [],
                tmp_path,
                "STAMP",
                None,
            )

            codes = {record["code"] for record in result["records"]}
            self.assertNotIn("P-NC-LOWER", codes)
            self.assertNotIn("P-NC-MIXED", codes)
            self.assertNotIn("P-NC-EXACT", codes)
            self.assertIn("P-NC-DASH", codes)

            nc_wb = load_workbook(result["nc_summary"], data_only=True)
            try:
                nc_rows = list(nc_wb.active.iter_rows(values_only=True))
            finally:
                nc_wb.close()
            excluded_codes = {row[2] for row in nc_rows[1:]}
            self.assertIn("P-NC-LOWER", excluded_codes)
            self.assertIn("P-NC-MIXED", excluded_codes)
            self.assertIn("P-NC-EXACT", excluded_codes)
            self.assertNotIn("P-NC-DASH", excluded_codes)


if __name__ == "__main__":
    unittest.main()
