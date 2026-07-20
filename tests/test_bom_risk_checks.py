from __future__ import annotations

import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

from openpyxl import Workbook

from app.backend.tools.analysis_tools import run_bom_risk_check
from app.backend.tools import bom_process, common
from app.backend.tools.bom_rules import evaluate_bom_risks
from app.backend.tools.bom_risk import run_generic_bom_import
from app.backend.tools.common import _read_bom_rows


def _write_bom(path: Path, rows: list[list[object]]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.append(["Reference", "Part Number", "Description", "Quantity", "Name", "Model", "Grade"])
    for row in rows:
        ws.append(row)
    wb.save(path)


def _quantity_finding(quantity: object, refs: list[str]) -> dict[str, str]:
    findings = evaluate_bom_risks([{"part_number": "P1", "quantity": quantity, "refs": refs}])
    return next(item for item in findings if item["name"] == "数量=位号数")


class BomRiskCheckTests(unittest.TestCase):
    def test_generic_import_keeps_ncp_parts(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = root / "ncp.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.append(["Reference", "Part Number", "Description", "Quantity", "Name", "Value"])
            ws.append(["U1", "U.001", "LDO regulator", 1, "LDO", "NCP1117"])
            wb.save(source)

            result = run_generic_bom_import(root, {"source_bom": str(source)})

        self.assertEqual(result["status"], "ok")
        self.assertEqual(result["summary"]["main_rows"], 1)
        self.assertEqual(result["summary"]["excluded_rows"], 0)

    def test_generic_import_reuses_exclusion_reason_for_every_row(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = root / "rows.xlsx"
            _write_bom(
                source,
                [
                    ["R1", "R.001", "Resistor", 1, "电阻", "10K", "优选"],
                    ["C1", "C.001", "Capacitor", 1, "电容", "1uF", "优选"],
                ],
            )

            with patch("app.backend.tools.bom_process.exclusion_reason", wraps=bom_process.exclusion_reason) as check:
                result = run_generic_bom_import(root, {"source_bom": str(source)})

        self.assertEqual(result["status"], "ok")
        self.assertEqual(check.call_count, 2)

    def test_generic_import_wraps_user_input_errors(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = root / "invalid.xlsx"
            wb = Workbook()
            wb.active.append(["unexpected header"])
            wb.save(source)

            result = run_generic_bom_import(root, {"source_bom": str(source)})

        self.assertEqual(result["status"], "error")
        self.assertIn("表头", result["message"])

    def test_fractional_quantity_mismatch_is_reported(self) -> None:
        finding = _quantity_finding(3.9, ["R1", "R2", "R3"])

        self.assertEqual(finding["status"], "warn")

    def test_quantity_helpers_are_public_and_bom_rules_avoids_private_imports(self) -> None:
        self.assertTrue(hasattr(common, "to_qty"))
        self.assertTrue(hasattr(common, "qty_matches"))
        self.assertEqual(common.to_qty("3.9"), 3)
        self.assertFalse(common.qty_matches("3.9", 3))
        self.assertTrue(common.qty_matches("", 3))

        source = (Path(__file__).parents[1] / "app" / "backend" / "tools" / "bom_rules.py").read_text(encoding="utf-8")
        self.assertNotIn("common import _to_qty", source)

    def test_merged_quantity_does_not_create_a_false_mismatch(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "merged-quantity.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.append(["Reference", "Part Number", "Description", "Quantity"])
            ws.append(["R1,R2,R3", "P1", "First group", 3])
            ws.append(["R4,R5", "P2", "Second group", None])
            ws.merge_cells("D2:D3")
            wb.save(path)

            rows = _read_bom_rows(path)
            finding = next(item for item in evaluate_bom_risks(rows) if item["name"] == "数量=位号数")

        self.assertEqual(finding["status"], "ok")

    def test_float_quantity_equal_to_reference_count_is_ok(self) -> None:
        finding = _quantity_finding(3.0, ["R1", "R2", "R3"])

        self.assertEqual(finding["status"], "ok")

    def test_float_quantity_mismatch_is_still_reported(self) -> None:
        finding = _quantity_finding(3.0, ["R1", "R2"])

        self.assertEqual(finding["status"], "warn")

    def test_blank_quantity_remains_non_blocking(self) -> None:
        for quantity in (None, ""):
            with self.subTest(quantity=quantity):
                finding = _quantity_finding(quantity, ["R1", "R2", "R3"])
                self.assertEqual(finding["status"], "ok")

    def test_integer_quantity_equal_to_reference_count_is_ok(self) -> None:
        finding = _quantity_finding(3, ["R1", "R2", "R3"])

        self.assertEqual(finding["status"], "ok")

    def test_confirmed_sh_bracket_satisfies_shield_bracket_check(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            bom = root / "processed.xlsx"
            _write_bom(
                bom,
                [
                    ["SH1", "SH-PN", "屏蔽支架", 1, "屏蔽支架", "shield bracket", "正常"],
                    ["R1", "R-PN", "电阻", 1, "电阻", "10K", "正常"],
                ],
            )

            result = run_bom_risk_check(root, {"bom": str(bom)})

            finding = next(item for item in result["risk_report"]["findings"] if item["name"] == "屏蔽支架")
            self.assertEqual(finding["status"], "ok")
            self.assertIn("SH-PN", finding["message"])

    def test_emmc_and_ddr_warn_about_hardware_version(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            bom = root / "processed.xlsx"
            _write_bom(
                bom,
                [
                    ["U1", "EMMC-PN", "eMMC 128GB", 1, "eMMC", "KLMAG1JETD", "正常"],
                    ["U2", "DDR-PN", "LPDDR4X 8Gb", 1, "DDR", "MT53", "正常"],
                ],
            )

            result = run_bom_risk_check(root, {"bom": str(bom)})

            finding = next(item for item in result["risk_report"]["findings"] if item["name"] == "硬件版本敏感物料")
            self.assertEqual(finding["status"], "info")
            self.assertIn("EMMC-PN", finding["message"])
            self.assertIn("DDR-PN", finding["message"])
            self.assertIn("硬件版本号", finding["message"])


if __name__ == "__main__":
    unittest.main()
