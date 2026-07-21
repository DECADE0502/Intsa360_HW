from __future__ import annotations

import shutil
from pathlib import Path

from fastapi.testclient import TestClient
from openpyxl import load_workbook

from app.backend.main import create_app


ROOT = Path(__file__).resolve().parents[1]
FIXTURE = ROOT / "tests" / "fixtures" / "smt" / "synthetic"
BASE_URL = "http://127.0.0.1:8765"
SESSION_HEADER = "X-Insta360-Session"


def _runtime_root(tmp_path: Path) -> Path:
    root = tmp_path / "runtime"
    shutil.copytree(ROOT / "config", root / "config")
    (root / "VERSION").write_text("0.5.0\n", encoding="utf-8")
    (root / "REVISION").write_text("smt-layout-e2e\n", encoding="utf-8")
    return root


def _mutation_headers(client: TestClient) -> dict[str, str]:
    session = client.get("/api/session")
    assert session.status_code == 200
    return {SESSION_HEADER: session.json()["token"], "Origin": BASE_URL}


def _full_params() -> dict[str, str]:
    return {
        "smt_folder": str(FIXTURE),
        "processed_bom": str(FIXTURE / "bom_processed" / "PLM.xlsx"),
        "netlist_folder": str(FIXTURE / "netlist"),
        "name": "SYNTHETIC_BOARD",
    }


def test_e2e_smt_layout_full_flow(tmp_path: Path) -> None:
    root = _runtime_root(tmp_path)
    with TestClient(create_app(root), base_url=BASE_URL) as client:
        response = client.post("/api/tools/smt_layout/run", json=_full_params(), headers=_mutation_headers(client))

    assert response.status_code == 200, response.text
    payload = response.json()
    assert payload["status"] == "ok"
    assert len(payload["components"]) == 20
    assert payload["nc_summary"]["total"] == 2
    assert "R99" in {item["ref"] for item in payload["sanity"]["missing_layout"]}
    assert payload["fai_table"]["headers"][0:2] == ["位号", "面"]
    assert len(payload["fai_table"]["rows"]) == 20

    assert len(payload["outputs"]) == 1
    output = Path(payload["outputs"][0])
    assert output.exists()
    workbook = load_workbook(output, read_only=True, data_only=True)
    try:
        assert workbook.active.max_row == 21
        assert workbook.active.cell(1, 1).value == "位号"
    finally:
        workbook.close()


def test_e2e_smt_layout_missing_netlist_marks_sanity_skipped(tmp_path: Path) -> None:
    root = _runtime_root(tmp_path)
    params = _full_params()
    params.pop("netlist_folder")
    with TestClient(create_app(root), base_url=BASE_URL) as client:
        response = client.post("/api/tools/smt_layout/run", json=params, headers=_mutation_headers(client))

    assert response.status_code == 200, response.text
    assert response.json()["sanity"] == {"status": "skipped_no_netlist"}


def test_e2e_smt_layout_captures_input_error_as_chinese_message(tmp_path: Path) -> None:
    root = _runtime_root(tmp_path)
    with TestClient(create_app(root), base_url=BASE_URL) as client:
        response = client.post("/api/tools/smt_layout/run", json={}, headers=_mutation_headers(client))

    assert response.status_code == 400, response.text
    payload = response.json()
    assert payload["status"] == "error"
    assert "缺少必填输入" in payload["message"]
    assert "SMT 资料文件夹" in payload["message"]
