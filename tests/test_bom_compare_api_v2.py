from __future__ import annotations

import runpy
import shutil
from pathlib import Path

from fastapi.testclient import TestClient
from openpyxl import load_workbook

from app.backend.main import create_app
from app.backend.tools.bom_compare import run_bom_compare


ROOT = Path(__file__).resolve().parents[1]
BUILDER = ROOT / "tests" / "fixtures" / "bom_semantics" / "build_fixtures.py"
BASE_URL = "http://127.0.0.1:8765"


def _fixtures(tmp_path: Path) -> dict[str, Path]:
    return runpy.run_path(str(BUILDER))["build_all"](tmp_path)


def _runtime(tmp_path: Path) -> Path:
    root = tmp_path / "runtime"
    (root / "config").mkdir(parents=True)
    (root / "data" / "uploads").mkdir(parents=True)
    (root / "VERSION").write_text("0.5.5\n", encoding="utf-8")
    shutil.copy2(ROOT / "config" / "capabilities.json", root / "config" / "capabilities.json")
    return root


def _headers(client: TestClient) -> dict[str, str]:
    token = client.get("/api/v1/session").json()["token"]
    return {
        "X-Insta360-Session": token,
        "Origin": BASE_URL,
    }


def test_inspect_returns_schema_v2_quality_and_substitute_summary(tmp_path: Path) -> None:
    path = _fixtures(tmp_path)["substitutes"]

    result = run_bom_compare(tmp_path, {"action": "inspect", "source": str(path)})

    assert result["status"] == "ok"
    assert result["schema_version"] == 2
    assert result["action"] == "inspect"
    assert result["summary"]["parent_codes"] == ["BOARD-A"]
    assert result["summary"]["actual_reference_count"] == 4
    assert result["summary"]["substitute_group_count"] == 1
    assert result["inspection"]["can_compare"] is True


def test_legacy_two_file_request_keeps_workbench_and_adds_semantic_layers(tmp_path: Path) -> None:
    paths = _fixtures(tmp_path)

    result = run_bom_compare(
        tmp_path,
        {
            "bom1": str(paths["ordinary"]),
            "bom2": str(paths["substitutes"]),
            "output_dir": str(tmp_path / "outputs"),
        },
    )

    assert result["status"] == "ok"
    assert result["schema_version"] == 2
    assert result["compare"]["key_label"] == "位号"
    assert result["semantic"]["summary"]["actual_reference_count_new"] == 4
    assert len(result["semantic"]["placement_diff"]) > 0
    assert result["source_inspections"]["new"]["boards"][0]["placement_count"] == 4
    assert "rows" not in result["source_inspections"]["new"]["boards"][0]
    assert len(result["outputs"]) == 3
    assert all(Path(path).is_file() for path in result["outputs"])


def test_api_compare_records_semantic_summary_without_refresh(tmp_path: Path) -> None:
    runtime = _runtime(tmp_path)
    paths = _fixtures(runtime / "data" / "uploads")
    app = create_app(runtime)

    with TestClient(app, base_url=BASE_URL) as client:
        response = client.post(
            "/api/v1/tools/bom_compare/run",
            json={
                "action": "compare",
                "bom1": str(paths["ordinary"]),
                "bom2": str(paths["substitutes"]),
                "output_dir": "bom_compare",
            },
            headers=_headers(client),
        )
        history = client.get("/api/v1/history").json()["runs"]

    assert response.status_code == 200
    assert response.json()["schema_version"] == 2
    assert history[0]["tool"] == "bom_compare"
    assert history[0]["summary"]["semantic"]["analysis_fingerprint"]
    assert history[0]["summary"]["semantic"]["substitute_group_count_new"] == 1


def test_different_single_board_parent_requires_scope_confirmation(tmp_path: Path) -> None:
    paths = _fixtures(tmp_path)
    renamed = tmp_path / "renamed-parent.xlsx"
    workbook = load_workbook(paths["ordinary"])
    worksheet = workbook["BOM导入模版"]
    for row in range(3, worksheet.max_row + 1):
        worksheet.cell(row, 1).value = "BOARD-B"
        worksheet.cell(row, 2).value = "Board B"
    workbook.save(renamed)
    workbook.close()

    pending = run_bom_compare(
        tmp_path,
        {
            "bom1": str(paths["ordinary"]),
            "bom2": str(renamed),
            "output_dir": str(tmp_path / "pending-outputs"),
        },
    )

    assert pending["status"] == "ok"
    assert pending["needs_scope_confirmation"] is True
    assert pending["can_export"] is False
    assert pending["outputs"] == []
    assert "semantic" not in pending
    pair = pending["comparison_scope"]["pairs"][0]
    assert pair["old_parent_code"] == "BOARD-A"
    assert pair["new_parent_code"] == "BOARD-B"
    assert pair["status"] == "suggested"
    assert pair["evidence"]["shared_reference_count"] == 2

    confirmed = run_bom_compare(
        tmp_path,
        {
            "bom1": str(paths["ordinary"]),
            "bom2": str(renamed),
            "scope_confirmation": True,
            "output_dir": str(tmp_path / "confirmed-outputs"),
        },
    )

    assert confirmed["status"] == "ok"
    assert confirmed["needs_scope_confirmation"] is False
    assert confirmed["comparison_scope"]["status"] == "confirmed"
    assert confirmed["semantic"]["placement_diff"] == []
    assert len(confirmed["semantic"]["board_metadata_diff"]) == 1
    assert len(confirmed["outputs"]) == 3
