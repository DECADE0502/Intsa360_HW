from __future__ import annotations

import io
import shutil
import zipfile
from pathlib import Path

from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook

from app.backend.main import create_app


ROOT = Path(__file__).resolve().parents[1]
BASE_URL = "http://127.0.0.1:8765"
SESSION_HEADER = "X-Insta360-Session"


def _runtime_root(tmp_path: Path) -> Path:
    root = tmp_path / "runtime"
    shutil.copytree(ROOT / "config", root / "config")
    (root / "VERSION").write_text("0.4.0\n", encoding="utf-8")
    (root / "REVISION").write_text("0123456789abcdef\n", encoding="utf-8")
    return root


def _source_bom() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(
        [
            "Item",
            "Quantity",
            "Reference",
            "Part Number",
            "Value",
            "Model",
            "Description",
            "Name",
            "Unit",
        ]
    )
    sheet.append([1, 1, "SH1", "SH-PN", "SHIELD", "BRACKET-A", "Shield bracket", "Shield bracket", "pcs"])
    sheet.append([2, 1, "R1", "P1", "10K", "M1", "Description A", "Resistor", "pcs"])
    sheet.append([3, 1, "R2", "P1", "10K", "M2", "Description B", "Resistor", "pcs"])
    buffer = io.BytesIO()
    workbook.save(buffer)
    workbook.close()
    return buffer.getvalue()


def _process_material_source_bom() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Reference", "Part Number", "Value", "Model", "Description", "Name", "Unit", "PCB Footprint"])
    sheet.append(["TP5", "TP-PN", "", "PROBE-A", "测试点 探针", "探针", "pcs", "TESTPOINT_TP0P4"])
    sheet.append(["R1", "R-PN", "10K", "R0402", "普通贴片电阻", "电阻", "pcs", "R0402"])
    buffer = io.BytesIO()
    workbook.save(buffer)
    workbook.close()
    return buffer.getvalue()


def _mutation_headers(client: TestClient) -> dict[str, str]:
    session = client.get("/api/v1/session")
    assert session.status_code == 200
    return {SESSION_HEADER: session.json()["token"], "Origin": BASE_URL}


def _placement_resolution(group: dict[str, object], action: str) -> dict[str, object]:
    inferred = group.get("inferred_fields") if isinstance(group.get("inferred_fields"), dict) else {}
    role = str(group.get("role") or "unknown")
    destination = "smt" if action == "keep" else "non_smt"
    subtype = ""
    exclusion_kind = ""
    if role == "shield":
        subtype = "bracket" if destination == "smt" else "cover"
        exclusion_kind = "" if destination == "smt" else "scope_excluded"
    elif destination == "non_smt":
        exclusion_kind = "process_only" if role in {"test_point", "short_symbol", "mounting_hole", "fiducial"} else "user_excluded"
    return {
        "destination": destination,
        "exclusion_kind": exclusion_kind,
        "role": role,
        "subtype": subtype,
        "part_number_override": str(inferred.get("part_number") or ""),
        "field_patch": {},
        "decision_source": "user",
    }


def test_bom_process_confirmation_package_and_history_flow(tmp_path: Path) -> None:
    root = _runtime_root(tmp_path)
    with TestClient(create_app(root), base_url=BASE_URL) as client:
        headers = _mutation_headers(client)
        upload = client.post(
            "/api/v1/upload",
            files={"files": ("source.xlsx", _source_bom(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            headers=headers,
        )
        assert upload.status_code == 200, upload.text
        source = upload.json()["files"][0]["path"]
        params: dict[str, object] = {
            "source_bom": source,
            "formats": ["plm"],
            "parent_code": "203010100819",
            "parent_desc": "E2E board",
            "name": "E2E_BOARD",
        }

        placement_review = client.post("/api/v1/tools/bom_process/run", json=params, headers=headers)
        assert placement_review.status_code == 200, placement_review.text
        assert placement_review.json()["status"] == "needs_confirmation"
        assert placement_review.json()["reason"] == "placement_review"
        shield_group = placement_review.json()["groups"][0]
        assert shield_group["category"] == "shield"
        assert shield_group["refs"] == ["SH1"]

        params["placement_resolutions"] = {
            shield_group["key"]: _placement_resolution(shield_group, "keep")
        }
        conflict_review = client.post("/api/v1/tools/bom_process/run", json=params, headers=headers)
        assert conflict_review.status_code == 200, conflict_review.text
        assert conflict_review.json()["status"] == "needs_confirmation"
        assert conflict_review.json()["reason"] == "part_property_conflicts"
        conflict = conflict_review.json()["conflicts"][0]
        assert conflict["code"] == "P1"

        params["merge_conflicts"] = True
        params["conflict_choices"] = {
            "P1": {"action": "select_variant", "variant_index": 0}
        }
        completed = client.post("/api/v1/tools/bom_process/run", json=params, headers=headers)
        assert completed.status_code == 200, completed.text
        result = completed.json()
        assert result["status"] == "ok"
        assert any(Path(item).name.endswith("_PLM_BOM.xlsx") for item in result["outputs"])
        assert Path(result["semantic_manifest"]).is_file()
        assert result["semantic_manifest"] in result["outputs"]
        assert any(Path(item).name.endswith("_NC未贴汇总.xlsx") for item in result["outputs"])

        package = client.post(
            "/api/package",
            json={"name": "E2E_BOARD", "files": result["outputs"]},
            headers=headers,
        )
        assert package.status_code == 200, package.text
        assert package.headers["content-type"] == "application/zip"
        with zipfile.ZipFile(io.BytesIO(package.content)) as archive:
            names = archive.namelist()
            assert any(name.endswith("_PLM_BOM.xlsx") for name in names)
            assert any(name.endswith("_BOM语义模型.json") for name in names)
            assert any(name.endswith("_NC未贴汇总.xlsx") for name in names)

        history = client.get("/api/history")
        assert history.status_code == 200, history.text
        assert any(
            run.get("tool") == "bom_process" and run.get("status") == "succeeded"
            for run in history.json()["runs"]
        )


def test_e2e_rejected_shield_stays_out_of_plm_and_keeps_raw_nc_name(tmp_path: Path) -> None:
    root = _runtime_root(tmp_path)
    with TestClient(create_app(root), base_url=BASE_URL) as client:
        headers = _mutation_headers(client)
        upload = client.post(
            "/api/v1/upload",
            files={"files": ("source.xlsx", _source_bom(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            headers=headers,
        )
        assert upload.status_code == 200, upload.text
        params: dict[str, object] = {
            "source_bom": upload.json()["files"][0]["path"],
            "formats": ["plm"],
            "parent_code": "203010100819",
            "parent_desc": "E2E board",
            "name": "E2E_REJECT_SH",
        }

        placement_review = client.post("/api/v1/tools/bom_process/run", json=params, headers=headers)
        assert placement_review.json()["reason"] == "placement_review"
        shield_group = placement_review.json()["groups"][0]

        params["placement_resolutions"] = {
            shield_group["key"]: _placement_resolution(shield_group, "exclude")
        }
        conflict_review = client.post("/api/v1/tools/bom_process/run", json=params, headers=headers)
        conflict = conflict_review.json()["conflicts"][0]
        params["merge_conflicts"] = True
        params["conflict_choices"] = {
            "P1": {"action": "select_variant", "variant_index": 0}
        }
        completed = client.post("/api/v1/tools/bom_process/run", json=params, headers=headers)
        assert completed.status_code == 200, completed.text
        result = completed.json()
        assert result["status"] == "ok"

        plm_path = next(Path(path) for path in result["outputs"] if path.endswith("_PLM_BOM.xlsx"))
        nc_path = Path(result["non_smt_summary"])
        plm = load_workbook(plm_path, data_only=True)
        nc = load_workbook(nc_path, data_only=True)
        try:
            plm_refs = [str(row[8] or "") for row in plm.active.iter_rows(min_row=3, values_only=True)]
            nc_rows = list(nc.active.iter_rows(min_row=2, values_only=True))
        finally:
            plm.close()
            nc.close()

        assert all("SH1" not in refs for refs in plm_refs)
        shield_nc = next(row for row in nc_rows if row[2] == "SH-PN")
        assert shield_nc[1] == "SH1"
        assert shield_nc[3] == "Shield bracket"


def test_e2e_coded_tp_with_correlated_package_is_auto_non_smt(tmp_path: Path) -> None:
    root = _runtime_root(tmp_path)
    with TestClient(create_app(root), base_url=BASE_URL) as client:
        headers = _mutation_headers(client)
        upload = client.post(
            "/api/v1/upload",
            files={"files": ("source.xlsx", _process_material_source_bom(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            headers=headers,
        )
        source = upload.json()["files"][0]["path"]
        params: dict[str, object] = {"source_bom": source, "formats": ["plm"], "name": "KEEP_TP"}

        result = client.post("/api/v1/tools/bom_process/run", json=params, headers=headers).json()
        assert result["status"] == "ok"

        plm_path = next(Path(path) for path in result["outputs"] if path.endswith("_PLM_BOM.xlsx"))
        plm = load_workbook(plm_path, data_only=True)
        try:
            refs = [str(row[8] or "") for row in plm.active.iter_rows(min_row=3, values_only=True)]
        finally:
            plm.close()
        assert all("TP5" not in item for item in refs)


def test_e2e_coded_tp_is_process_only_not_nc(tmp_path: Path) -> None:
    root = _runtime_root(tmp_path)
    with TestClient(create_app(root), base_url=BASE_URL) as client:
        headers = _mutation_headers(client)
        upload = client.post(
            "/api/v1/upload",
            files={"files": ("source.xlsx", _process_material_source_bom(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            headers=headers,
        )
        params: dict[str, object] = {
            "source_bom": upload.json()["files"][0]["path"],
            "formats": ["plm"],
            "name": "DEFAULT_NC",
        }

        result = client.post("/api/v1/tools/bom_process/run", json=params, headers=headers).json()
        non_smt = load_workbook(Path(result["non_smt_summary"]), data_only=True)
        nc = load_workbook(Path(result["nc_summary"]), data_only=True)
        try:
            rows = list(non_smt.active.iter_rows(min_row=2, values_only=True))
            nc_rows = list(nc.active.iter_rows(min_row=2, values_only=True))
        finally:
            non_smt.close()
            nc.close()
        tp = next(row for row in rows if row[1] == "TP5")
        assert tp[7] == "非贴片工艺项"
        assert tp[8] == "process_only"
        assert all(row[1] != "TP5" for row in nc_rows)
