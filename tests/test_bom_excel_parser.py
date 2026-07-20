from __future__ import annotations

import importlib.util
import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

from openpyxl import Workbook

from app.backend.parsers.bom_table import read_bom_rows
from app.backend.tools import bom_process
from app.backend.tools.common import _read_bom_rows


class BomExcelParserTests(unittest.TestCase):
    def test_merged_quantity_cell_is_not_inherited(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "merged-quantity.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.append(["Reference", "Part Number", "Description", "Quantity"])
            ws.append(["R1,R2,R3", "P1", "First group", 3])
            ws.append(["R4,R5", "P2", "Second group", None])
            ws.merge_cells("D2:D3")
            wb.save(path)

            rows = read_bom_rows(path)

        self.assertEqual([row["quantity"] for row in rows], [3, None])

    def test_merged_reference_cell_is_not_inherited(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "merged-reference.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.append(["Reference", "Part Number", "Description", "Quantity"])
            ws.append(["R1,R2", "P1", "First group", 2])
            ws.append([None, "P2", "Second group", None])
            ws.merge_cells("A2:A3")
            wb.save(path)

            rows = read_bom_rows(path, require_refs=False)

        self.assertEqual(len(rows), 2)
        self.assertEqual(rows[0]["refs"], ["R1", "R2"])
        self.assertEqual(rows[1]["refs"], [])

    def test_all_readers_share_one_header_normalization(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "normalized-headers.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.append(["{Reference}", "Part Number*", "描述##ms", "数量"])
            ws.append(["R1", "P1", "Chip resistor", 1])
            wb.save(path)

            parser_rows = read_bom_rows(path)
            process_rows, _ = bom_process.load_source(path)

        self.assertEqual(parser_rows[0]["part_number"], "P1")
        self.assertEqual(process_rows[0]["part_number"], "P1")
        self.assertEqual(parser_rows[0]["description"], process_rows[0]["desc"])

    def test_bom_excel_module_is_removed(self) -> None:
        self.assertIsNone(importlib.util.find_spec("app.backend.parsers.bom_excel"))

    def test_vertical_merged_part_number_is_inherited_by_all_bom_readers(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "merged-bom.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.append(["Reference", "Part Number", "Description", "Quantity"])
            ws.append(["R1", "P1", "Chip resistor", 1])
            ws.append(["R2", None, "Chip resistor", 1])
            ws.merge_cells("B2:B3")
            wb.save(path)

            process_rows, _ = bom_process.load_source(path)
            parser_rows = read_bom_rows(path)
            common_rows = _read_bom_rows(path)

        for label, rows in (
            ("bom_process", process_rows),
            ("bom_table", parser_rows),
            ("common", common_rows),
        ):
            with self.subTest(reader=label):
                self.assertEqual(len(rows), 2)
                self.assertEqual([row["part_number"] for row in rows], ["P1", "P1"])

    def test_read_bom_rows_prefers_child_columns_after_part_number(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "bom.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.append(["父项编码", "描述", "子项编码", "名称", "型号", "描述", "单位", "数量", "位号", "物料优选等级"])
            ws.append(["PARENT", "整机", "R.001", "电阻", "10K", "贴片电阻", "ea", 2, "R1,R2", "优选"])
            wb.save(path)

            rows = read_bom_rows(path)

            self.assertEqual(len(rows), 1)
            self.assertEqual(rows[0]["part_number"], "R.001")
            self.assertEqual(rows[0]["description"], "贴片电阻")
            self.assertEqual(rows[0]["quantity"], 2)
            self.assertEqual(rows[0]["refs"], ["R1", "R2"])
            self.assertEqual(rows[0]["name"], "电阻")
            self.assertEqual(rows[0]["model"], "10K")
            self.assertEqual(rows[0]["grade"], "优选")

    def test_read_bom_rows_can_keep_no_ref_board_rows(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "bom.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.append(["Reference", "Part Number", "Description", "Quantity"])
            ws.append(["", "PCB.001", "PCB board", 1])
            wb.save(path)

            self.assertEqual(read_bom_rows(path), [])
            self.assertEqual(read_bom_rows(path, require_refs=False)[0]["part_number"], "PCB.001")

    def test_read_bom_rows_closes_workbook_when_header_detection_fails(self) -> None:
        class TrackingWorkbook:
            def __init__(self, worksheet) -> None:
                self.active = worksheet
                self.closed = False

            def close(self) -> None:
                self.closed = True

        source = Workbook()
        source.active.append(["unexpected header"])
        workbook = TrackingWorkbook(source.active)

        with patch("app.backend.parsers._workbook.load_workbook", return_value=workbook):
            with self.assertRaises(ValueError):
                read_bom_rows(Path("missing.xlsx"))

        self.assertTrue(workbook.closed)


if __name__ == "__main__":
    unittest.main()
