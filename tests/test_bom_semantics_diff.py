from __future__ import annotations

import runpy
from pathlib import Path

from openpyxl import load_workbook

from app.backend.bom_semantics.diff import compare_board_boms
from app.backend.bom_semantics.models import (
    ChangeKind,
    FindingSeverity,
    FunctionalImpact,
    ValidationFinding,
)
from app.backend.bom_semantics.normalization import normalize_workbook
from app.backend.bom_semantics.substitutes import build_board_boms


BUILDER = Path(__file__).resolve().parent / "fixtures" / "bom_semantics" / "build_fixtures.py"


def _fixtures(tmp_path: Path) -> dict[str, Path]:
    return runpy.run_path(str(BUILDER))["build_all"](tmp_path)


def _boards(path: Path):
    return build_board_boms(normalize_workbook(path))


def test_three_member_group_does_not_inflate_physical_count(tmp_path: Path) -> None:
    paths = _fixtures(tmp_path)
    result = compare_board_boms(_boards(paths["substitutes"]), _boards(paths["substitutes"]))

    assert result.summary.actual_reference_count_old == 4
    assert result.summary.actual_reference_count_new == 4
    assert result.events == ()
    assert result.can_export


def test_adding_alternative_is_supply_change_not_placement_change(tmp_path: Path) -> None:
    paths = _fixtures(tmp_path)
    new_path = tmp_path / "two_members.xlsx"
    workbook = load_workbook(paths["substitutes"])
    worksheet = workbook["BOM导入模版"]
    worksheet.delete_rows(5)
    workbook.save(new_path)
    workbook.close()

    result = compare_board_boms(_boards(new_path), _boards(paths["substitutes"]))

    assert result.summary.actual_reference_count_old == result.summary.actual_reference_count_new == 4
    assert [event.kind for event in result.events] == [ChangeKind.ALTERNATIVE_ADDED]
    assert result.events[0].impact == FunctionalImpact.SUPPLY


def test_main_priority_swap_is_one_business_event(tmp_path: Path) -> None:
    paths = _fixtures(tmp_path)
    swapped = tmp_path / "swapped.xlsx"
    workbook = load_workbook(paths["substitutes"])
    worksheet = workbook["BOM导入模版"]
    main_refs = worksheet["I3"].value
    worksheet["I3"] = ""
    worksheet["M3"] = "MAT-B"
    worksheet["P3"] = 1
    worksheet["I4"] = main_refs
    worksheet["M4"] = "MAT-B"
    worksheet["P4"] = 0
    worksheet["M5"] = "MAT-B"
    workbook.save(swapped)
    workbook.close()

    result = compare_board_boms(_boards(paths["substitutes"]), _boards(swapped))
    functional = [event for event in result.events if event.impact != FunctionalImpact.METADATA]

    assert len(functional) == 1
    assert functional[0].kind == ChangeKind.MAIN_CHANGED_REFS_MIGRATED
    assert len(functional[0].references) == 4


def test_missing_relation_configuration_does_not_create_false_change(
    tmp_path: Path,
) -> None:
    paths = _fixtures(tmp_path)
    system_export = tmp_path / "system-export-with-empty-relation-fields.xlsx"
    workbook = load_workbook(paths["substitutes"])
    worksheet = workbook.active
    worksheet["N4"] = ""
    worksheet["O4"] = ""
    workbook.save(system_export)
    workbook.close()

    result = compare_board_boms(
        _boards(paths["substitutes"]),
        _boards(system_export),
    )

    assert result.substitute_diff == ()
    assert result.events == ()
    assert result.can_export
    assert {
        finding.code for finding in result.warnings
    } >= {"substitute_strategy_missing", "substitute_mode_missing"}


def test_two_known_relation_configurations_are_compared_semantically(
    tmp_path: Path,
) -> None:
    paths = _fixtures(tmp_path)
    changed = tmp_path / "changed-relation-configuration.xlsx"
    workbook = load_workbook(paths["substitutes"])
    worksheet = workbook.active
    worksheet["N4"] = "project-specific-policy"
    worksheet["O4"] = "project-specific-mode"
    workbook.save(changed)
    workbook.close()

    result = compare_board_boms(
        _boards(paths["substitutes"]),
        _boards(changed),
    )

    assert len(result.substitute_diff) == 1
    assert set(result.substitute_diff[0]["change_dimensions"]) == {
        "strategies",
        "modes",
    }
    assert [event.kind for event in result.events] == [
        ChangeKind.SUBSTITUTE_CONFIGURATION_CHANGED
    ]


def test_description_only_change_stays_in_metadata_layer(tmp_path: Path) -> None:
    paths = _fixtures(tmp_path)
    changed = tmp_path / "metadata.xlsx"
    workbook = load_workbook(paths["ordinary"])
    worksheet = workbook["BOM导入模版"]
    worksheet["F3"] = "电阻 10K 更新描述"
    workbook.save(changed)
    workbook.close()

    result = compare_board_boms(_boards(paths["ordinary"]), _boards(changed))

    assert result.placement_diff == ()
    assert len(result.metadata_diff) == 1
    assert result.events[0].kind == ChangeKind.METADATA_ONLY
    assert result.events[0].impact == FunctionalImpact.METADATA
    assert result.summary.changed_event_count == 1
    assert result.summary.review_event_count == 0
    assert result.summary.metadata_event_count == 1
    assert result.summary.metadata_change_count == 1
    assert result.summary.metadata_field_count == 1


def test_complete_material_replacement_requires_old_material_to_exit(tmp_path: Path) -> None:
    paths = _fixtures(tmp_path)
    changed = tmp_path / "replacement.xlsx"
    workbook = load_workbook(paths["ordinary"])
    worksheet = workbook["BOM导入模版"]
    worksheet["C3"] = "MAT-R-NEW"
    workbook.save(changed)
    workbook.close()

    result = compare_board_boms(_boards(paths["ordinary"]), _boards(changed))

    replacement = [event for event in result.events if event.kind == ChangeKind.REPLACEMENT]
    assert len(replacement) == 1
    assert replacement[0].oa_change_type == "更换(A换成B)"
    assert replacement[0].references == ("R1", "R2")
    assert len(result.placement_diff) == 2
    assert len(result.placement_groups) == 1
    assert result.placement_groups[0]["references"] == ["R1", "R2"]
    assert result.placement_groups[0]["reference_count"] == 2
    assert result.summary.placement_change_group_count == 1
    assert result.summary.placement_changed_reference_count == 2


def test_placement_groups_do_not_merge_different_business_outcomes(tmp_path: Path) -> None:
    paths = _fixtures(tmp_path)
    changed = tmp_path / "mixed-placement-changes.xlsx"
    workbook = load_workbook(paths["ordinary"])
    worksheet = workbook["BOM导入模版"]
    worksheet["C3"] = "MAT-R-NEW"
    worksheet["I4"] = "C1,C3"
    workbook.save(changed)
    workbook.close()

    result = compare_board_boms(_boards(paths["ordinary"]), _boards(changed))

    keys = {
        (
            group["status"],
            group["old_material_code"],
            group["new_material_code"],
        )
        for group in result.placement_groups
    }
    assert ("migrated", "MAT-R", "MAT-R-NEW") in keys
    assert any(group["status"] in {"added", "removed"} for group in result.placement_groups)


def test_complete_replacement_allows_target_material_to_preexist(tmp_path: Path) -> None:
    paths = _fixtures(tmp_path)
    old_path = tmp_path / "target-preexists-old.xlsx"
    new_path = tmp_path / "target-preexists-new.xlsx"
    workbook = load_workbook(paths["ordinary"])
    worksheet = workbook["BOM导入模版"]
    worksheet["H3"] = 1
    worksheet["I3"] = "R1"
    worksheet["C4"] = "MAT-TARGET"
    worksheet["D4"] = "电阻"
    worksheet["E4"] = "10K"
    worksheet["F4"] = "目标物料"
    worksheet["H4"] = 1
    worksheet["I4"] = "R2"
    workbook.save(old_path)
    worksheet["C3"] = "MAT-TARGET"
    worksheet["H3"] = 2
    worksheet["I3"] = "R1,R2"
    worksheet.delete_rows(4)
    workbook.save(new_path)
    workbook.close()

    result = compare_board_boms(_boards(old_path), _boards(new_path))

    replacements = [
        event for event in result.events
        if event.kind == ChangeKind.REPLACEMENT
    ]
    assert len(replacements) == 1
    assert replacements[0].references == ("R1",)
    assert replacements[0].old_snapshot["material_code"] == "MAT-R"
    assert replacements[0].new_snapshot["material_code"] == "MAT-TARGET"


def test_same_material_reference_delta_is_one_reference_set_event(tmp_path: Path) -> None:
    paths = _fixtures(tmp_path)
    changed = tmp_path / "reference-set.xlsx"
    workbook = load_workbook(paths["ordinary"])
    worksheet = workbook["BOM导入模版"]
    worksheet["I3"] = "R2,R3"
    workbook.save(changed)
    workbook.close()

    result = compare_board_boms(_boards(paths["ordinary"]), _boards(changed))

    events = [
        event for event in result.events
        if event.kind == ChangeKind.REFERENCE_SET_CHANGED
    ]
    assert len(events) == 1
    assert events[0].references == ("R1", "R3")
    assert not any(
        event.kind in {ChangeKind.REFERENCE_ADDED, ChangeKind.REFERENCE_REMOVED}
        for event in result.events
    )


def test_issue_method_only_change_is_visible_in_metadata(tmp_path: Path) -> None:
    paths = _fixtures(tmp_path)
    changed = tmp_path / "issue-method.xlsx"
    workbook = load_workbook(paths["ordinary"])
    worksheet = workbook["BOM导入模版"]
    worksheet["Q3"] = "直接领料"
    workbook.save(changed)
    workbook.close()

    result = compare_board_boms(_boards(paths["ordinary"]), _boards(changed))

    assert len(result.metadata_diff) == 1
    assert result.metadata_diff[0]["changed_fields"] == ["issue_method"]
    assert any(
        event.kind == ChangeKind.METADATA_ONLY
        for event in result.events
    )


def test_summary_distinguishes_blocker_findings_from_affected_records() -> None:
    findings = (
        ValidationFinding(
            code="substitute_strategy_missing",
            severity=FindingSeverity.BLOCKER,
            message="缺少替代策略",
            parent_code="BOARD-A",
            source_ids=("old:BOM:20",),
        ),
        ValidationFinding(
            code="substitute_mode_missing",
            severity=FindingSeverity.BLOCKER,
            message="缺少替代方式",
            parent_code="BOARD-A",
            source_ids=("old:BOM:20",),
        ),
        ValidationFinding(
            code="substitute_strategy_missing",
            severity=FindingSeverity.BLOCKER,
            message="缺少替代策略",
            parent_code="BOARD-A",
            source_ids=("new:BOM:21",),
        ),
    )

    result = compare_board_boms((), (), additional_findings=findings)

    assert result.summary.blocker_count == 3
    assert result.summary.blocking_record_count == 2
    assert result.summary.review_event_count == 1
    assert result.summary.metadata_event_count == 0
