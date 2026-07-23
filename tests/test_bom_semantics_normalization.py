from __future__ import annotations

import runpy
from pathlib import Path

from openpyxl import Workbook, load_workbook

from app.backend.bom_semantics.models import FindingSeverity
from app.backend.bom_semantics.normalization import normalize_workbook


BUILDER = Path(__file__).resolve().parent / "fixtures" / "bom_semantics" / "build_fixtures.py"


def _fixtures(tmp_path: Path) -> dict[str, Path]:
    return runpy.run_path(str(BUILDER))["build_all"](tmp_path)


def test_normalization_skips_repeated_headers_and_isolates_parents(tmp_path: Path) -> None:
    source = normalize_workbook(_fixtures(tmp_path)["multi_parent"])

    assert len(source.rows) == 3
    assert {row.parent_code for row in source.rows} == {"BOARD-A", "BOARD-B"}
    assert any(item.code == "repeated_header_skipped" for item in source.findings)
    numeric = [item for item in source.findings if item.code == "numeric_reference_suspected"]
    assert len(numeric) == 1
    assert numeric[0].severity == FindingSeverity.BLOCKER


def test_numeric_reference_resolution_is_bound_to_source_row(tmp_path: Path) -> None:
    path = _fixtures(tmp_path)["multi_parent"]
    source = normalize_workbook(path)
    row = next(row for row in source.rows if row.raw_reference == "72")

    resolved = normalize_workbook(
        path,
        reference_resolutions={row.source_id: "empty"},
    )
    resolved_row = next(item for item in resolved.rows if item.material_code == "PCB-B")
    assert resolved_row.references == ()
    assert not any(item.code == "numeric_reference_suspected" for item in resolved.findings)


def test_substitute_priority_and_decimal_quantity_are_preserved(tmp_path: Path) -> None:
    source = normalize_workbook(_fixtures(tmp_path)["substitutes"])

    assert [row.substitute_priority for row in source.rows] == [0, 1, 2]
    assert source.rows[0].quantity is not None
    assert str(source.rows[0].quantity) == "4"
    assert source.rows[2].references == ()


def test_template_instruction_rows_are_reported_without_becoming_materials(tmp_path: Path) -> None:
    path = _fixtures(tmp_path)["ordinary"]
    workbook = load_workbook(path)
    worksheet = workbook.active
    worksheet.append(["填写说明", "", "填写说明", "", "", "", "", "填写说明", "填写说明"])
    workbook.save(path)
    workbook.close()

    source = normalize_workbook(path)

    assert len(source.rows) == 2
    skipped = [item for item in source.findings if item.code == "template_instruction_skipped"]
    assert len(skipped) == 1


def test_capture_parent_can_be_aligned_for_two_file_comparison(tmp_path: Path) -> None:
    path = tmp_path / "capture.xlsx"
    workbook = load_workbook(_fixtures(tmp_path)["ordinary"])
    worksheet = workbook.active
    worksheet.delete_rows(1, worksheet.max_row)
    worksheet.append(["Reference", "Part Number", "Quantity", "Value"])
    worksheet.append(["R1", "MAT-R", 1, "10K"])
    workbook.save(path)
    workbook.close()

    source = normalize_workbook(path, default_parent_code="COMPARE-BOARD")

    assert {row.parent_code for row in source.rows} == {"COMPARE-BOARD"}


def test_capture_rows_with_blank_part_number_keep_value_evidence(tmp_path: Path) -> None:
    path = tmp_path / "blank-code.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["Reference", "Part Number", "Quantity", "Value"])
    worksheet.append(["H1", "", 1, "SC.DAX0501"])
    worksheet.append(["D8", "", 1, "NC/TE_2306654-3"])
    workbook.save(path)
    workbook.close()

    source = normalize_workbook(path)

    assert len(source.rows) == 2
    assert source.rows[0].material_code == ""
    assert source.rows[0].value == "SC.DAX0501"
    assert source.rows[0].references == ("H1",)
    assert source.rows[0].is_nc is False
    assert source.rows[1].value == "NC/TE_2306654-3"
    assert source.rows[1].is_nc is True
