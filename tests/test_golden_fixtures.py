from __future__ import annotations

import hashlib
import json
import re
import tempfile
import unittest
import zipfile
from pathlib import Path
from xml.etree import ElementTree

from openpyxl import load_workbook

from app.backend.parsers.bom_excel import read_bom_rows
from fixture_builders import build_capture_bom, build_processed_bom


FIXTURES = Path(__file__).parent / "fixtures"
SIGNATURE_FIELDS = ("model", "description", "name", "grade", "unit")
CORE_PROPERTIES = {
    "creator": "{http://purl.org/dc/elements/1.1/}creator",
    "last_modified_by": "{http://schemas.openxmlformats.org/package/2006/metadata/core-properties}lastModifiedBy",
    "created": "{http://purl.org/dc/terms/}created",
    "modified": "{http://purl.org/dc/terms/}modified",
}


def load_fixture(relative_path: str) -> dict[str, object]:
    return json.loads((FIXTURES / relative_path).read_text(encoding="utf-8"))


def parse_net_nodes(text: str) -> dict[str, set[str]]:
    nets: dict[str, set[str]] = {}
    current: str | None = None
    expect_name = False
    for raw in text.splitlines():
        line = raw.strip()
        if line == "NET_NAME":
            expect_name = True
            continue
        if expect_name:
            match = re.fullmatch(r"'([^']+)'", line)
            if match:
                current = match.group(1)
                nets[current] = set()
                expect_name = False
            continue
        tokens = line.split()
        if current and len(tokens) >= 3 and tokens[0] == "NODE_NAME":
            nets[current].add(f"{tokens[1]}.{tokens[2]}")
    return nets


def parse_part_packages(text: str) -> dict[str, str]:
    packages: dict[str, str] = {}
    for raw in text.splitlines():
        match = re.fullmatch(r"\s*([A-Za-z]+\d+)\s+'([^']+)':;", raw)
        if match:
            packages[match.group(1)] = match.group(2)
    return packages


def read_capture_records(path: Path) -> list[dict[str, object]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook.active
        headers = [cell.value for cell in sheet[1]]
        return [dict(zip(headers, row)) for row in sheet.iter_rows(min_row=2, values_only=True)]
    finally:
        workbook.close()


def sha256(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def read_core_properties(path: Path) -> tuple[dict[str, str], set[tuple[int, int, int, int, int, int]]]:
    with zipfile.ZipFile(path) as archive:
        root = ElementTree.fromstring(archive.read("docProps/core.xml"))
        properties = {field: root.findtext(tag) for field, tag in CORE_PROPERTIES.items()}
        timestamps = {entry.date_time for entry in archive.infolist()}
    return properties, timestamps


def signature_from_source(row: dict[str, object]) -> tuple[str, str, str, str, str]:
    return tuple(str(row[field]) for field in ("Model", "Description", "Name", "Grade", "Unit"))


def signature_from_oracle(signature: dict[str, object]) -> tuple[str, str, str, str, str]:
    return tuple(str(signature[field]) for field in SIGNATURE_FIELDS)


def derive_conflict_reason(signatures: set[tuple[str, str, str, str, str]]) -> str:
    for model, description, *_ in signatures:
        for other_model, other_description, *_ in signatures:
            if (model, description) != (other_model, other_description) and other_model.startswith(model) and other_description.startswith(description):
                return "truncation_prefix_completion"
    if len({(model, description, name, unit) for model, description, name, _, unit in signatures}) == 1:
        return "grade_only_conflict"
    return "multiple_complete_candidates"


class GoldenFixtureTests(unittest.TestCase):
    def test_sanitized_bom_cases_preserve_conflicts_and_risk_expectations(self) -> None:
        cases = load_fixture("bom/conflict_cases.json")
        expected = load_fixture("bom/expected_recommendations.json")
        rows = cases["rows"]
        self.assertIsInstance(rows, dict)
        row_values = list(rows.values())

        recommendations = expected["recommendations"]
        self.assertEqual(set(recommendations), set(expected["conflict_codes"]))
        for code, expectation in recommendations.items():
            source_signatures = {signature_from_source(row) for row in row_values if row["Part Number"] == code}
            allowed = {signature_from_oracle(signature) for signature in expectation["allowed_signatures"]}
            self.assertEqual(allowed, source_signatures)
            self.assertEqual(expectation["expected_reason"], derive_conflict_reason(source_signatures))
            high_confidence = expectation["expected_confidence"] == "high"
            self.assertEqual(expectation["high_confidence"], high_confidence)
            self.assertEqual(expectation["manual_choice_required"], not high_confidence)
            if high_confidence:
                selected = expectation["selected_signature"]
                self.assertEqual(set(selected), set(SIGNATURE_FIELDS))
                self.assertIn(signature_from_oracle(selected), source_signatures)
            else:
                self.assertNotIn("selected_signature", expectation)

        risks = expected["risk_expectations"]
        shield = risks["shield_candidate"]
        self.assertTrue(shield["requires_confirmation"])
        self.assertTrue(shield["enters_final_bom"])
        self.assertTrue(shield["excluded_from_nc"])
        shield_row = next(row for row in row_values if row["Reference"] == shield["reference"])
        self.assertEqual(shield_row["Part Number"], shield["part_number"])
        duplicate = risks["duplicate_reference"]
        duplicate_rows = [row for row in row_values if duplicate["reference"] in row["Reference"].split()]
        self.assertEqual({row["Part Number"] for row in duplicate_rows}, set(duplicate["part_numbers"]))

    def test_capture_builder_writes_the_full_sanitized_case(self) -> None:
        expected = load_fixture("bom/expected_recommendations.json")
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "capture.xlsx"
            second_path = Path(tmp) / "capture-second.xlsx"
            result = build_capture_bom(path, "golden")
            build_capture_bom(second_path, "golden")

            self.assertEqual(result, path)
            self.assertTrue(path.exists())
            self.assertEqual(sha256(path), sha256(second_path))
            core_properties, timestamps = read_core_properties(path)
            self.assertEqual(core_properties, expected["workbook_expectations"]["core_properties"])
            self.assertEqual(timestamps, {tuple(expected["workbook_expectations"]["zip_timestamp"])})
            workbook = load_workbook(path, read_only=True, data_only=True)
            try:
                sheet = workbook.active
                headers = [cell.value for cell in sheet[1]]
                records = [dict(zip(headers, row)) for row in sheet.iter_rows(min_row=2, values_only=True)]
            finally:
                workbook.close()

        self.assertIn("{Reference}", headers)
        self.assertIn("{Part Number}", headers)
        self.assertTrue(set(expected["conflict_codes"]) <= {record["{Part Number}"] for record in records})
        duplicate = expected["risk_expectations"]["duplicate_reference"]
        self.assertEqual(
            sum(duplicate["reference"] in record["{Reference}"].split() for record in records),
            len(duplicate["part_numbers"]),
        )
        shield_expectation = expected["risk_expectations"]["shield_candidate"]
        shield = next(record for record in records if record["{Reference}"] == shield_expectation["reference"])
        self.assertEqual(shield["{Part Number}"], shield_expectation["part_number"])

    def test_capture_builder_emits_concrete_ddr_and_emmc_rows(self) -> None:
        expected = load_fixture("bom/expected_recommendations.json")
        warnings = expected["risk_expectations"]["version_warnings"]
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "version_sensitive.xlsx"
            build_capture_bom(path, "version_sensitive")
            rows = read_capture_records(path)

        by_reference = {row["{Reference}"]: row for row in rows}
        self.assertEqual(set(by_reference), {warning["reference"] for warning in warnings})
        for warning in warnings:
            row = by_reference[warning["reference"]]
            self.assertEqual(row["{Part Number}"], warning["part_number"])
            self.assertEqual(row["{规格型号}"], warning["model"])
            self.assertEqual(row["{器件描述（新整理）}"], warning["description"])
            self.assertIn(warning["kind"].casefold(), " ".join(str(value) for value in row.values()).casefold())

    def test_package_size_conflict_joins_the_same_bom_and_netlist_reference(self) -> None:
        expected = load_fixture("bom/expected_recommendations.json")
        conflict = expected["risk_expectations"]["package_size_conflict"]
        netlist_parts = parse_part_packages((FIXTURES / "netlist/pstxprt_sample.dat").read_text(encoding="utf-8"))
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "package_conflict.xlsx"
            build_capture_bom(path, "package_size_conflict")
            rows = read_capture_records(path)

        self.assertEqual(len(rows), 1)
        bom_row = rows[0]
        self.assertEqual(bom_row["{Reference}"], conflict["reference"])
        self.assertEqual(netlist_parts[conflict["reference"]], conflict["netlist_package"])
        self.assertEqual(bom_row["{Part Number}"], conflict["bom_part_number"])
        self.assertEqual(bom_row["{规格型号}"], conflict["bom_model"])
        self.assertEqual(bom_row["{器件描述（新整理）}"], conflict["bom_description"])
        self.assertEqual(bom_row["{PCB封装}"], conflict["bom_package"])
        netlist_size = re.search(r"R(\d{4})", conflict["netlist_package"])
        bom_size = re.search(r"\b(\d{4})\b", str(bom_row["{PCB封装}"]))
        self.assertIsNotNone(netlist_size)
        self.assertIsNotNone(bom_size)
        self.assertEqual(netlist_size.group(1), conflict["netlist_size"])
        self.assertEqual(bom_size.group(1), conflict["bom_size"])
        self.assertNotEqual(netlist_size.group(1), bom_size.group(1))

    def test_processed_bom_builders_match_oracle_reader_and_reproducible_bytes(self) -> None:
        expected = load_fixture("bom/expected_recommendations.json")
        workbook_expectations = expected["workbook_expectations"]
        for template, oracle in expected["processed_templates"].items():
            with self.subTest(template=template), tempfile.TemporaryDirectory() as tmp:
                path = Path(tmp) / f"{template}-first.xlsx"
                second_path = Path(tmp) / f"{template}-second.xlsx"
                result = build_processed_bom(path, template)
                build_processed_bom(second_path, template)

                self.assertEqual(result, path)
                self.assertEqual(sha256(path), sha256(second_path))
                workbook = load_workbook(path, read_only=True, data_only=True)
                try:
                    sheet = workbook.active
                    self.assertEqual(sheet.title, oracle["sheet_title"])
                    if "group_header" in oracle:
                        self.assertEqual(
                            ["" if cell.value is None else cell.value for cell in sheet[1]],
                            oracle["group_header"],
                        )
                    headers = [cell.value for cell in sheet[oracle["header_row"]]]
                    row = [
                        "" if cell.value is None else cell.value
                        for cell in sheet[oracle["data_row"]]
                    ]
                finally:
                    workbook.close()

                self.assertEqual(headers, oracle["headers"])
                self.assertEqual(row, oracle["row"])
                self.assertEqual(len(headers), oracle["column_count"])
                reader_rows = read_bom_rows(path)
                self.assertEqual(len(reader_rows), 1)
                self.assertEqual(
                    {field: reader_rows[0][field] for field in oracle["reader"]},
                    oracle["reader"],
                )
                core_properties, timestamps = read_core_properties(path)
                self.assertEqual(core_properties, workbook_expectations["core_properties"])
                self.assertEqual(timestamps, {tuple(workbook_expectations["zip_timestamp"])})

    def test_netlist_samples_keep_real_format_and_expected_edge_case_tokens(self) -> None:
        expected = load_fixture("bom/expected_recommendations.json")
        package_conflict = expected["risk_expectations"]["package_size_conflict"]
        netlist = (FIXTURES / "netlist/pstxnet_sample.dat").read_text(encoding="utf-8")
        parts = (FIXTURES / "netlist/pstxprt_sample.dat").read_text(encoding="utf-8")
        nets = parse_net_nodes(netlist)
        packages = parse_part_packages(parts)

        self.assertIn("FILE_TYPE = EXPANDEDNETLIST;", netlist)
        self.assertEqual(nets["RENAME_OLD"], nets["RENAME_NEW"])
        self.assertEqual(nets["SPLIT_OLD"], nets["SPLIT_NEW_A"] | nets["SPLIT_NEW_B"])
        self.assertEqual(nets["MERGE_NEW"], nets["MERGE_OLD_A"] | nets["MERGE_OLD_B"])
        self.assertEqual(nets["MIXED_TOPOLOGY"], {"R401.1", "U401.A1", "U401.B1", "C401.1"})
        self.assertIn("NC_7", nets)
        self.assertIn("NCLK", nets)
        self.assertIn("nCS0", nets)
        self.assertEqual(nets["NC_7"], {"C701.1"})
        self.assertIn("FILE_TYPE = EXPANDEDPARTLIST;", parts)
        self.assertEqual(packages[package_conflict["reference"]], package_conflict["netlist_package"])
        self.assertIn(package_conflict["netlist_size"], package_conflict["netlist_package"])
        self.assertIn(package_conflict["bom_size"], parts)
        self.assertEqual(packages["C701"], "CAP_NP_C0201_NC")


if __name__ == "__main__":
    unittest.main()
