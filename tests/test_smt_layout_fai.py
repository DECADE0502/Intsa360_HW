from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from app.backend.tools.smt_layout import _build_fai_table, _write_fai_xlsx


EXPECTED_HEADERS = [
    "位号",
    "面",
    "X(mm)",
    "Y(mm)",
    "封装",
    "应贴料号",
    "应贴型号",
    "应贴描述",
    "优选等级",
    "QC",
    "备注",
]


def _component(
    ref: str,
    *,
    side: str = "top",
    x_mm: float = 0.0,
    y_mm: float = 0.0,
    grade: str = "优选",
    part_number: str = "PN-1",
    status: str = "installed",
) -> dict[str, object]:
    return {
        "ref": ref,
        "side": side,
        "x_mm": x_mm,
        "y_mm": y_mm,
        "rotation": 0,
        "footprint": "R0402",
        "part_number": part_number,
        "model": "10K",
        "description": "Resistor",
        "grade": grade,
        "status": status,
        "high_risk": False,
    }


def test_fai_table_columns_are_fixed_chinese() -> None:
    table = _build_fai_table([_component("R1")])

    assert table["headers"] == EXPECTED_HEADERS


def test_fai_serpentine_ordering_by_side_then_yx() -> None:
    table = _build_fai_table(
        [
            _component("B2", side="bottom", x_mm=20, y_mm=3.0),
            _component("T1", x_mm=20, y_mm=3.0),
            _component("T3", x_mm=5, y_mm=6.0),
            _component("B1", side="bottom", x_mm=10, y_mm=3.1),
            _component("T2", x_mm=10, y_mm=3.1),
        ]
    )

    assert [row[0] for row in table["rows"]] == ["T3", "T2", "T1", "B1", "B2"]


def test_fai_flags_grade_warn_rows() -> None:
    table = _build_fai_table(
        [
            _component("R1", grade="优选"),
            _component("R2", grade="限制使用"),
        ]
    )

    rows = {row[0]: row for row in table["rows"]}
    assert rows["R1"][10] == ""
    assert rows["R2"][10] == "⚠ 等级"


def test_fai_xlsx_writes_all_rows_and_headers_bold(tmp_path: Path) -> None:
    table = _build_fai_table(
        [
            _component("R1"),
            _component("R2", grade="限制使用", part_number="", status="missing_bom"),
        ]
    )

    output = _write_fai_xlsx(tmp_path, "BOARD_A", "20260721_120000_000", table)

    assert output == tmp_path / "首件核对表_BOARD_A_20260721_120000_000.xlsx"
    workbook = load_workbook(output, read_only=False, data_only=True)
    try:
        sheet = workbook.active
        assert [cell.value for cell in sheet[1]] == EXPECTED_HEADERS
        assert sheet.max_row == 3
        assert all(cell.font.bold for cell in sheet[1])
        assert sheet[3][0].fill.fgColor.rgb.endswith("FFF7E6")
        assert sheet.page_setup.orientation == "landscape"
    finally:
        workbook.close()
