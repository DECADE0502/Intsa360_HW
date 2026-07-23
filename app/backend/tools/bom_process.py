from __future__ import annotations

import hashlib
import json
import re
from collections import Counter, OrderedDict
from dataclasses import dataclass, field, replace
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Mapping

from app.backend.capture_fields import BOM_OPTIONAL_FIELDS, FIELD_DEFAULTS, PLM_TEMPLATE_HEADERS
from app.backend.parsers._workbook import build_merged_cell_lookup, open_bom_workbook
from app.backend.parsers.bom_table import INHERIT_FIELDS, normalize_header, split_refs
from app.backend.parsers.refs import natural_key
from app.backend.tools.bom_classify import (
    CODE_CANDIDATE_FIELDS,
    PROCESS_MATERIAL_RE,
    NormalizedBomRow,
    build_normalized_row,
    clean_field_text,
    classification_config,
    code_shape_matches,
    process_keyword,
)
from app.backend.tools.bom_domain import (
    BOM_RULE_VERSION,
    PhysicalPart,
    RefOccurrence,
    SourceQualityIssue,
    SourceQualityReport,
    stable_fingerprint,
)
from app.backend.tools.bom_rules import NC_VALUE_RE
from app.backend.tools.bom_semantic_manifest import write_semantic_manifest

# BOM 处理工具：把 Capture 导出的原始 BOM 处理成可导入的 PLM / OA 成品。
# - 自动定位表头行（Capture 导出前面常有标题块），表头带不带 {} 花括号都能识别
# - 按物料属性过滤 NC/未贴和疑似工艺件；SH 屏蔽支架单独确认；按料号合并位号、统计数量
#   （从 Capture 按版本导出的 BOM 本身已是该版本的真实器件，无需再做版本换料）
# - 输出 PLM（19 列）和/或 OA（16 列）成品，并可追加 PCB/屏蔽支架等附加物料
# - 同时产出 NC/未贴器件汇总

# 源列别名 -> 规范字段
SRC_ALIASES = {
    "item": ["Item", "序号", "项次"],
    "quantity": ["Quantity", "数量", "Qty", "用量"],
    "reference": ["Reference", "位号", "Designator", "RefDes"],
    "part_number": ["Part Number", "子项编码", "料号", "物料编码", "PN"],
    "value": ["Value", "值"],
    "model": ["规格型号", "型号", "Model", "MPN"],
    "desc": ["器件描述（新整理）", "器件描述", "内容", "描述", "Description"],
    "old_desc": ["器件描述（旧）", "旧描述", "Old Description"],
    "name": ["物料名称", "名称", "Part Type", "Name"],
    "grade": ["等级", "物料优选等级", "优选等级"],
    "unit": ["单位", "Unit", "UOM"],
    "remark": ["备注", "Remark", "Note"],
    "grade_remark": ["物料优选等级备注", "等级备注"],
    "alt_group": ["替代组编码"],
    "alt_strategy": ["替代策略"],
    "alt_method": ["替代方式"],
    "alt_priority": ["替代优先级"],
    "issue_method": ["发料方式", "领料方式", "发料"],
    "mrp": ["是否参与MRP运算"],
    "jump_level": ["是否跳层"],
    "pcb_footprint": ["PCB Footprint", "Footprint"],
    "pcb_package": ["PCB封装", "封装"],
    "source_package": ["Source Package"],
    "source_part": ["Source Part"],
    "part_type": ["Part Type"],
    "part_reference": ["Part Reference"],
    # Capture trace fields are retained for diagnostics and future rules. They do
    # not become columns in the final PLM/OA output unless explicitly mapped.
    "manufacturer": ["制造商", "Manufacturer"],
    "datasheet": ["datasheet", "Datasheet"],
    "source_library": ["Source Library"],
    "designator": ["Designator"],
    "color": ["Color"],
    "implementation": ["Implementation"],
    "implementation_path": ["Implementation Path"],
    "implementation_type": ["Implementation Type"],
    "primitive": ["Primitive"],
    "graphic": ["Graphic", "Graphic ID"],
    "capture_id": ["ID"],
    "original_symbol_origin": ["OriginalSymbolOrigin"],
    "power_pins_visible": ["Power Pins Visible"],
    "location_x": ["Location X-Coordinate"],
    "location_y": ["Location Y-Coordinate"],
    "split_inst": ["SPLIT_INST"],
    "swap_info": ["SWAP_INFO"],
}

PLM_R1 = ["父项信息", "", "子项信息", "", "", "", "", "", "BOM明细", "", "", "", "", "", "", "", "", "", ""]
PLM_HEADERS = PLM_TEMPLATE_HEADERS
OA_HEADERS = [
    "序号", "编码（父）*##bmf", "描述（父）##msf", "编码（子）*##bm", "描述（子）##ms", "数量*##sl",
    "单位*##dw", "位号##wh", "备注##bz", "替代组编码##tdzbm", "替代策略##tdcl", "发料方式##tdfs",
    "物料优选等级##tdyxj", "领料方式*##flfs", "是否参与MRP运算*##sfcymrpys", "是否跳层*##sftc",
]
NC_HEADERS = ["原始行号", "位号", "子项编码", "物料名称", "型号", "描述", "Value", "过滤原因", "判定类型"]
NON_SMT_HEADERS = NC_HEADERS
DECISION_HEADERS = [
    "位号",
    "子项编码",
    "身份状态",
    "分类状态",
    "角色",
    "子类型",
    "目标区域",
    "排除类型",
    "规则",
    "规则版本",
    "决策来源",
    "决策指纹",
    "原因/证据",
]
CONFLICT_FIELDS = (
    "name",
    "value",
    "model",
    "desc",
    "pcb_footprint",
    "pcb_package",
    "manufacturer",
    "grade",
    "unit",
)
_PROCESS_MATERIAL_RE = PROCESS_MATERIAL_RE
_NUMERIC_VALUE_RE = re.compile(
    r"^\d+(?:\.\d+)?(?:\s*[RrKkMmGgTtUuNnPpFfHhVv]\d*)?"
    r"\s*(?:Ω|ohms?|%|℃|°C|[VvAaWwFfHh])?$",
    re.IGNORECASE,
)
def _looks_numeric(value: str) -> bool:
    """Return whether a value is a numeric component specification."""
    return bool(_NUMERIC_VALUE_RE.fullmatch(value.strip()))
@dataclass(frozen=True)
class ParsedSource:
    source_path: Path
    raw_rows: list[dict[str, str]]
    row_numbers: list[int]
    normalized_rows: tuple[NormalizedBomRow, ...] = ()
    source_fingerprint: str = ""
    quality_report: SourceQualityReport = SourceQualityReport()
    occurrences: tuple[RefOccurrence, ...] = ()
    physical_parts: tuple[PhysicalPart, ...] = ()
    physical_refs_by_row: dict[int, tuple[str, ...]] = field(default_factory=dict)


def normalize_ref(ref: str) -> str:
    return re.sub(r"\s+", "", str(ref or "").strip()).upper()


def detect_header(ws) -> tuple[int, dict[str, int]]:
    """扫描前若干行，定位含 Reference + Part Number 的表头行，返回行号与字段列映射。"""
    alias_norm = {key: [normalize_header(a) for a in names] for key, names in SRC_ALIASES.items()}
    best_row, best_map, best_score = 1, {}, -1
    for row in range(1, min(ws.max_row, 40) + 1):
        values = [normalize_header(ws.cell(row, col).value) for col in range(1, ws.max_column + 1)]
        mapping: dict[str, int] = {}
        for key, names in alias_norm.items():
            for alias in names:
                for idx, value in enumerate(values, start=1):
                    if value and value == alias:
                        mapping[key] = idx
                        break
                if key in mapping:
                    break
        score = sum(1 for k in ("reference", "part_number") if k in mapping) * 10 + len(mapping)
        if score > best_score:
            best_row, best_map, best_score = row, mapping, score
    if "reference" not in best_map or "part_number" not in best_map:
        raise ValueError("表头识别失败：未找到 Reference / Part Number 列，请确认按导出配置生成的原始 BOM。")
    return best_row, best_map


def _cell(
    ws,
    row: int,
    mapping: dict[str, int],
    key: str,
    merged_lookup: dict[tuple[int, int], object] | None = None,
) -> str:
    value, _ = _cell_details(ws, row, mapping, key, merged_lookup)
    return value


def _cell_details(
    ws,
    row: int,
    mapping: dict[str, int],
    key: str,
    merged_lookup: dict[tuple[int, int], object] | None = None,
) -> tuple[str, str]:
    col = mapping.get(key)
    if not col:
        return "", "cell"
    raw_value = ws.cell(row, col).value
    provenance = "cell"
    source_field = {
        "desc": "description",
        "pcb_footprint": "package",
        "pcb_package": "package",
        "source_package": "package",
    }.get(key, key)
    if raw_value is None and merged_lookup is not None and source_field in INHERIT_FIELDS:
        raw_value = merged_lookup.get((row, col))
        if raw_value is not None:
            provenance = "merged_inherit"
    return str(raw_value or "").strip(), provenance


def has_shield_refs(refs: list[str]) -> bool:
    return any(ref.upper().startswith("SH") for ref in refs)


def _material_text(row: dict[str, str]) -> str:
    return " ".join(
        str(row.get(field) or "").strip()
        for field in ("desc", "name", "model", "pcb_package", "pcb_footprint")
    )


def _matches_process_material(row: dict[str, str]) -> str | None:
    return process_keyword(_material_text(row)) or None


def _process_candidate_key(part_number: str, refs: list[str]) -> str:
    normalized = sorted({normalize_ref(ref) for ref in refs}, key=natural_key)
    return f"{part_number.strip()}|{','.join(normalized)}"


def _classify_no_partnumber(refs: list[str], row: dict[str, str]) -> str:
    if row.get("_missing_part_number_decision") == "exclude":
        return "用户确认不装（空编码）"
    upper = [ref.upper() for ref in refs]
    value = str(row.get("value") or "").strip()
    if NC_VALUE_RE.fullmatch(value):
        return "NC/未贴（无料号）"
    if any(ref.startswith("SH") for ref in upper):
        return "疑似屏蔽支架 SH*（无料号）"
    if any(ref.startswith(("TP", "Z_TP")) for ref in upper):
        return "疑似测试点 TP*/Z_TP*（无料号）"
    if any(ref.startswith("JP") for ref in upper):
        return "疑似跳线 JP*（无料号）"
    if any(re.fullmatch(r"H\d+", ref) for ref in upper):
        return "疑似安装孔 H*（无料号）"
    if any(ref.startswith(("MH", "MTG")) for ref in upper):
        return "疑似安装孔 MH*/MTG*（无料号）"
    return "子项编码为空"


def _normalize_shield_row(row: dict[str, str], refs: list[str]) -> None:
    if (
        not has_shield_refs(refs)
        or str(row.get("_placement_role") or "") != "shield"
        or str(row.get("_placement_subtype") or "") != "bracket"
    ):
        return
    if not row.get("name", "").strip():
        row["name"] = "屏蔽支架"
    if not row.get("desc", "").strip():
        row["desc"] = row.get("value", "").strip() or "屏蔽支架"


def _processing_row(row: dict[str, str], refs: list[str]) -> dict[str, str]:
    normalized = dict(row)
    raw_flags = normalized.get("_field_flags")
    field_flags = raw_flags if isinstance(raw_flags, dict) else {}
    value_flags = set(field_flags.get("value") or [])
    safe_value = not value_flags.intersection({
        "code_shape",
        "nc_keyword",
        "process_keyword",
        "placeholder_residue",
        "mojibake",
    })
    if not normalized.get("name"):
        normalized["name"] = normalized.get("part_type") or ""
    if not normalized.get("model"):
        normalized["model"] = (
            (normalized.get("value") if safe_value else "")
            or normalized.get("pcb_package")
            or normalized.get("pcb_footprint")
            or ""
        )
    if not normalized.get("desc") and safe_value:
        normalized["desc"] = normalized.get("value") or ""
    _normalize_shield_row(normalized, refs)
    return normalized


def exclusion_reason(
    row: dict[str, str],
    refs: list[str],
    include_shields: bool = False,
    process_material_keeps: set[str] | None = None,
) -> str | None:
    placement_destination = str(row.get("_placement_destination") or "").strip().lower()
    if placement_destination == "non_smt":
        return str(row.get("_placement_reason") or "用户确认移出贴片 BOM")
    if placement_destination == "smt":
        return None
    placement_action = str(row.get("_placement_action") or "").strip().lower()
    if placement_action == "exclude":
        return str(row.get("_placement_reason") or "用户确认不装")
    if placement_action in {"keep", "keep_as_is"}:
        return None
    part_number = str(row.get("part_number") or "").strip()
    if not part_number:
        return _classify_no_partnumber(refs, row)
    upper = [r.upper() for r in refs]
    if any(r.startswith("SH") for r in upper):
        if include_shields:
            return None
        return "屏蔽支架 SH*"
    value = str(row.get("value") or "").strip()
    if NC_VALUE_RE.fullmatch(value):
        return "NC/未贴"
    keyword = _matches_process_material(row)
    if keyword:
        keeps = process_material_keeps or set()
        if _process_candidate_key(part_number, refs) in keeps:
            return None
        return f"工艺件（描述含 {keyword}）"
    return None


def _exclusion_reason_kind(row: dict[str, str], reason: str) -> str:
    explicit = str(
        row.get("_placement_exclusion_kind")
        or row.get("_placement_reason_kind")
        or ""
    ).strip()
    if explicit:
        return explicit
    if reason.startswith("NC/未贴"):
        return "system_nc"
    if reason.startswith("工艺件"):
        return "process_default"
    if reason.startswith("用户确认不装"):
        return "user_excluded"
    return "legacy_filter"


_REF_TOKEN_SPLIT_RE = re.compile(r"[,，;；\s]+")
_REF_RANGE_RE = re.compile(
    r"^(?P<prefix>[A-Za-z_]+)(?P<start>\d+)[-~～–—](?:(?P<end_prefix>[A-Za-z_]+))?(?P<end>\d+)$"
)
_MULTI_UNIT_REF_RE = re.compile(r"^(?P<base>[A-Za-z_]+\d+)(?P<unit>[A-Za-z])$")


def _quantity_as_int(value: object) -> int | None:
    text = str(value or "").strip()
    if not text:
        return None
    try:
        number = Decimal(text)
    except InvalidOperation:
        return None
    if number != number.to_integral_value() or number < 0:
        return None
    return int(number)


def _reference_tokens(
    raw_value: object,
    quantity: object,
    row_number: int,
) -> tuple[list[str], list[SourceQualityIssue]]:
    raw = str(raw_value or "").strip()
    if not raw:
        return [], []
    tokens = [token for token in _REF_TOKEN_SPLIT_RE.split(raw) if token]
    quantity_value = _quantity_as_int(quantity)
    issues: list[SourceQualityIssue] = []
    expanded: list[str] = []
    for token in tokens:
        match = _REF_RANGE_RE.fullmatch(token)
        if match is None:
            expanded.append(normalize_ref(token))
            continue
        prefix = match.group("prefix").upper()
        end_prefix = (match.group("end_prefix") or prefix).upper()
        start = int(match.group("start"))
        end = int(match.group("end"))
        count = end - start + 1
        can_expand = (
            len(tokens) == 1
            and prefix == end_prefix
            and start <= end
            and count <= 10000
            and quantity_value == count
        )
        if can_expand:
            expanded.extend(f"{prefix}{value}" for value in range(start, end + 1))
            continue
        expanded.append(normalize_ref(token))
        issues.append(SourceQualityIssue(
            "unexpanded_reference_range",
            "warning",
            "范围位号未满足同前缀、递增和 Quantity 一致条件，已保留原文等待审查。",
            (row_number,),
            (token,),
            {"quantity": str(quantity or ""), "calculated_count": max(count, 0)},
        ))
    return expanded, issues


def _package_signature(row: dict[str, str]) -> str:
    return "\x1f".join(
        str(row.get(field) or "").strip().casefold()
        for field in ("pcb_footprint", "pcb_package", "source_package", "source_part")
    )


def _physical_reference_model(
    row_payloads: list[tuple[int, dict[str, str], list[str]]],
) -> tuple[
    tuple[RefOccurrence, ...],
    tuple[PhysicalPart, ...],
    dict[int, tuple[str, ...]],
    dict[int, tuple[str, ...]],
    list[SourceQualityIssue],
]:
    candidates: list[dict[str, object]] = []
    for row_number, row, refs in row_payloads:
        part_number = str(row.get("part_number") or "").strip()
        package_signature = _package_signature(row)
        for ref in refs:
            unit_match = _MULTI_UNIT_REF_RE.fullmatch(ref)
            candidates.append({
                "source_row": row_number,
                "raw_ref": ref,
                "normalized_ref": normalize_ref(ref),
                "physical_ref": normalize_ref(ref),
                "part_number": part_number,
                "package_signature": package_signature,
                "unit_base": unit_match.group("base").upper() if unit_match else "",
                "unit_marker": unit_match.group("unit").upper() if unit_match else "",
            })

    by_unit_base: dict[str, list[dict[str, object]]] = {}
    for candidate in candidates:
        unit_base = str(candidate["unit_base"])
        if unit_base:
            by_unit_base.setdefault(unit_base, []).append(candidate)

    issues: list[SourceQualityIssue] = []
    for base, items in by_unit_base.items():
        distinct_refs = {str(item["normalized_ref"]) for item in items}
        part_numbers = {str(item["part_number"]).casefold() for item in items if str(item["part_number"])}
        packages = {str(item["package_signature"]) for item in items if str(item["package_signature"]).strip("\x1f")}
        can_merge = len(distinct_refs) >= 2 and len(part_numbers) == 1 and len(packages) == 1
        if not can_merge:
            continue
        for item in items:
            item["physical_ref"] = base
        issues.append(SourceQualityIssue(
            "multi_unit_merged",
            "info",
            "多单元器件在料号和封装证据一致时归并为一个物理位号。",
            tuple(sorted({int(item["source_row"]) for item in items})),
            tuple(sorted(distinct_refs, key=natural_key)),
            {"physical_ref": base},
        ))

    grouped: "OrderedDict[str, list[dict[str, object]]]" = OrderedDict()
    for candidate in candidates:
        grouped.setdefault(str(candidate["physical_ref"]), []).append(candidate)

    physical_parts: list[PhysicalPart] = []
    conflicts_by_row: dict[int, set[str]] = {}
    for physical_ref, items in grouped.items():
        source_rows = tuple(sorted({int(item["source_row"]) for item in items}))
        source_refs = tuple(sorted({str(item["normalized_ref"]) for item in items}, key=natural_key))
        part_numbers = {str(item["part_number"]).strip().casefold() for item in items if str(item["part_number"]).strip()}
        packages = {
            str(item["package_signature"]) for item in items
            if str(item["package_signature"]).strip("\x1f")
        }
        conflicts: list[str] = []
        if len(part_numbers) > 1:
            conflicts.append("same_physical_ref_multiple_part_numbers")
        if len(packages) > 1 and len(items) > 1:
            conflicts.append("same_physical_ref_multiple_packages")
        if conflicts:
            for row_number in source_rows:
                conflicts_by_row.setdefault(row_number, set()).update(conflicts)
            issues.append(SourceQualityIssue(
                "physical_identity_conflict",
                "error",
                "同一物理位号出现不同料号或封装，禁止自动归并。",
                source_rows,
                (physical_ref,),
                {"conflicts": conflicts},
            ))
        merge_kind = "single"
        if len(source_refs) > 1 and any(str(item["unit_marker"]) for item in items):
            merge_kind = "multi_unit"
        elif len(items) > 1:
            merge_kind = "duplicate"
            issues.append(SourceQualityIssue(
                "duplicate_occurrence",
                "info",
                "同一物理位号在多行出现，已保留全部来源并按一个物理器件计数。",
                source_rows,
                (physical_ref,),
                {"occurrences": len(items)},
            ))
        physical_parts.append(PhysicalPart(
            physical_ref,
            source_rows,
            len(items),
            merge_kind,
            source_refs,
            tuple(conflicts),
        ))

    occurrences = tuple(
        RefOccurrence(
            int(item["source_row"]),
            str(item["raw_ref"]),
            str(item["normalized_ref"]),
            str(item["physical_ref"]),
            str(item["part_number"]),
            str(item["package_signature"]),
            str(item["unit_marker"]),
        )
        for item in candidates
    )
    refs_by_row: dict[int, set[str]] = {}
    for occurrence in occurrences:
        refs_by_row.setdefault(occurrence.source_row, set()).add(occurrence.physical_ref)
    return (
        occurrences,
        tuple(sorted(physical_parts, key=lambda item: natural_key(item.reference))),
        {row: tuple(sorted(refs, key=natural_key)) for row, refs in refs_by_row.items()},
        {row: tuple(sorted(values)) for row, values in conflicts_by_row.items()},
        issues,
    )


def parse_source(path: Path) -> ParsedSource:
    with open_bom_workbook(path, data_only=True) as wb:
        ws = wb.active
        merged_lookup = build_merged_cell_lookup(ws)
        header_row, mapping = detect_header(ws)
        rows: list[dict[str, str]] = []
        row_numbers: list[int] = []
        provenances: list[dict[str, str]] = []
        row_payloads: list[tuple[int, dict[str, str], list[str]]] = []
        quality_issues: list[SourceQualityIssue] = []
        for row_num in range(header_row + 1, ws.max_row + 1):
            details = {
                key: _cell_details(ws, row_num, mapping, key, merged_lookup)
                for key in SRC_ALIASES
            }
            row = {key: value for key, (value, _) in details.items()}
            if not row.get("desc") and row.get("old_desc"):
                row["desc"] = row["old_desc"]
                details["desc"] = (row["desc"], details["old_desc"][1])
            if not any(row.values()):
                continue
            if not any(
                str(value or "").strip()
                for key, value in row.items()
                if key not in {"item", "quantity"}
            ):
                quality_issues.append(SourceQualityIssue(
                    "noise_row_skipped",
                    "info",
                    "仅含 Item/Quantity 的标题、分页或备注噪声行已跳过。",
                    (row_num,),
                    details={"item": row.get("item", ""), "quantity": row.get("quantity", "")},
                ))
                continue
            refs, ref_issues = _reference_tokens(row.get("reference"), row.get("quantity"), row_num)
            quality_issues.extend(ref_issues)
            quantity_value = _quantity_as_int(row.get("quantity"))
            if str(row.get("quantity") or "").strip() and quantity_value is None:
                quality_issues.append(SourceQualityIssue(
                    "invalid_quantity",
                    "warning",
                    "Quantity 不是非负整数，原值已保留且不会改变装机结论。",
                    (row_num,),
                    tuple(refs),
                    {"quantity": row.get("quantity", "")},
                ))
            elif quantity_value is not None and quantity_value != len(refs):
                quality_issues.append(SourceQualityIssue(
                    "quantity_mismatch",
                    "warning",
                    "Quantity 与拆分后的位号数不一致，原值已保留且不会改变装机结论。",
                    (row_num,),
                    tuple(refs),
                    {"quantity": quantity_value, "reference_count": len(refs)},
                ))
            if not refs and any(str(row.get(field) or "").strip() for field in ("part_number", "value", "name", "model", "desc")):
                quality_issues.append(SourceQualityIssue(
                    "missing_reference",
                    "warning",
                    "物料行缺少可识别位号，保留该行并要求人工审查。",
                    (row_num,),
                ))
            rows.append(row)
            row_numbers.append(row_num)
            provenances.append({key: provenance for key, (_, provenance) in details.items()})
            row_payloads.append((row_num, row, refs))

        occurrences, physical_parts, refs_by_row, conflicts_by_row, physical_issues = _physical_reference_model(row_payloads)
        quality_issues.extend(physical_issues)
        normalized_rows: list[NormalizedBomRow] = []
        for row, row_num, provenance, (_, _, source_refs) in zip(rows, row_numbers, provenances, row_payloads):
            physical_refs = refs_by_row.get(row_num, ())
            row["_source_reference"] = row.get("reference", "")
            row["reference"] = ",".join(physical_refs)
            normalized_rows.append(build_normalized_row(
                row_num,
                physical_refs,
                row,
                provenance,
                source_refs=source_refs,
                quantity=row.get("quantity", ""),
                physical_conflicts=conflicts_by_row.get(row_num, ()),
            ))

        source_path = Path(path)
        source_fingerprint = hashlib.sha256(source_path.read_bytes()).hexdigest()[:24]
        quality_report = SourceQualityReport(
            source_rows=max(ws.max_row - header_row, 0),
            parsed_rows=len(rows),
            occurrence_count=len(occurrences),
            physical_part_count=len(physical_parts),
            issues=tuple(quality_issues),
        )
        return ParsedSource(
            source_path=source_path,
            raw_rows=rows,
            row_numbers=row_numbers,
            normalized_rows=tuple(normalized_rows),
            source_fingerprint=source_fingerprint,
            quality_report=quality_report,
            occurrences=occurrences,
            physical_parts=physical_parts,
            physical_refs_by_row=refs_by_row,
        )


def filter_rows(
    parsed: ParsedSource,
    include_shields: bool = False,
    process_material_keeps: set[str] | None = None,
) -> tuple[list[dict[str, str]], list[list[object]]]:
    rows: list[dict[str, str]] = []
    excluded: list[list[object]] = []
    for row_num, row in zip(parsed.row_numbers, parsed.raw_rows):
        refs = [normalize_ref(r) for r in split_refs(row.get("reference"))]
        reason = exclusion_reason(
            row,
            refs,
            include_shields=include_shields,
            process_material_keeps=process_material_keeps,
        )
        if reason:
            excluded.append([
                row_num,
                ",".join(refs),
                row.get("part_number"),
                row.get("name"),
                row.get("model"),
                row.get("desc"),
                row.get("value"),
                reason,
                _exclusion_reason_kind(row, reason),
            ])
            continue
        rows.append(_processing_row(row, refs))
    return rows, excluded


def load_source(
    path: Path,
    include_shields: bool = False,
    process_material_keeps: set[str] | None = None,
) -> tuple[list[dict[str, str]], list[list[object]]]:
    return filter_rows(
        parse_source(path),
        include_shields=include_shields,
        process_material_keeps=process_material_keeps,
    )


def detect_shield_candidates(source_rows: list[dict[str, str]]) -> list[dict[str, object]]:
    candidates = []
    for row in source_rows:
        refs = sorted({normalize_ref(r) for r in split_refs(row.get("reference"))}, key=natural_key)
        shield_refs = [ref for ref in refs if ref.upper().startswith("SH")]
        if not shield_refs or not str(row.get("part_number") or "").strip():
            continue
        candidates.append(
            {
                "code": row.get("part_number", "").strip(),
                "name": row.get("name", "").strip() or "屏蔽支架",
                "model": row.get("model", "").strip(),
                "desc": row.get("desc", "").strip() or "屏蔽支架",
                "grade": row.get("grade", "").strip(),
                "refs": shield_refs,
                "count": len(shield_refs),
            }
        )
    return candidates


def _clean_candidate_field(value: object) -> str:
    text = str(value or "").strip()
    if not text or text.startswith("{") or "\ufffd" in text:
        return ""
    return text


def _suggest_part_number(row: dict[str, str]) -> str:
    config = classification_config()
    for field in CODE_CANDIDATE_FIELDS:
        if field == "part_number":
            continue
        value = _clean_candidate_field(row.get(field))
        if value and code_shape_matches(value, config):
            return value
    return ""


def _missing_candidate_signature(row: dict[str, str], suggested_code: str) -> tuple[str, ...]:
    if suggested_code:
        return ("suggested", suggested_code.upper())
    return (
        "properties",
        *(
            _clean_candidate_field(row.get(field)).casefold()
            for field in (
                "value",
                "name",
                "model",
                "desc",
                "pcb_footprint",
                "pcb_package",
                "source_package",
                "source_part",
            )
        ),
    )


def detect_missing_part_number_candidates(parsed: ParsedSource) -> list[dict[str, object]]:
    """Group blank-code rows that still carry evidence of a physical material."""
    grouped: "OrderedDict[tuple[str, ...], dict[str, object]]" = OrderedDict()
    for row_num, row in zip(parsed.row_numbers, parsed.raw_rows):
        if str(row.get("part_number") or "").strip() or row.get("_missing_part_number_decision"):
            continue
        refs = sorted({normalize_ref(ref) for ref in split_refs(row.get("reference"))}, key=natural_key)
        value = str(row.get("value") or "").strip()
        if not refs or NC_VALUE_RE.fullmatch(value):
            continue

        suggested_code = _suggest_part_number(row)
        evidence: list[str] = []
        if suggested_code:
            evidence.append("Value/型号疑似物料编码")
        if any(_clean_candidate_field(row.get(field)) for field in ("name", "model", "desc")):
            evidence.append("存在名称、型号或描述")
        if any(
            _clean_candidate_field(row.get(field))
            for field in ("pcb_footprint", "pcb_package", "source_package", "source_part")
        ):
            evidence.append("存在封装或原理图库信息")
        if not evidence:
            continue

        signature = _missing_candidate_signature(row, suggested_code)
        candidate = grouped.setdefault(
            signature,
            {
                "row_numbers": [],
                "refs": set(),
                "suggested_code": suggested_code,
                "name": _clean_candidate_field(row.get("name")),
                "model": _clean_candidate_field(row.get("model")),
                "desc": _clean_candidate_field(row.get("desc")),
                "value": value,
                "pcb_footprint": _clean_candidate_field(row.get("pcb_footprint") or row.get("pcb_package")),
                "source_package": _clean_candidate_field(row.get("source_package")),
                "evidence": [],
            },
        )
        candidate["row_numbers"].append(row_num)
        candidate["refs"].update(refs)
        for item in evidence:
            if item not in candidate["evidence"]:
                candidate["evidence"].append(item)
        for field in ("name", "model", "desc", "value", "pcb_footprint", "source_package"):
            if not candidate[field]:
                candidate[field] = _clean_candidate_field(row.get(field))

    candidates: list[dict[str, object]] = []
    for candidate in grouped.values():
        refs = sorted(candidate["refs"], key=natural_key)
        row_numbers = sorted(candidate["row_numbers"])
        suggested_code = str(candidate["suggested_code"] or "")
        candidates.append(
            {
                **candidate,
                "key": f"rows:{','.join(str(value) for value in row_numbers)}",
                "row_numbers": row_numbers,
                "refs": refs,
                "position_count": len(refs),
                "recommended_action": "keep" if suggested_code else "review",
            }
        )
    return candidates


def apply_missing_part_number_resolutions(
    parsed: ParsedSource,
    resolutions: dict[str, object],
) -> tuple[ParsedSource, dict[str, int]]:
    candidates = detect_missing_part_number_candidates(parsed)
    missing_keys = [str(candidate["key"]) for candidate in candidates if candidate["key"] not in resolutions]
    if missing_keys:
        raise ValueError(f"仍有 {len(missing_keys)} 组空编码物料未确认。")

    rows = [dict(row) for row in parsed.raw_rows]
    row_indexes = {row_num: index for index, row_num in enumerate(parsed.row_numbers)}
    summary = {
        "missing_part_number_candidates": len(candidates),
        "missing_part_number_positions": sum(int(candidate["position_count"]) for candidate in candidates),
        "missing_part_number_kept": 0,
        "missing_part_number_excluded": 0,
    }
    for candidate in candidates:
        key = str(candidate["key"])
        resolution = resolutions.get(key)
        if not isinstance(resolution, dict):
            raise ValueError(f"空编码物料 {','.join(candidate['refs'])} 的确认结果无效。")
        action = str(resolution.get("action") or "").strip().lower()
        if action not in {"keep", "exclude"}:
            raise ValueError(f"请选择空编码物料 {','.join(candidate['refs'])} 是纳入 BOM 还是确认不装。")

        if action == "keep":
            part_number = str(resolution.get("part_number") or "").strip()
            if not part_number:
                raise ValueError(f"空编码物料 {','.join(candidate['refs'])} 选择纳入 BOM 时必须填写子项编码。")
            resolved_fields = {
                field: _clean_candidate_field(resolution.get(field))
                for field in ("name", "model", "desc", "grade", "unit")
            }
            if not any(resolved_fields[field] for field in ("name", "model", "desc")):
                raise ValueError(f"空编码物料 {','.join(candidate['refs'])} 至少需要填写名称、型号或描述之一。")
            summary["missing_part_number_kept"] += 1
        else:
            part_number = ""
            resolved_fields = {}
            summary["missing_part_number_excluded"] += 1

        for row_num in candidate["row_numbers"]:
            row = rows[row_indexes[int(row_num)]]
            row["_missing_part_number_decision"] = action
            if action == "exclude":
                continue
            row["part_number"] = part_number
            for field, value in resolved_fields.items():
                if value:
                    row[field] = value
            if str(row.get("value") or "").strip() == part_number:
                row["value"] = ""

    return ParsedSource(parsed.source_path, rows, list(parsed.row_numbers), parsed.normalized_rows), summary


def detect_process_material_candidates(source_rows: list[dict[str, str]]) -> list[dict[str, object]]:
    candidates: list[dict[str, object]] = []
    for row in source_rows:
        part_number = str(row.get("part_number") or "").strip()
        refs = sorted({normalize_ref(r) for r in split_refs(row.get("reference"))}, key=natural_key)
        value = str(row.get("value") or "").strip()
        keyword = _matches_process_material(row)
        if not part_number or not refs or NC_VALUE_RE.fullmatch(value) or has_shield_refs(refs) or not keyword:
            continue
        candidates.append(
            {
                "key": _process_candidate_key(part_number, refs),
                "part_number": part_number,
                "refs": refs,
                "description": str(row.get("desc") or "").strip(),
                "name": str(row.get("name") or "").strip(),
                "model": str(row.get("model") or "").strip(),
                "matched_keyword": keyword,
            }
        )
    return candidates


def _representative(values: list[str], _field: str) -> str:
    non_empty = [str(value or "").strip() for value in values if str(value or "").strip()]
    if not non_empty:
        return ""
    counts = Counter(non_empty)
    return max(non_empty, key=lambda value: (counts[value], -non_empty.index(value)))


def _row_signature(row: dict[str, str]) -> tuple[str, ...]:
    return tuple(row.get(field, "").strip() for field in CONFLICT_FIELDS)


def _signature_payload(signature: tuple[str, ...]) -> dict[str, str]:
    return dict(zip(CONFLICT_FIELDS, signature))


def _fallback_recommendation(
    variants: list[dict[str, object]],
    signatures: list[tuple[str, ...]],
    reason: str,
) -> dict[str, object]:
    return {
        "confidence": "none",
        "reason": reason,
        "high_confidence": False,
        "manual_choice_required": True,
        "recommended_index": None,
        "recommended_signature": None,
    }


def _cosmetic_key(value: str) -> tuple[str, tuple[str, ...]]:
    text = clean_field_text(value).casefold()
    chunks = [chunk for chunk in re.split(r"[,;；，]+", text) if chunk.strip()]
    if len(chunks) > 1:
        text = "|".join(sorted(re.sub(r"[^\w\u3400-\u9fff]+", "", chunk) for chunk in chunks))
    compact = re.sub(r"[^\w\u3400-\u9fff]+", "", text)
    return compact, tuple(re.findall(r"\d+(?:\.\d+)?", text))


def _cosmetically_equal(signatures: list[tuple[str, ...]]) -> bool:
    return all(
        len({_cosmetic_key(signature[index]) for signature in signatures}) <= 1
        for index in range(len(CONFLICT_FIELDS))
    )


def _manual_conflict_reason(signatures: list[tuple[str, ...]]) -> str:
    values_by_field = {
        field: {signature[index] for signature in signatures if signature[index]}
        for index, field in enumerate(CONFLICT_FIELDS)
    }
    if len(values_by_field["pcb_footprint"]) > 1 or len(values_by_field["pcb_package"]) > 1:
        return "footprint_conflict"
    for field in ("value", "model", "desc"):
        values = values_by_field[field]
        if len(values) > 1 and any(re.search(r"\d", value) for value in values):
            return "numeric_or_version_conflict"
    if len(values_by_field["model"]) > 1 or len(values_by_field["manufacturer"]) > 1:
        return "model_or_manufacturer_conflict"
    if len(values_by_field["grade"]) > 1:
        return "grade_conflict"
    if len(values_by_field["desc"]) > 1:
        return "full_description_conflict"
    return "multiple_complete_candidates"


def _conflict_recommendation(variants: list[dict[str, object]]) -> dict[str, object]:
    signatures = [tuple(str(variant.get(field) or "") for field in CONFLICT_FIELDS) for variant in variants]

    if _cosmetically_equal(signatures):
        index = max(range(len(variants)), key=lambda item: (int(variants[item].get("count") or 0), -item))
        return {
            "confidence": "high",
            "reason": "cosmetic_equivalence",
            "high_confidence": True,
            "manual_choice_required": False,
            "recommended_index": index,
            "recommended_signature": _signature_payload(signatures[index]),
        }

    non_empty_by_field = [
        {signature[index] for signature in signatures if signature[index]}
        for index in range(len(CONFLICT_FIELDS))
    ]
    has_blank = any(not value for signature in signatures for value in signature)
    if has_blank and all(len(values) <= 1 for values in non_empty_by_field):
        target = tuple(next(iter(values), "") for values in non_empty_by_field)
        complete = [index for index, signature in enumerate(signatures) if signature == target]
        if len(complete) == 1:
            index = complete[0]
            return {
                "confidence": "high",
                "reason": "blank_completion",
                "high_confidence": True,
                "manual_choice_required": False,
                "recommended_index": index,
                "recommended_signature": _signature_payload(signatures[index]),
            }
        return _fallback_recommendation(variants, signatures, "complementary_incomplete_candidates")

    dominant: list[int] = []
    for candidate_index, candidate in enumerate(signatures):
        dominates_all = True
        has_strict_prefix = False
        for other_index, other in enumerate(signatures):
            if candidate_index == other_index:
                continue
            for field, candidate_value, other_value in zip(CONFLICT_FIELDS, candidate, other):
                if candidate_value == other_value:
                    continue
                if field not in {"name", "desc"}:
                    dominates_all = False
                    break
                if not candidate_value or not other_value:
                    dominates_all = False
                    break
                if not candidate_value.startswith(other_value):
                    dominates_all = False
                    break
                extension = candidate_value[len(other_value):]
                if extension and extension[0].isdigit():
                    dominates_all = False
                    break
                has_strict_prefix = True
            if not dominates_all:
                break
        if dominates_all and has_strict_prefix:
            dominant.append(candidate_index)
    if len(dominant) == 1:
        index = dominant[0]
        return {
            "confidence": "high",
            "reason": "truncation_prefix_completion",
            "high_confidence": True,
            "manual_choice_required": False,
            "recommended_index": index,
            "recommended_signature": _signature_payload(signatures[index]),
        }

    return _fallback_recommendation(variants, signatures, _manual_conflict_reason(signatures))


def detect_part_conflicts(source_rows: list[dict[str, str]]) -> list[dict[str, object]]:
    by_code: "OrderedDict[str, list[dict[str, str]]]" = OrderedDict()
    for row in source_rows:
        code = row.get("part_number", "").strip()
        if code:
            by_code.setdefault(code, []).append(row)

    conflicts: list[dict[str, object]] = []
    for code, rows in by_code.items():
        variants: "OrderedDict[tuple[str, ...], dict[str, object]]" = OrderedDict()
        for row in rows:
            refs = sorted({normalize_ref(r) for r in split_refs(row.get("reference"))}, key=natural_key)
            signature = _row_signature(row)
            if signature not in variants:
                variants[signature] = {
                    **_signature_payload(signature),
                    "refs": [],
                    "count": 0,
                    "variant_id": stable_fingerprint("bom-material-variant", {
                        "code": code,
                        **_signature_payload(signature),
                    }),
                }
            variants[signature]["refs"].extend(refs)
            variants[signature]["count"] += len(refs) or 1
        if len(variants) > 1:
            variant_list = [
                {
                    **variant,
                    "refs": sorted(set(variant["refs"]), key=natural_key)[:30],
                }
                for variant in variants.values()
            ]
            conflicts.append(
                {
                    "code": code,
                    "total_refs": sum(int(v["count"]) for v in variants.values()),
                    "variants": variant_list,
                    **_conflict_recommendation(variant_list),
                }
            )
    return conflicts


def conflict_summary(source_rows: list[dict[str, str]], limit: int = 50) -> list[dict[str, object]]:
    conflicts = detect_part_conflicts(source_rows)
    return conflicts[:limit]


def _selected_signature(
    code: str,
    source_rows: list[dict[str, str]],
    conflict_choices: dict[str, object] | None,
) -> tuple[str, ...] | None:
    if not conflict_choices or code not in conflict_choices:
        return None
    raw_choice = conflict_choices.get(code)
    if isinstance(raw_choice, Mapping):
        action = str(raw_choice.get("action") or "select_variant")
        if action != "select_variant":
            return None
        raw_choice = raw_choice.get("variant_index")
    if isinstance(raw_choice, bool):
        return None
    if isinstance(raw_choice, int):
        choice = raw_choice
    elif isinstance(raw_choice, str) and raw_choice.isdigit():
        choice = int(raw_choice)
    else:
        return None
    signatures: list[tuple[str, ...]] = []
    for row in source_rows:
        if row.get("part_number", "").strip() != code:
            continue
        signature = _row_signature(row)
        if signature not in signatures:
            signatures.append(signature)
    if 0 <= choice < len(signatures):
        return signatures[choice]
    return None


def unresolved_part_conflicts(
    source_rows: list[dict[str, str]],
    conflict_choices: dict[str, object] | None = None,
) -> list[dict[str, object]]:
    unresolved = []
    for conflict in detect_part_conflicts(source_rows):
        code = str(conflict["code"])
        if conflict.get("high_confidence"):
            continue
        raw_choice = conflict_choices.get(code) if isinstance(conflict_choices, Mapping) else None
        if not _conflict_choice_is_complete(conflict, raw_choice, source_rows):
            unresolved.append(conflict)
    return unresolved


def _variant_index_by_signature(conflict: Mapping[str, object]) -> dict[tuple[str, ...], int]:
    variants = conflict.get("variants")
    if not isinstance(variants, list):
        return {}
    return {
        tuple(str(variant.get(field) or "") for field in CONFLICT_FIELDS): index
        for index, variant in enumerate(variants)
        if isinstance(variant, Mapping)
    }


def _choice_indices(raw: object, key: str) -> list[int]:
    if not isinstance(raw, Mapping):
        return []
    values = raw.get(key)
    if not isinstance(values, list):
        return []
    return [int(value) for value in values if isinstance(value, int) and not isinstance(value, bool)]


def _split_assignments(raw_choice: object) -> dict[int, str]:
    if not isinstance(raw_choice, Mapping) or str(raw_choice.get("action") or "") != "split_refs":
        return {}
    assignments = raw_choice.get("assignments")
    if not isinstance(assignments, list):
        return {}
    mapped: dict[int, str] = {}
    for assignment in assignments:
        if not isinstance(assignment, Mapping):
            continue
        index = assignment.get("variant_index")
        part_number = str(assignment.get("part_number") or "").strip()
        if isinstance(index, int) and not isinstance(index, bool) and part_number:
            mapped[index] = part_number
    return mapped


def _conflict_choice_is_complete(
    conflict: Mapping[str, object],
    raw_choice: object,
    source_rows: list[dict[str, str]],
) -> bool:
    variants = conflict.get("variants")
    variant_count = len(variants) if isinstance(variants, list) else 0
    if variant_count < 2:
        return True
    if _selected_signature(str(conflict.get("code") or ""), source_rows, {str(conflict.get("code") or ""): raw_choice}) is not None:
        return True
    if not isinstance(raw_choice, Mapping):
        return False
    action = str(raw_choice.get("action") or "")
    if action == "split_refs":
        assignments = _split_assignments(raw_choice)
        return set(assignments) == set(range(variant_count)) and len(set(assignments.values())) == variant_count
    if action == "move_non_smt":
        moved = set(_choice_indices(raw_choice, "variant_indices"))
        return bool(moved) and moved.issubset(set(range(variant_count))) and variant_count - len(moved) <= 1
    return False


def apply_conflict_choices(
    parsed: ParsedSource,
    conflict_choices: Mapping[str, object] | None,
) -> ParsedSource:
    choices = conflict_choices if isinstance(conflict_choices, Mapping) else {}
    source_rows, _ = filter_rows(parsed)
    conflicts = {str(item["code"]): item for item in detect_part_conflicts(source_rows)}
    if not conflicts:
        return parsed

    rows: list[dict[str, str]] = []
    row_numbers: list[int] = []
    for row_number, source_row in zip(parsed.row_numbers, parsed.raw_rows):
        row = dict(source_row)
        code = str(row.get("part_number") or "").strip()
        conflict = conflicts.get(code)
        choice = choices.get(code)
        if conflict is None or not isinstance(choice, Mapping):
            rows.append(row)
            row_numbers.append(row_number)
            continue
        variant_index = _variant_index_by_signature(conflict).get(_row_signature(row))
        if variant_index is None:
            rows.append(row)
            row_numbers.append(row_number)
            continue
        action = str(choice.get("action") or "select_variant")
        if action == "select_variant":
            row["_conflict_action"] = "select_variant"
        elif action == "split_refs":
            replacement = _split_assignments(choice).get(variant_index)
            if replacement:
                row["part_number"] = replacement
                touched = set(row.get("_user_touched") or [])
                touched.add("part_number")
                row["_user_touched"] = sorted(touched)
                row["_conflict_action"] = "split_refs"
        elif action == "move_non_smt" and variant_index in set(_choice_indices(choice, "variant_indices")):
            exclusion_kind = str(choice.get("exclusion_kind") or "user_excluded")
            if exclusion_kind not in {"scope_excluded", "user_excluded"}:
                exclusion_kind = "user_excluded"
            row["_placement_action"] = "exclude"
            row["_placement_destination"] = "non_smt"
            row["_placement_exclusion_kind"] = exclusion_kind
            row["_placement_reason_kind"] = exclusion_kind
            row["_placement_reason"] = (
                "不属于当前 PCBA/SMT 范围"
                if exclusion_kind == "scope_excluded"
                else "用户确认移出贴片 BOM"
            )
            row["_conflict_action"] = "move_non_smt"
        rows.append(row)
        row_numbers.append(row_number)
    return replace(parsed, raw_rows=rows, row_numbers=row_numbers)


def build_records(
    parsed: ParsedSource,
    merge_conflicts: bool = False,
    conflict_choices: dict[str, object] | None = None,
    include_shields: bool = False,
    process_material_keeps: set[str] | None = None,
) -> list[dict[str, object]]:
    groups: "OrderedDict[tuple, dict[str, object]]" = OrderedDict()
    source_rows, _ = filter_rows(
        parsed,
        include_shields=include_shields,
        process_material_keeps=process_material_keeps,
    )
    conflicts_by_code = {str(conflict["code"]): conflict for conflict in detect_part_conflicts(source_rows)}
    selected_by_code: dict[str, tuple[str, ...]] = {}
    if merge_conflicts:
        for code, conflict in conflicts_by_code.items():
            selected = _selected_signature(code, source_rows, conflict_choices)
            if selected is None and conflict.get("high_confidence"):
                recommended = conflict.get("recommended_signature")
                if isinstance(recommended, dict):
                    selected = tuple(str(recommended.get(field) or "") for field in CONFLICT_FIELDS)
            if selected is not None:
                selected_by_code[code] = selected
    for row in source_rows:
        refs = sorted({normalize_ref(r) for r in split_refs(row.get("reference"))}, key=natural_key)
        code = row.get("part_number", "").strip()
        signature = _row_signature(row)
        if merge_conflicts and (code not in conflicts_by_code or code in selected_by_code):
            key = (code,)
        else:
            key = (code, *signature)
        if key not in groups:
            groups[key] = {
                "code": code,
                "refs": set(),
                "user_touched": set(),
                **{field: [] for field in (*CONFLICT_FIELDS, *BOM_OPTIONAL_FIELDS)},
            }
        for field in CONFLICT_FIELDS:
            groups[key][field].append(row.get(field, "").strip())
        for field in BOM_OPTIONAL_FIELDS:
            if field not in CONFLICT_FIELDS:
                groups[key][field].append(row.get(field, "").strip())
        groups[key]["refs"].update(refs)
        raw_touched = row.get("_user_touched")
        if isinstance(raw_touched, (list, tuple, set)):
            groups[key]["user_touched"].update(str(field) for field in raw_touched)
    records = []
    for group in groups.values():
        refs = sorted(group["refs"], key=natural_key)
        selected = selected_by_code.get(str(group["code"])) if merge_conflicts else None
        selected_values = _signature_payload(selected) if selected is not None else {}
        record = {
            "code": group["code"],
            "name": selected_values.get("name", "") if selected else _representative(group["name"], "name"),
            "model": selected_values.get("model", "") if selected else _representative(group["model"], "model"),
            "desc": selected_values.get("desc", "") if selected else _representative(group["desc"], "desc"),
            "grade": selected_values.get("grade", "") if selected else _representative(group["grade"], "grade"),
            "refs": refs,
            "qty": len(refs),
            "user_touched": sorted(group["user_touched"]),
        }
        for field in BOM_OPTIONAL_FIELDS:
            if selected is not None and field in CONFLICT_FIELDS:
                record[field] = selected_values.get(field, "")
            else:
                record[field] = _representative(group[field], field)
        records.append(record)
    return records


def _extra_records(extras: list[dict[str, object]] | None) -> list[dict[str, object]]:
    out = []
    for extra in extras or []:
        code = str(extra.get("code") or "").strip()
        if not code:
            continue
        refs = sorted({normalize_ref(r) for r in split_refs(extra.get("refs"))}, key=natural_key)
        qty_raw = str(extra.get("qty") or "").strip()
        qty = int(qty_raw) if qty_raw.isdigit() else (len(refs) if refs else 1)
        out.append({
            "code": code, "name": str(extra.get("name") or "").strip(), "model": str(extra.get("model") or "").strip(),
            "desc": str(extra.get("desc") or "").strip(), "grade": "", "refs": refs, "qty": qty,
            "user_touched": [],
            **{field: FIELD_DEFAULTS.get(field, "") for field in BOM_OPTIONAL_FIELDS},
        })
    return out


def _autosize(ws) -> None:
    for column in ws.columns:
        width = min(max((len(str(cell.value or "")) for cell in column), default=4) + 2, 60)
        ws.column_dimensions[column[0].column_letter].width = width


def _sheet_title(name: str, suffix: str) -> str:
    base = re.sub(r'[\\/*?:\[\]]', "_", name or "BOM")
    return (base + suffix)[:31] or "BOM"


def _safe_filename_component(value: str) -> str:
    text = str(value or "BOM").strip()
    text = re.sub(r'[\\/:*?"<>|\x00-\x1f]+', "_", text)
    text = re.sub(r"\s+", "_", text).strip("._ ")
    return text or "BOM"


def _rec_value(rec: dict[str, object], key: str) -> object:
    value = rec.get(key)
    if value is None or value == "":
        return FIELD_DEFAULTS.get(key, "")
    return value


def _plm_row(rec: dict[str, object], parent_code: str, parent_desc: str) -> list[object]:
    return [
        parent_code,
        parent_desc,
        rec["code"],
        rec["name"],
        rec["model"],
        rec["desc"],
        _rec_value(rec, "unit"),
        rec["qty"],
        ",".join(rec["refs"]),
        _rec_value(rec, "remark"),
        rec["grade"],
        _rec_value(rec, "grade_remark"),
        _rec_value(rec, "alt_group"),
        _rec_value(rec, "alt_strategy"),
        _rec_value(rec, "alt_method"),
        _rec_value(rec, "alt_priority"),
        _rec_value(rec, "issue_method"),
        _rec_value(rec, "mrp"),
        _rec_value(rec, "jump_level"),
    ]


def write_plm(
    path: Path,
    records: list[dict[str, object]],
    parent_code: str,
    parent_desc: str,
    name: str,
    template: Path | None = None,
) -> None:
    """优先复制 PLM 模板（保留合并表头、配色、边框、默认值），把数据套用数据行样式填入。"""
    from copy import copy
    from openpyxl import Workbook
    from openpyxl.styles import Font

    rows_data = [_plm_row(rec, parent_code, parent_desc) for rec in records]

    if template and Path(template).exists():
        with open_bom_workbook(template) as wb:
            ws = wb.worksheets[0]
            # 捕获模板首个数据行（第3行）每列样式，用于新数据行
            styles: dict[int, tuple] = {}
            if ws.max_row >= 3:
                for col in range(1, ws.max_column + 1):
                    cell = ws.cell(3, col)
                    styles[col] = (copy(cell.font), copy(cell.fill), copy(cell.border), copy(cell.alignment), cell.number_format)
                ws.delete_rows(3, ws.max_row - 2)  # 删除模板自带示例数据 + 脚注，保留 1-2 行表头
            for i, values in enumerate(rows_data):
                for col, val in enumerate(values, start=1):
                    cell = ws.cell(3 + i, col, val)
                    style = styles.get(col)
                    if style:
                        cell.font, cell.fill, cell.border, cell.alignment, cell.number_format = (
                            copy(style[0]), copy(style[1]), copy(style[2]), copy(style[3]), style[4],
                        )
            try:
                ws.title = _sheet_title(name, "")
            except Exception:
                pass
            wb.save(path)
        return

    # 无模板时从零生成（降级）
    wb = Workbook()
    try:
        ws = wb.active
        ws.title = _sheet_title(name, "")
        ws.append(PLM_R1)
        ws.append(PLM_HEADERS)
        for values in rows_data:
            ws.append(values)
        for cell in ws[2]:
            cell.font = Font(bold=True)
        _autosize(ws)
        wb.save(path)
    finally:
        wb.close()


def write_oa(path: Path, records: list[dict[str, object]], parent_code: str, parent_desc: str, name: str) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    try:
        ws = wb.active
        ws.title = _sheet_title(name, "_OA")
        ws.append(OA_HEADERS)
        for rec in records:
            ws.append([
                "",
                parent_code,
                parent_desc,
                rec["code"],
                rec["desc"],
                rec["qty"],
                _rec_value(rec, "unit"),
                ",".join(rec["refs"]),
                _rec_value(rec, "remark"),
                _rec_value(rec, "alt_group"),
                _rec_value(rec, "alt_strategy"),
                _rec_value(rec, "alt_method"),
                rec["grade"],
                _rec_value(rec, "issue_method"),
                _rec_value(rec, "mrp"),
                _rec_value(rec, "jump_level"),
            ])
        for cell in ws[1]:
            cell.font = Font(bold=True)
        _autosize(ws)
        wb.save(path)
    finally:
        wb.close()


def write_nc_summary(path: Path, excluded: list[list[object]], name: str, suffix: str = "_NC") -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    try:
        ws = wb.active
        ws.title = _sheet_title(name, suffix)
        ws.append(NC_HEADERS)
        for row in excluded:
            ws.append(row)
        for cell in ws[1]:
            cell.font = Font(bold=True)
        _autosize(ws)
        wb.save(path)
    finally:
        wb.close()


def write_decision_report(path: Path, decisions: list[dict[str, object]], name: str) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font

    workbook = Workbook()
    try:
        sheet = workbook.active
        sheet.title = _sheet_title(name, "_判定记录")
        sheet.append(DECISION_HEADERS)
        for decision in decisions:
            snapshot = decision.get("material_snapshot")
            material = snapshot if isinstance(snapshot, Mapping) else {}
            evidence = decision.get("evidence")
            evidence_items = evidence if isinstance(evidence, list) else []
            evidence_text = "；".join(
                str(item.get("display") or "")
                for item in evidence_items
                if isinstance(item, Mapping) and str(item.get("display") or "")
            )
            refs = decision.get("refs")
            for ref in refs if isinstance(refs, list) else []:
                sheet.append([
                    ref,
                    material.get("part_number", ""),
                    decision.get("identity_status", ""),
                    decision.get("classification_state", ""),
                    decision.get("role", ""),
                    decision.get("subtype", ""),
                    decision.get("destination", ""),
                    decision.get("exclusion_kind", ""),
                    decision.get("rule_id", ""),
                    decision.get("rule_version", ""),
                    decision.get("decision_source", ""),
                    decision.get("decision_fingerprint", ""),
                    evidence_text,
                ])
        for cell in sheet[1]:
            cell.font = Font(bold=True)
        _autosize(sheet)
        workbook.save(path)
    finally:
        workbook.close()


def write_decision_manifest(
    path: Path,
    source_fingerprint: str,
    decisions: list[dict[str, object]],
) -> None:
    payload = {
        "schema_version": 2,
        "rule_version": BOM_RULE_VERSION,
        "source_fingerprint": source_fingerprint,
        "placements": decisions,
    }
    temporary = path.with_suffix(f"{path.suffix}.tmp")
    try:
        temporary.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
        temporary.replace(path)
    finally:
        temporary.unlink(missing_ok=True)


def _finalized_decisions(
    parsed: ParsedSource,
    placement_summary: Mapping[str, object] | None,
    output_records: list[dict[str, object]] | None = None,
    excluded_records: list[list[object]] | None = None,
) -> list[dict[str, object]]:
    raw_records = placement_summary.get("decision_records") if isinstance(placement_summary, Mapping) else None
    records = [dict(item) for item in raw_records if isinstance(item, Mapping)] if isinstance(raw_records, list) else []
    rows_by_ref: dict[str, list[dict[str, str]]] = {}
    for row in parsed.raw_rows:
        for ref in (normalize_ref(value) for value in split_refs(row.get("reference"))):
            rows_by_ref.setdefault(ref, []).append(row)
    records_by_ref: dict[str, list[dict[str, object]]] = {}
    for record in records:
        raw_refs = record.get("refs")
        refs = [normalize_ref(value) for value in raw_refs] if isinstance(raw_refs, list) else []
        for ref in refs:
            records_by_ref.setdefault(ref, []).append(record)

    output_by_ref: dict[str, dict[str, object]] = {}
    for output in output_records or []:
        for raw_ref in output.get("refs") or []:
            ref = normalize_ref(raw_ref)
            previous = output_by_ref.get(ref)
            if previous is not None and str(previous.get("code") or "") != str(output.get("code") or ""):
                raise ValueError(f"BOM 判定记录生成失败：位号 {ref} 同时归入多个输出料号。")
            output_by_ref[ref] = output

    excluded_by_ref: dict[str, list[object]] = {}
    for excluded in excluded_records or []:
        raw_refs = excluded[1] if len(excluded) > 1 else ""
        for raw_ref in split_refs(raw_refs):
            ref = normalize_ref(raw_ref)
            if ref:
                excluded_by_ref[ref] = excluded

    def one_value(ref: str, matched_rows: list[dict[str, str]], field: str, *, default: str = "") -> str:
        values = {
            str(row.get(field) or "").strip()
            for row in matched_rows
            if str(row.get(field) or "").strip()
        }
        if len(values) > 1:
            raise ValueError(f"BOM 判定记录生成失败：物理位号 {ref} 的 {field} 决议不一致。")
        return next(iter(values), default)

    snapshot_fields = (
        "part_number",
        "value",
        "model",
        "name",
        "desc",
        "pcb_footprint",
        "pcb_package",
        "manufacturer",
        "grade",
        "unit",
    )
    finalized: list[dict[str, object]] = []
    for physical_part in sorted(parsed.physical_parts, key=lambda item: natural_key(item.reference)):
        ref = normalize_ref(physical_part.reference)
        matched_rows = rows_by_ref.get(ref, [])
        candidates = records_by_ref.get(ref, [])
        if not matched_rows:
            raise ValueError(f"BOM 判定记录生成失败：物理位号 {ref} 缺少决议。")

        output = output_by_ref.get(ref)
        excluded = excluded_by_ref.get(ref)
        destination = one_value(ref, matched_rows, "_placement_destination")
        role = one_value(ref, matched_rows, "_placement_role", default="unknown")
        exclusion = one_value(ref, matched_rows, "_placement_exclusion_kind")
        subtype = one_value(ref, matched_rows, "_placement_subtype")

        if not candidates:
            legacy_kind = str(excluded[8] or "") if excluded is not None and len(excluded) > 8 else ""
            destination = destination or ("smt" if output is not None else "non_smt" if excluded is not None else "")
            exclusion = exclusion or {
                "system_nc": "nc",
                "nc": "nc",
                "process_default": "process_only",
                "process_only": "process_only",
                "scope_excluded": "scope_excluded",
                "user_excluded": "user_excluded",
            }.get(legacy_kind, "user_excluded" if destination == "non_smt" else "")
            is_shield = ref.startswith("SH")
            if role == "unknown" and is_shield:
                role = "shield"
            if role == "shield" and not subtype:
                subtype = "bracket" if destination == "smt" else "other"
            part_number = str((output or {}).get("code") or matched_rows[0].get("part_number") or "").strip()
            classification_state = (
                "confirmed_nc"
                if exclusion == "nc"
                else "suspected_process"
                if exclusion == "process_only"
                else "confirmed_material"
                if destination == "smt" and part_number
                else "suspected_material"
            )
            fingerprint = stable_fingerprint("bom-placement-legacy-migration", {
                "rule_version": BOM_RULE_VERSION,
                "reference": ref,
                "part_number": part_number,
                "destination": destination,
                "exclusion_kind": exclusion,
                "role": role,
                "subtype": subtype,
            })
            candidates = [{
                "group_id": fingerprint,
                "decision_fingerprint": fingerprint,
                "identity_status": "identity_confirmed" if part_number else "identity_missing",
                "classification_state": classification_state,
                "rule_id": "LEGACY-MIGRATION",
                "rule_version": BOM_RULE_VERSION,
                "decision_source": "user" if any(row.get("_conflict_action") for row in matched_rows) else "rule",
                "evidence": [{
                    "kind": "legacy_migration",
                    "field": "destination",
                    "priority": 0,
                    "display": "旧调用参数已迁移为统一装机决议",
                }],
            }]

        if destination not in {"smt", "non_smt"}:
            raise ValueError(f"BOM 判定记录生成失败：物理位号 {ref} 缺少有效目标区域。")
        if destination == "smt" and exclusion:
            raise ValueError(f"BOM 判定记录生成失败：贴片位号 {ref} 不能带排除类型。")
        if destination == "non_smt" and exclusion not in {"nc", "process_only", "scope_excluded", "user_excluded"}:
            raise ValueError(f"BOM 判定记录生成失败：非贴片位号 {ref} 缺少排除类型。")

        base = dict(candidates[0])
        source_group_ids = sorted({str(item.get("group_id") or "") for item in candidates if str(item.get("group_id") or "")})
        source_fingerprints = sorted({
            str(item.get("decision_fingerprint") or "")
            for item in candidates
            if str(item.get("decision_fingerprint") or "")
        })
        decision_fingerprint = (
            source_fingerprints[0]
            if len(source_fingerprints) == 1
            else stable_fingerprint("bom-placement-physical", {
                "rule_version": BOM_RULE_VERSION,
                "source_fingerprints": source_fingerprints,
            })
        )
        material: dict[str, str] = {}
        for field in snapshot_fields:
            output_field = "code" if field == "part_number" else field
            output_value = str(output.get(output_field) or "").strip() if output is not None else ""
            row_values = sorted({
                str(row.get(field) or "").strip()
                for row in matched_rows
                if str(row.get(field) or "").strip()
            }, key=str.casefold)
            material[field] = output_value or " | ".join(row_values)

        evidence: list[dict[str, object]] = []
        seen_evidence: set[str] = set()
        for candidate in candidates:
            raw_evidence = candidate.get("evidence")
            if not isinstance(raw_evidence, list):
                continue
            for item in raw_evidence:
                if not isinstance(item, Mapping):
                    continue
                marker = json.dumps(dict(item), ensure_ascii=False, sort_keys=True, default=str)
                if marker in seen_evidence:
                    continue
                seen_evidence.add(marker)
                evidence.append(dict(item))

        base.update({
            "group_id": stable_fingerprint("bom-physical-decision-group", [*source_group_ids, ref]),
            "source_group_ids": source_group_ids,
            "refs": [ref],
            "destination": destination,
            "exclusion_kind": exclusion,
            "role": role,
            "subtype": subtype,
            "decision_fingerprint": decision_fingerprint,
            "decision_source": "user" if any(row.get("_conflict_action") for row in matched_rows) else str(base.get("decision_source") or "rule"),
            "evidence": evidence,
            "material_snapshot": material,
        })
        finalized.append(base)
    return finalized


def _validate_bom_output(path: Path, records: list[dict[str, object]], format_name: str) -> None:
    from openpyxl import load_workbook

    expected = Counter(
        (
            str(record.get("code") or ""),
            int(record.get("qty") or 0),
            ",".join(str(ref) for ref in record.get("refs") or []),
        )
        for record in records
    )
    workbook = load_workbook(path, data_only=True, read_only=True)
    try:
        sheet = workbook.active
        if format_name == "plm":
            headers = [str(sheet.cell(2, col).value or "") for col in range(1, len(PLM_HEADERS) + 1)]
            if headers != PLM_HEADERS:
                raise ValueError(f"PLM 输出回读失败：表头不完整（{path.name}）")
            start_row, code_col, qty_col, ref_col = 3, 3, 8, 9
        else:
            headers = [str(sheet.cell(1, col).value or "") for col in range(1, len(OA_HEADERS) + 1)]
            if headers != OA_HEADERS:
                raise ValueError(f"OA 输出回读失败：表头不完整（{path.name}）")
            start_row, code_col, qty_col, ref_col = 2, 4, 6, 8
        actual: Counter[tuple[str, int, str]] = Counter()
        for row_number in range(start_row, sheet.max_row + 1):
            code = str(sheet.cell(row_number, code_col).value or "").strip()
            if not code:
                continue
            quantity = _quantity_as_int(sheet.cell(row_number, qty_col).value)
            refs = ",".join(sorted({normalize_ref(ref) for ref in split_refs(sheet.cell(row_number, ref_col).value)}, key=natural_key))
            actual[(code, int(quantity or 0), refs)] += 1
    finally:
        workbook.close()
    if actual != expected:
        raise ValueError(f"{format_name.upper()} 输出回读失败：料号、数量或位号集合与判定结果不一致。")


def _validate_placement_partition(
    parsed: ParsedSource,
    records: list[dict[str, object]],
    excluded: list[list[object]],
) -> None:
    expected = {part.reference for part in parsed.physical_parts}
    smt_occurrences = [normalize_ref(ref) for record in records for ref in record.get("refs") or [] if normalize_ref(ref) in expected]
    non_smt_occurrences = [normalize_ref(ref) for row in excluded for ref in split_refs(row[1]) if normalize_ref(ref) in expected]
    smt = set(smt_occurrences)
    non_smt = set(non_smt_occurrences)
    if smt.intersection(non_smt):
        raise ValueError("输出校验失败：同一物理位号同时出现在贴片区和非贴片区。")
    if smt.union(non_smt) != expected:
        missing = sorted(expected - smt - non_smt, key=natural_key)
        raise ValueError(f"输出校验失败：{len(missing)} 个物理位号未进入任何目标区域。")
    if any(count > 1 for count in Counter(smt_occurrences).values()):
        raise ValueError("输出校验失败：贴片 BOM 中存在重复物理位号。")


def process(
    parsed: ParsedSource,
    formats: list[str],
    parent_code: str,
    parent_desc: str,
    name: str,
    extras: list[dict[str, object]] | None,
    out_dir: Path,
    stamp: str,
    template: Path | None = None,
    merge_conflicts: bool = False,
    conflict_choices: dict[str, object] | None = None,
    confirm_shields: bool = False,
    process_material_keeps: set[str] | None = None,
    placement_summary: dict[str, object] | None = None,
) -> dict[str, object]:
    if merge_conflicts:
        parsed = apply_conflict_choices(parsed, conflict_choices)
    source_path = parsed.source_path
    name = (name or source_path.stem).strip() or "BOM"
    parent_code = (parent_code or "").strip()
    parent_desc = (parent_desc or name).strip()

    if placement_summary:
        category_counts = placement_summary.get("category_counts")
        state_counts = placement_summary.get("state_counts")
        category_counts = category_counts if isinstance(category_counts, dict) else {}
        state_counts = state_counts if isinstance(state_counts, dict) else {}
        shield_candidates: list[dict[str, object]] = []
        process_material_candidates: list[dict[str, object]] = []
        shield_candidate_count = int(category_counts.get("shield") or 0)
        process_material_candidate_count = int(state_counts.get("suspected_process") or 0)
    else:
        shield_candidates = detect_shield_candidates(parsed.raw_rows)
        process_material_candidates = detect_process_material_candidates(parsed.raw_rows)
        shield_candidate_count = len(shield_candidates)
        process_material_candidate_count = len(process_material_candidates)
    source_rows, excluded = filter_rows(
        parsed,
        include_shields=confirm_shields,
        process_material_keeps=process_material_keeps,
    )
    conflicts = detect_part_conflicts(source_rows)
    unresolved_conflicts = (
        unresolved_part_conflicts(source_rows, conflict_choices)
        if merge_conflicts
        else []
    )
    records = _extra_records(extras) + build_records(
        parsed,
        merge_conflicts=merge_conflicts,
        conflict_choices=conflict_choices,
        include_shields=confirm_shields,
        process_material_keeps=process_material_keeps,
    )
    _validate_placement_partition(parsed, records, excluded)
    finalized_decisions = _finalized_decisions(parsed, placement_summary, records, excluded)
    explicit_nc = [row for row in excluded if str(row[8] or "") in {"nc", "system_nc"}]
    non_smt = [row for row in excluded if str(row[8] or "") not in {"nc", "system_nc"}]

    out_dir.mkdir(parents=True, exist_ok=True)
    base = _safe_filename_component(f"{name}_{stamp}")
    outputs: list[Path] = []
    if "plm" in formats:
        plm = out_dir / f"{base}_PLM_BOM.xlsx"
        write_plm(plm, records, parent_code, parent_desc, name, template)
        _validate_bom_output(plm, records, "plm")
        outputs.append(plm)
    if "oa" in formats:
        oa = out_dir / f"{base}_OA_BOM.xlsx"
        write_oa(oa, records, parent_code, parent_desc, name)
        _validate_bom_output(oa, records, "oa")
        outputs.append(oa)
    nc = out_dir / f"{base}_NC未贴汇总.xlsx"
    write_nc_summary(nc, explicit_nc, name, "_NC")
    non_smt_path = out_dir / f"{base}_非贴片项汇总.xlsx"
    write_nc_summary(non_smt_path, non_smt, name, "_非贴片")
    decision_report = out_dir / f"{base}_BOM判定记录.xlsx"
    write_decision_report(decision_report, finalized_decisions, name)
    decision_manifest = out_dir / f"{base}_BOM决策清单.json"
    write_decision_manifest(decision_manifest, parsed.source_fingerprint, finalized_decisions)
    from app.backend.tools.bom_decisions import load_decision_manifest

    validated_decisions = load_decision_manifest(decision_manifest)
    if not outputs:
        raise ValueError("BOM 处理未生成 PLM 或 OA 成品，无法建立语义模型。")
    semantic_manifest = out_dir / f"{base}_BOM语义模型.json"
    write_semantic_manifest(
        semantic_manifest,
        outputs[0],
        validated_decisions,
    )

    return {
        "outputs": outputs,
        "nc_summary": nc,
        "non_smt_summary": non_smt_path,
        "decision_report": decision_report,
        "decision_manifest": decision_manifest,
        "semantic_manifest": semantic_manifest,
        "decision_records": finalized_decisions,
        "records": records,
        "summary": {
            "name": name,
            "parent_code": parent_code or "(未填)",
            "records": len(records),
            "total_positions": sum(rec["qty"] for rec in records),
            "excluded": len(excluded),
            "explicit_nc": len(explicit_nc),
            "non_smt": len(non_smt),
            "nc_reason_counts": dict(Counter(str(row[8] or "legacy_filter") for row in explicit_nc)),
            "non_smt_reason_counts": dict(Counter(str(row[8] or "legacy_filter") for row in non_smt)),
            "extras": len(_extra_records(extras)),
            "conflicts": len(conflicts),
            "recommended_conflicts": sum(1 for conflict in conflicts if conflict.get("high_confidence")),
            "unresolved_conflicts": len(unresolved_conflicts),
            "merge_conflicts": merge_conflicts,
            "shield_candidates": shield_candidate_count,
            "confirm_shields": confirm_shields,
            "process_material_candidates": process_material_candidate_count,
            "process_material_keeps": len(process_material_keeps or set()),
            "placement_review": dict(placement_summary or {}),
        },
        "conflicts": conflicts,
        "unresolved_conflicts": unresolved_conflicts,
        "shield_candidates": shield_candidates,
        "process_material_candidates": process_material_candidates,
    }
