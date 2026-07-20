from __future__ import annotations

import re
import zipfile
from datetime import datetime
from pathlib import Path
from typing import Iterable

from openpyxl.utils.exceptions import InvalidFileException

from app.backend.parsers.bom_table import (
    FIELD_ALIASES,
    find_header as _find_header,
    normalize_header as _normalize_header,
    read_bom_rows,
    refine_mapping as _refine_bom_mapping,
    split_refs as _split_refs,
)
from app.backend.parsers.refs import natural_key
from app.backend.paths import AppPaths


def _timestamp() -> str:
    return datetime.now().strftime("%Y%m%d_%H%M%S_%f")[:-3]


def _output_dir(params: dict[str, object], root: Path, subdir: str) -> Path:
    raw = params.get("output_dir")
    base = AppPaths(root).outputs_dir.resolve()
    if raw:
        requested = Path(str(raw))
        out = requested if requested.is_absolute() else base / requested
    else:
        out = base / subdir
    out.mkdir(parents=True, exist_ok=True)
    return out


def to_qty(value: object) -> int:
    try:
        return int(float(str(value).strip()))
    except (ValueError, TypeError):
        return 0


_to_qty = to_qty


def qty_matches(value: object, expected: int) -> bool:
    if value is None or str(value).strip() == "":
        return True
    normalized = re.sub(
        r"\s*(?:pcs?|片|个|颗)$",
        "",
        str(value).strip(),
        flags=re.IGNORECASE,
    ).strip()
    try:
        return float(normalized) == expected
    except (ValueError, TypeError):
        return True


def _read_bom_rows(path: Path, require_refs: bool = True) -> list[dict[str, object]]:
    return read_bom_rows(path, require_refs=require_refs)


def _write_table(path: Path, title: str, headers: list[str], rows: list[list[object]]) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    try:
        ws = wb.active
        ws.title = title[:31]
        ws.append(headers)
        for row in rows:
            ws.append(row)
        header_fill = PatternFill("solid", fgColor="D9EAF7")
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
        for column in ws.columns:
            width = min(max(len(str(cell.value or "")) for cell in column) + 2, 60)
            ws.column_dimensions[column[0].column_letter].width = width
        wb.save(path)
    finally:
        wb.close()


MAX_INLINE_ROWS = 5000


def _jsonable(value: object) -> object:
    if value is None:
        return ""
    if isinstance(value, (str, int, float, bool)):
        return value
    return str(value)


def _table(
    headers: list[str],
    rows: list[list[object]],
    status_col: int | None = None,
    diff_pairs: list[list[int]] | None = None,
) -> dict[str, object]:
    """构造可内联返回给前端渲染的表格视图（与导出的 xlsx 同源）。"""
    shown = rows[:MAX_INLINE_ROWS]
    return {
        "headers": list(headers),
        "rows": [[_jsonable(value) for value in row] for row in shown],
        "status_col": status_col,
        "diff_pairs": [list(pair) for pair in (diff_pairs or [])],
        "total_rows": len(rows),
        "shown_rows": len(shown),
    }


def _compare(
    key_label: str,
    left_label: str,
    right_label: str,
    fields: list[str],
    items: list[dict[str, object]],
) -> dict[str, object]:
    """构造左右并排对照视图（BOM1 vs BOM2 / 网表1 vs 网表2）。"""
    shown = items[:MAX_INLINE_ROWS]
    return {
        "key_label": key_label,
        "left_label": left_label,
        "right_label": right_label,
        "fields": list(fields),
        "items": shown,
        "total_rows": len(items),
        "shown_rows": len(shown),
    }


def _write_sheets(path: Path, sheets: list[tuple[str, list[str], list[list[object]]]]) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    try:
        header_fill = PatternFill("solid", fgColor="D9EAF7")
        for index, (title, headers, rows) in enumerate(sheets):
            ws = wb.active if index == 0 else wb.create_sheet()
            ws.title = title[:31]
            ws.append(headers)
            for row in rows:
                ws.append(row)
            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.fill = header_fill
            for column in ws.columns:
                width = min(max(len(str(cell.value or "")) for cell in column) + 2, 60)
                ws.column_dimensions[column[0].column_letter].width = width
        wb.save(path)
    finally:
        wb.close()


def _result(
    tool_id: str,
    outputs: list[Path],
    summary: dict[str, object],
    table: dict[str, object] | None = None,
    compare: dict[str, object] | None = None,
) -> dict[str, object]:
    result: dict[str, object] = {
        "status": "ok",
        "tool": tool_id,
        "outputs": [str(path) for path in outputs],
        "summary": summary,
    }
    if table is not None:
        result["table"] = table
    if compare is not None:
        result["compare"] = compare
    return result


def _error(tool_id: str, message: str) -> dict[str, object]:
    return {"status": "error", "tool": tool_id, "error": message}


def _required_path(params: dict[str, object], key: str, label: str) -> tuple[Path | None, str | None]:
    raw = str(params.get(key, "") or "").strip()
    if not raw:
        return None, f"缺少必填输入：{label}"
    path = Path(raw)
    if not path.exists():
        return None, f"输入不存在：{label}（{path}）"
    return path, None


def _required_file(params: dict[str, object], key: str, label: str) -> tuple[Path | None, str | None]:
    path, error = _required_path(params, key, label)
    if error:
        return None, error
    if path is None or not path.is_file():
        return None, f"输入必须是文件：{label}（{path}）"
    return path, None


def _required_folder(params: dict[str, object], key: str, label: str) -> tuple[Path | None, str | None]:
    raw_value = params.get(key)
    if isinstance(raw_value, list):
        candidates = [Path(str(item)) for item in raw_value if str(item or "").strip()]
        folder = _folder_from_uploaded_netlist_files(candidates)
        if folder is not None:
            return folder, None
        return None, f"输入必须包含同一文件夹下的 pstxnet.dat / pstxprt.dat：{label}"
    path, error = _required_path(params, key, label)
    if error:
        return None, error
    if path is None or not path.is_dir():
        return None, f"输入必须是文件夹：{label}（{path}）"
    return path, None


def _folder_from_uploaded_netlist_files(paths: list[Path]) -> Path | None:
    by_parent: dict[Path, set[str]] = {}
    for path in paths:
        by_parent.setdefault(path.parent, set()).add(path.name.lower())
    for parent, names in by_parent.items():
        if {"pstxnet.dat", "pstxprt.dat"} <= names:
            return parent
    for parent, names in by_parent.items():
        if "pstxnet.dat" in names:
            return parent
    return None
def _read_text_guess(path: Path) -> str:
    for encoding in ("utf-8", "gb18030", "cp936"):
        try:
            return path.read_text(encoding=encoding)
        except UnicodeDecodeError:
            continue
    return path.read_text(encoding="utf-8", errors="ignore")


def _clean_pst_string(value: str) -> str:
    text = value.strip().rstrip(";").rstrip(":").strip()
    if len(text) >= 2 and text[0] == "'" and text[-1] == "'":
        text = text[1:-1]
    return text.strip()


def _natural_join(values: Iterable[str]) -> str:
    return ",".join(sorted(set(values), key=natural_key))


def _parse_node_tokens(line: str) -> tuple[str, str] | None:
    tokens = line.strip().split()
    if len(tokens) >= 3 and tokens[0].upper() == "NODE_NAME":
        return tokens[1], tokens[2]
    return None


def _parse_net_file(folder: Path) -> dict[str, dict[str, list[str]]]:
    path = folder / "pstxnet.dat"
    if not path.exists():
        raise ValueError(f"缺少 pstxnet.dat: {folder}")
    nets: dict[str, dict[str, set[str]]] = {}
    current: str | None = None
    pending_name = False
    for raw in _read_text_guess(path).splitlines():
        line = raw.strip()
        if not line:
            continue
        upper = line.upper()
        if upper == "NET_NAME":
            pending_name = True
            current = None
            continue
        if pending_name:
            name = _clean_pst_string(line)
            if name and not name.startswith("@") and "=" not in name:
                current = name
                nets.setdefault(current, {"refs": set(), "pins": set(), "nodes": set()})
                pending_name = False
            continue
        node = _parse_node_tokens(line)
        if node and current:
            ref, pin = node
            nets[current]["refs"].add(ref)
            nets[current]["pins"].add(pin)
            nets[current]["nodes"].add(f"{ref}.{pin}")
            continue

        # Fallback for simple whitespace netlists: NET N1 R1.1 C1.2
        parts = line.split()
        if not parts:
            continue
        if parts[0].upper() == "NET" and len(parts) >= 2:
            name = parts[1]
            tokens = parts[2:]
        else:
            name = parts[0]
            tokens = parts[1:]
        if not tokens or name in {"FILE_TYPE", "C_SIGNAL"} or "=" in name:
            continue
        entry = nets.setdefault(_clean_pst_string(name), {"refs": set(), "pins": set(), "nodes": set()})
        for token in tokens:
            clean = _clean_pst_string(token)
            ref, _, pin = clean.partition(".")
            if not ref:
                continue
            entry["refs"].add(ref)
            if pin:
                entry["pins"].add(pin)
                entry["nodes"].add(f"{ref}.{pin}")
            else:
                entry["nodes"].add(ref)
    return {
        name: {
            "refs": sorted(data["refs"], key=natural_key),
            "pins": sorted(data["pins"], key=natural_key),
            "nodes": sorted(data["nodes"], key=natural_key),
        }
        for name, data in nets.items()
    }


def _parse_part_file(folder: Path) -> dict[str, str]:
    path = folder / "pstxprt.dat"
    if not path.exists():
        raise ValueError(f"缺少 pstxprt.dat: {folder}")
    parts: dict[str, str] = {}
    pending_part = False
    part_re = re.compile(r"^([A-Za-z]+\d+[A-Za-z0-9_-]*)\s+'([^']+)'")
    for raw in _read_text_guess(path).splitlines():
        line = raw.strip()
        if not line:
            continue
        if line.upper() == "PART_NAME":
            pending_part = True
            continue
        match = part_re.match(line)
        if match:
            parts[match.group(1)] = match.group(2).strip()
            pending_part = False
            continue
        if pending_part:
            tokens = line.split(None, 1)
            if len(tokens) >= 2:
                parts[tokens[0]] = _clean_pst_string(tokens[1])
                pending_part = False
    return parts


def _parse_part_file_optional(folder: Path) -> tuple[dict[str, str], str | None]:
    try:
        return _parse_part_file(folder), None
    except ValueError as exc:
        if "pstxprt.dat" in str(exc):
            return {}, f"缺少 pstxprt.dat：{folder}，已跳过器件封装变化检查，仅执行网络节点对比。"
        raise
_CRITICAL_NET_RE = re.compile(r"(GND|VSS|VDD|VCC|POWER|BUCK|SYS_|USB|MIPI|CSI|DSI|DDR|EMMC|CLK|CLOCK|RST|RESET|I2C|SPI|UART)", re.IGNORECASE)
_NC_NET_RE = re.compile(r"(^|[^A-Z0-9])(NC|NOCONNECT|NO_CONNECT|DNP)([^A-Z0-9]|$)", re.IGNORECASE)
_POWER_NET_RE = re.compile(r"(^|[^A-Z0-9])(GND|AGND|DGND|PGND|VSS|VDD|VCC|VBAT|VSYS|SYS|POWER|BUCK|LDO|3V3|1V8|5V)([^A-Z0-9]|$)", re.IGNORECASE)
_TEST_NET_RE = re.compile(r"(TP|TEST|PROBE|FID|MARK|JTAG|SWD|UART_DBG|DEBUG)", re.IGNORECASE)
_MECHANICAL_NET_RE = re.compile(r"(HOLE|MOUNT|MTG|SCREW|SHIELD|CHASSIS|GASKET)", re.IGNORECASE)


def _net_signature(net: dict[str, list[str]]) -> tuple[str, ...]:
    return tuple(sorted(net.get("nodes", []), key=natural_key))


def _is_critical_net(name: str) -> bool:
    return bool(_CRITICAL_NET_RE.search(str(name or "")))
USER_INPUT_EXCEPTIONS = (
    ValueError,
    KeyError,
    FileNotFoundError,
    PermissionError,
    zipfile.BadZipFile,
    InvalidFileException,
)


def _user_error(tool: str, exc: Exception) -> dict[str, object]:
    message = str(exc) or type(exc).__name__
    return {"status": "error", "tool": tool, "error": message, "message": message, "user_message": message, "error_kind": type(exc).__name__}
