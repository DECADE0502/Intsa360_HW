from __future__ import annotations

import re
from dataclasses import dataclass
from pathlib import Path

from app.backend.parsers.board_outline import resolve_board_outline
from app.backend.parsers.cadence_pst import parse_net_file, parse_part_file
from app.backend.parsers.refs import natural_key
from app.backend.parsers.xy import parse_xy_file
from app.backend.tools import bom_process
from app.backend.tools.common import (
    USER_INPUT_EXCEPTIONS,
    _output_dir,
    _read_bom_rows,
    _required_file,
    _required_folder,
    _timestamp,
    _user_error,
)
from app.backend.tools.smt_package import _is_high_risk_package, _package_matches


_SANITY_SEVERITY_ORDER = {"high": 0, "medium": 1, "low": 2}
_FAI_HEADERS = [
    "位号",
    "面",
    "X(mm)",
    "Y(mm)",
    "封装",
    "应贴料号",
    "应贴型号",
    "应贴描述",
    "优选等级",
    "QC",
    "备注",
]


@dataclass(frozen=True)
class NcEvidence:
    confirmed_refs: set[str]
    candidate_refs: set[str]
    unverified_refs: set[str]
    conflict_refs: set[str]
    inference_mode: str
    explicit_summary_used: bool


def _infer_nc_evidence(
    *,
    xy_refs: set[str],
    bom_refs: set[str],
    netlist_refs: set[str] | None,
    explicit_nc_refs: set[str],
    explicit_summary_used: bool,
) -> NcEvidence:
    missing_from_bom = xy_refs - bom_refs
    conflicts = explicit_nc_refs & bom_refs
    if netlist_refs is None:
        confirmed = missing_from_bom & explicit_nc_refs
        return NcEvidence(
            confirmed_refs=confirmed,
            candidate_refs=missing_from_bom - explicit_nc_refs,
            unverified_refs=set(),
            conflict_refs=conflicts,
            inference_mode="without_netlist",
            explicit_summary_used=explicit_summary_used,
        )

    confirmed = missing_from_bom & (netlist_refs | explicit_nc_refs)
    return NcEvidence(
        confirmed_refs=confirmed,
        candidate_refs=set(),
        unverified_refs=missing_from_bom - netlist_refs - explicit_nc_refs,
        conflict_refs=conflicts,
        inference_mode="with_netlist",
        explicit_summary_used=explicit_summary_used,
    )


def _find_xy_file(folder: Path) -> Path:
    for path in folder.iterdir():
        if path.is_file() and path.name.casefold() == "xy.txt":
            return path
    raise ValueError("SMT 文件夹缺少 XY.txt")


def _find_nc_summary(processed_bom: Path) -> Path | None:
    candidates = [
        path
        for path in processed_bom.parent.glob("*.xlsx")
        if path != processed_bom and "nc未贴" in path.name.casefold()
    ]
    return sorted(candidates, key=lambda path: path.name.casefold())[0] if candidates else None


def _bom_by_ref(processed_bom: Path) -> tuple[dict[str, dict[str, object]], set[str]]:
    rows = _read_bom_rows(processed_bom)
    by_ref = {
        str(ref).upper(): row
        for row in rows
        for ref in row.get("refs", [])
        if str(ref).strip()
    }
    nc_refs: set[str] = set()
    nc_summary = _find_nc_summary(processed_bom)
    if nc_summary:
        parsed = bom_process.parse_source(nc_summary)
        for row in parsed.raw_rows:
            nc_refs.update(ref.upper() for ref in bom_process.split_refs(row.get("reference")))
    return by_ref, nc_refs


def _component_payload(component: object, bom_row: dict[str, object] | None, nc_refs: set[str]) -> dict[str, object]:
    ref = str(component.ref)
    row = bom_row or {}
    status = "nc" if ref.upper() in nc_refs else ("installed" if bom_row else "missing_bom")
    description = str(row.get("description") or "")
    model = str(row.get("model") or "")
    footprint = str(component.footprint)
    return {
        "ref": ref,
        "x_mm": float(component.x_mm),
        "y_mm": float(component.y_mm),
        "rotation": int(component.rotation),
        "side": str(component.side),
        "footprint": footprint,
        "part_number": str(row.get("part_number") or ""),
        "description": description,
        "model": model,
        "grade": str(row.get("grade") or ""),
        "status": status,
        "high_risk": _is_high_risk_package(footprint, description, model),
    }


def _component_value(component: object, field: str) -> object:
    if isinstance(component, dict):
        return component.get(field, "")
    return getattr(component, field, "")


def _bom_footprint_text(row: dict[str, object]) -> str:
    return " ".join(
        str(row.get(field) or "").strip()
        for field in ("package", "description", "name", "model")
        if str(row.get(field) or "").strip()
    )


def _package_compatible(left: str, right: str) -> tuple[bool, str]:
    normalized_left = "".join(char for char in left.casefold() if char not in "-_")
    normalized_right = "".join(char for char in right.casefold() if char not in "-_")
    if normalized_left and normalized_left == normalized_right:
        return True, "封装规范化后完全一致"
    return _package_matches(left, right)


def _sort_sanity_items(items: list[dict[str, object]]) -> list[dict[str, object]]:
    return sorted(
        items,
        key=lambda item: (
            _SANITY_SEVERITY_ORDER.get(str(item.get("severity") or "low"), 3),
            natural_key(str(item.get("ref") or "")),
        ),
    )


def _compute_sanity(
    components: list[object],
    bom_rows: list[dict[str, object]],
    netlist_parts: dict[str, str],
) -> dict[str, list[dict[str, object]]]:
    component_by_ref = {
        str(_component_value(component, "ref")).strip().upper(): component
        for component in components
        if str(_component_value(component, "ref")).strip()
    }
    bom_by_ref = {
        str(ref).strip().upper(): row
        for row in bom_rows
        for ref in row.get("refs", [])
        if str(ref).strip()
    }
    netlist_by_ref = {
        str(ref).strip().upper(): str(package or "").strip()
        for ref, package in netlist_parts.items()
        if str(ref).strip()
    }
    xy_refs = set(component_by_ref)
    bom_refs = set(bom_by_ref)
    netlist_refs = set(netlist_by_ref)

    missing_layout = [
        {
            "ref": ref,
            "note": f"{ref} 在 BOM 或网表中存在，但 XY 布局中没有坐标，可能漏放。",
            "severity": "high",
        }
        for ref in (bom_refs | netlist_refs) - xy_refs
    ]
    missing_bom = [
        {
            "ref": ref,
            "note": f"{ref} 在 {'XY 布局' if ref in xy_refs else '网表'}中存在，但处理后 BOM 中没有记录。",
            "severity": "high" if ref in xy_refs else "medium",
        }
        for ref in (xy_refs | netlist_refs) - bom_refs
    ]
    missing_netlist = [
        {
            "ref": ref,
            "note": f"{ref} 在 {'XY 布局' if ref in xy_refs else 'BOM'}中存在，但 pstxprt.dat 中没有记录。",
            "severity": "medium" if ref in xy_refs else "low",
        }
        for ref in (xy_refs | bom_refs) - netlist_refs
    ]

    footprint_conflicts: list[dict[str, object]] = []
    for ref in sorted(xy_refs & (bom_refs | netlist_refs), key=natural_key):
        xy_footprint = str(_component_value(component_by_ref[ref], "footprint") or "").strip()
        netlist_footprint = netlist_by_ref.get(ref, "")
        bom_footprint = _bom_footprint_text(bom_by_ref[ref]) if ref in bom_by_ref else ""
        notes: list[str] = []
        if xy_footprint and netlist_footprint:
            matches, detail = _package_compatible(xy_footprint, netlist_footprint)
            if not matches:
                notes.append(f"XY 与网表封装不一致：{detail}")
        if xy_footprint and bom_footprint:
            matches, detail = _package_compatible(xy_footprint, bom_footprint)
            if not matches:
                notes.append(f"XY 与 BOM 封装信息不一致：{detail}")
        if notes:
            footprint_conflicts.append(
                {
                    "ref": ref,
                    "xy_footprint": xy_footprint,
                    "netlist_footprint": netlist_footprint,
                    "bom_footprint": bom_footprint,
                    "note": "；".join(notes),
                }
            )

    return {
        "missing_layout": _sort_sanity_items(missing_layout),
        "missing_bom": _sort_sanity_items(missing_bom),
        "missing_netlist": _sort_sanity_items(missing_netlist),
        "footprint_conflicts": footprint_conflicts,
    }


def _fai_sort_key(component: object) -> tuple[object, ...]:
    side = str(_component_value(component, "side") or "top").casefold()
    x_mm = float(_component_value(component, "x_mm") or 0.0)
    y_mm = float(_component_value(component, "y_mm") or 0.0)
    y_band = int(y_mm // 2.54)
    return (0 if side == "top" else 1, -y_band, x_mm, natural_key(str(_component_value(component, "ref"))))


def _build_fai_table(components: list[object]) -> dict[str, object]:
    rows: list[list[object]] = []
    for component in sorted(components, key=_fai_sort_key):
        side = str(_component_value(component, "side") or "top").casefold()
        status = str(_component_value(component, "status") or "")
        part_number = str(_component_value(component, "part_number") or "").strip()
        grade = str(_component_value(component, "grade") or "").strip()
        notes: list[str] = []
        if grade not in {"优选", "正常"}:
            notes.append("⚠ 等级")
        if status == "missing_bom" or not part_number:
            part_number = "⚠ BOM 缺料号"
        rows.append(
            [
                str(_component_value(component, "ref") or ""),
                "正面" if side == "top" else "背面",
                float(_component_value(component, "x_mm") or 0.0),
                float(_component_value(component, "y_mm") or 0.0),
                str(_component_value(component, "footprint") or ""),
                part_number,
                str(_component_value(component, "model") or ""),
                str(_component_value(component, "description") or ""),
                grade,
                "",
                "；".join(notes),
            ]
        )
    return {"headers": list(_FAI_HEADERS), "rows": rows}


def _write_fai_xlsx(output_dir: Path, name: str, stamp: str, table: dict[str, object]) -> Path:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    output_dir.mkdir(parents=True, exist_ok=True)
    safe_name = re.sub(r'[<>:"/\\|?*]+', "_", str(name or "SMT")).strip(" ._") or "SMT"
    output = output_dir / f"首件核对表_{safe_name}_{stamp}.xlsx"
    workbook = Workbook()
    try:
        sheet = workbook.active
        sheet.title = "首件核对表"
        headers = list(table.get("headers") or [])
        rows = list(table.get("rows") or [])
        sheet.append(headers)
        for row in rows:
            sheet.append(list(row))

        header_fill = PatternFill("solid", fgColor="D9EAF7")
        warning_fill = PatternFill("solid", fgColor="FFF7E6")
        for cell in sheet[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for row_index in range(2, sheet.max_row + 1):
            part_number = str(sheet.cell(row=row_index, column=6).value or "")
            note = str(sheet.cell(row=row_index, column=11).value or "")
            if part_number.startswith("⚠") or note:
                for cell in sheet[row_index]:
                    cell.fill = warning_fill

        widths = [14, 8, 12, 12, 18, 20, 20, 40, 14, 10, 18]
        for index, width in enumerate(widths, start=1):
            sheet.column_dimensions[sheet.cell(row=1, column=index).column_letter].width = width
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        sheet.print_title_rows = "1:1"
        sheet.page_setup.paperSize = sheet.PAPERSIZE_A3
        sheet.page_setup.orientation = "landscape"
        sheet.page_setup.fitToWidth = 1
        sheet.page_setup.fitToHeight = 0
        sheet.sheet_properties.pageSetUpPr.fitToPage = True
        workbook.save(output)
    finally:
        workbook.close()
    return output


def _run_smt_layout_impl(root: Path, params: dict[str, object]) -> dict[str, object]:
    smt_folder, error = _required_folder(params, "smt_folder", "SMT 资料文件夹")
    if error:
        raise ValueError(error)
    processed_bom, error = _required_file(params, "processed_bom", "处理后 BOM")
    if error:
        raise ValueError(error)
    assert smt_folder is not None and processed_bom is not None

    netlist_folder: Path | None = None
    if str(params.get("netlist_folder") or "").strip() or isinstance(params.get("netlist_folder"), list):
        netlist_folder, error = _required_folder(params, "netlist_folder", "Cadence 网表文件夹")
        if error:
            raise ValueError(error)

    _, parsed_components = parse_xy_file(_find_xy_file(smt_folder))
    outline_bbox = params.get("outline_bbox_mm")
    if outline_bbox is not None and not isinstance(outline_bbox, (list, tuple)):
        raise ValueError("outline_bbox_mm 必须是四个毫米坐标")
    outline = resolve_board_outline(
        smt_folder,
        outline_bbox_mm=outline_bbox,
        outline_dxf_layer=str(params.get("outline_dxf_layer") or "").strip() or None,
    )
    bom_by_ref, nc_refs = _bom_by_ref(processed_bom)
    components = [
        _component_payload(component, bom_by_ref.get(component.ref.upper()), nc_refs)
        for component in parsed_components
    ]
    components.sort(key=lambda item: natural_key(str(item["ref"])))

    if netlist_folder is None:
        sanity: dict[str, object] = {"status": "skipped_no_netlist"}
    else:
        parse_net_file(netlist_folder)
        netlist_parts = parse_part_file(netlist_folder)
        sanity = _compute_sanity(components, list(bom_by_ref.values()), netlist_parts)

    fai_table = _build_fai_table(components)
    fai_name = str(params.get("name") or processed_bom.stem)
    fai_output = _write_fai_xlsx(_output_dir(params, root, "smt"), fai_name, _timestamp(), fai_table)
    return {
        "status": "ok",
        "tool": "smt_layout",
        "outputs": [str(fai_output)],
        "board": {
            "outline_rings": outline.rings,
            "bbox_mm": outline.bbox,
            "source": outline.source,
        },
        "components": components,
        "nc_summary": {"total": len(nc_refs), "refs": sorted(nc_refs, key=natural_key)},
        "sanity": sanity,
        "fai_table": fai_table,
        "summary": {
            "total_components": len(components),
            "top_count": sum(item["side"] == "top" for item in components),
            "bottom_count": sum(item["side"] == "bottom" for item in components),
            "nc_count": sum(item["status"] == "nc" for item in components),
            "high_risk_count": sum(bool(item["high_risk"]) for item in components),
        },
    }


def run_smt_layout(root: Path, params: dict[str, object]) -> dict[str, object]:
    try:
        return _run_smt_layout_impl(root, params)
    except USER_INPUT_EXCEPTIONS as exc:
        return _user_error("smt_layout", exc)
