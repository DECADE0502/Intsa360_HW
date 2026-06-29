from __future__ import annotations

import re
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Iterable

from app.backend.tool_registry import Tool

# openpyxl 仅在实际运行工具时才需要，延迟到使用处导入，
# 让服务启动 / build_registry 免去约 1s 的导入开销，加快冷启动。


REF_SPLIT_RE = re.compile(r"[,;\s]+")
FIELD_ALIASES = {
    "reference": ["位号", "位置号", "Reference", "Ref", "RefDes", "Designator", "BOM1位号", "BOM2位号"],
    "part_number": ["编号", "料号", "物料编码", "子项编码", "编码子", "编码（子）", "编码(子)", "Part Number", "PN"],
    "description": ["描述", "物料描述", "子项描述", "描述子", "描述（子）", "描述(子)", "器件描述", "器件描述（新整理）", "Description"],
    "quantity": ["数量", "用量", "Qty", "Quantity"],
    "name": ["名称", "物料名称", "Part Type"],
    "package": ["封装名", "PCB封装", "PCB Footprint", "Package", "Footprint"],
    "value": ["Value", "值"],
    "model": ["型号", "规格型号", "规格", "Model", "MPN"],
    "grade": ["物料优选等级", "优选等级", "物料等级", "等级"],
}


def _timestamp() -> str:
    return datetime.now().strftime("%Y%m%d_%H%M%S_%f")[:-3]


def _output_dir(params: dict[str, object], root: Path, subdir: str) -> Path:
    raw = params.get("output_dir")
    out = Path(str(raw)) if raw else root / "data" / "outputs" / subdir
    out.mkdir(parents=True, exist_ok=True)
    return out


def _normalize_header(value: object) -> str:
    text = str(value or "").strip()
    text = re.sub(r"##.*$", "", text)
    text = text.replace("*", "")
    if text.startswith("{") and text.endswith("}"):
        text = text[1:-1]
    text = text.replace(" ", "")
    return text


def _find_header(ws, required: Iterable[str], scan_rows: int = 30) -> tuple[int, dict[str, int]]:
    required = list(required)
    best_row = 1
    best_map: dict[str, int] = {}
    for row in range(1, min(ws.max_row, scan_rows) + 1):
        values = [str(ws.cell(row, col).value or "").strip() for col in range(1, ws.max_column + 1)]
        normalized_values = [_normalize_header(value) for value in values]
        mapping: dict[str, int] = {}
        for canonical, aliases in FIELD_ALIASES.items():
            normalized_aliases = {_normalize_header(alias) for alias in aliases}
            for idx, value in enumerate(normalized_values, start=1):
                if value in normalized_aliases:
                    mapping[canonical] = idx
                    break
        score = sum(1 for key in required if key in mapping)
        if score > sum(1 for key in required if key in best_map):
            best_row = row
            best_map = mapping
    missing = [key for key in required if key not in best_map]
    if missing:
        raise ValueError(f"表头识别失败，缺少字段: {', '.join(missing)}")
    return best_row, best_map


def _choose_best_column(
    normalized_values: list[str],
    aliases: Iterable[str],
    anchor_col: int | None = None,
    prefer_after_anchor: bool = False,
) -> int | None:
    normalized_aliases = {_normalize_header(alias) for alias in aliases}
    matches = [idx for idx, value in enumerate(normalized_values, start=1) if value in normalized_aliases]
    if not matches:
        return None
    if anchor_col is not None and prefer_after_anchor:
        after_anchor = [idx for idx in matches if idx > anchor_col]
        if after_anchor:
            return after_anchor[0]
    return matches[0]


def _refine_bom_mapping(ws, header_row: int, mapping: dict[str, int]) -> dict[str, int]:
    values = [str(ws.cell(header_row, col).value or "").strip() for col in range(1, ws.max_column + 1)]
    normalized_values = [_normalize_header(value) for value in values]
    refined = dict(mapping)

    part_col = _choose_best_column(normalized_values, FIELD_ALIASES["part_number"])
    if part_col:
        refined["part_number"] = part_col

    for key in ["description", "quantity", "name", "package", "value", "model", "grade", "reference"]:
        col = _choose_best_column(
            normalized_values,
            FIELD_ALIASES[key],
            anchor_col=refined.get("part_number"),
            prefer_after_anchor=key in {"description", "quantity", "name", "package", "value"},
        )
        if col:
            refined[key] = col
    return refined


def _split_refs(value: object) -> list[str]:
    return [part for part in REF_SPLIT_RE.split(str(value or "").strip()) if part]


_NAT_RE = re.compile(r"(\d+)")


def _natural_key(ref: str) -> list[object]:
    return [int(token) if token.isdigit() else token.lower() for token in _NAT_RE.split(ref)]


def _to_qty(value: object) -> int:
    try:
        return int(float(str(value).strip()))
    except (ValueError, TypeError):
        return 0


def _read_bom_rows(path: Path, require_refs: bool = True) -> list[dict[str, object]]:
    from openpyxl import load_workbook

    wb = load_workbook(path, data_only=True)
    ws = wb.active
    header_row, mapping = _find_header(ws, ["reference", "part_number", "description", "quantity"])
    mapping = _refine_bom_mapping(ws, header_row, mapping)
    rows: list[dict[str, object]] = []
    for row in range(header_row + 1, ws.max_row + 1):
        refs = _split_refs(ws.cell(row, mapping["reference"]).value)
        part_number = str(ws.cell(row, mapping["part_number"]).value or "").strip()
        if require_refs and not refs:
            continue
        if not refs and not part_number:
            continue
        item = {
            "source_row": row,
            "reference": ",".join(refs),
            "refs": refs,
            "part_number": part_number,
            "model": str(ws.cell(row, mapping["model"]).value or "").strip() if "model" in mapping else "",
            "grade": str(ws.cell(row, mapping["grade"]).value or "").strip() if "grade" in mapping else "",
            "description": str(ws.cell(row, mapping["description"]).value or "").strip(),
            "quantity": ws.cell(row, mapping["quantity"]).value,
            "name": str(ws.cell(row, mapping.get("name", mapping["description"])).value or "").strip(),
            "package": str(ws.cell(row, mapping.get("package", mapping["description"])).value or "").strip(),
            "value": str(ws.cell(row, mapping.get("value", mapping["description"])).value or "").strip(),
        }
        rows.append(item)
    return rows


def _write_table(path: Path, title: str, headers: list[str], rows: list[list[object]]) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]
    ws.append(headers)
    for row in rows:
        ws.append(row)
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
    for column in ws.columns:
        width = min(max(len(str(cell.value or "")) for cell in column) + 2, 60)
        ws.column_dimensions[column[0].column_letter].width = width
    wb.save(path)


MAX_INLINE_ROWS = 5000


def _jsonable(value: object) -> object:
    if value is None:
        return ""
    if isinstance(value, (str, int, float, bool)):
        return value
    return str(value)


def _table(
    headers: list[str],
    rows: list[list[object]],
    status_col: int | None = None,
    diff_pairs: list[list[int]] | None = None,
) -> dict[str, object]:
    """构造可内联返回给前端渲染的表格视图（与导出的 xlsx 同源）。"""
    shown = rows[:MAX_INLINE_ROWS]
    return {
        "headers": list(headers),
        "rows": [[_jsonable(value) for value in row] for row in shown],
        "status_col": status_col,
        "diff_pairs": [list(pair) for pair in (diff_pairs or [])],
        "total_rows": len(rows),
        "shown_rows": len(shown),
    }


def _compare(
    key_label: str,
    left_label: str,
    right_label: str,
    fields: list[str],
    items: list[dict[str, object]],
) -> dict[str, object]:
    """构造左右并排对照视图（BOM1 vs BOM2 / 网表1 vs 网表2）。"""
    shown = items[:MAX_INLINE_ROWS]
    return {
        "key_label": key_label,
        "left_label": left_label,
        "right_label": right_label,
        "fields": list(fields),
        "items": shown,
        "total_rows": len(items),
        "shown_rows": len(shown),
    }


def _write_sheets(path: Path, sheets: list[tuple[str, list[str], list[list[object]]]]) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for index, (title, headers, rows) in enumerate(sheets):
        ws = wb.active if index == 0 else wb.create_sheet()
        ws.title = title[:31]
        ws.append(headers)
        for row in rows:
            ws.append(row)
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
        for column in ws.columns:
            width = min(max(len(str(cell.value or "")) for cell in column) + 2, 60)
            ws.column_dimensions[column[0].column_letter].width = width
    wb.save(path)


def _result(
    tool_id: str,
    outputs: list[Path],
    summary: dict[str, object],
    table: dict[str, object] | None = None,
    compare: dict[str, object] | None = None,
) -> dict[str, object]:
    result: dict[str, object] = {
        "status": "ok",
        "tool": tool_id,
        "outputs": [str(path) for path in outputs],
        "summary": summary,
    }
    if table is not None:
        result["table"] = table
    if compare is not None:
        result["compare"] = compare
    return result


def _error(tool_id: str, message: str) -> dict[str, object]:
    return {"status": "error", "tool": tool_id, "error": message}


def _required_path(params: dict[str, object], key: str, label: str) -> tuple[Path | None, str | None]:
    raw = str(params.get(key, "") or "").strip()
    if not raw:
        return None, f"缺少必填输入：{label}"
    path = Path(raw)
    if not path.exists():
        return None, f"输入不存在：{label}（{path}）"
    return path, None


def _required_file(params: dict[str, object], key: str, label: str) -> tuple[Path | None, str | None]:
    path, error = _required_path(params, key, label)
    if error:
        return None, error
    if path is None or not path.is_file():
        return None, f"输入必须是文件：{label}（{path}）"
    return path, None


def _required_folder(params: dict[str, object], key: str, label: str) -> tuple[Path | None, str | None]:
    raw_value = params.get(key)
    if isinstance(raw_value, list):
        candidates = [Path(str(item)) for item in raw_value if str(item or "").strip()]
        folder = _folder_from_uploaded_netlist_files(candidates)
        if folder is not None:
            return folder, None
        return None, f"输入必须包含同一文件夹下的 pstxnet.dat / pstxprt.dat：{label}"
    path, error = _required_path(params, key, label)
    if error:
        return None, error
    if path is None or not path.is_dir():
        return None, f"输入必须是文件夹：{label}（{path}）"
    return path, None


def _folder_from_uploaded_netlist_files(paths: list[Path]) -> Path | None:
    by_parent: dict[Path, set[str]] = {}
    for path in paths:
        by_parent.setdefault(path.parent, set()).add(path.name.lower())
    for parent, names in by_parent.items():
        if {"pstxnet.dat", "pstxprt.dat"} <= names:
            return parent
    for parent, names in by_parent.items():
        if "pstxnet.dat" in names:
            return parent
    return None


def _index_by_ref(rows: list[dict[str, object]]) -> dict[str, dict[str, object]]:
    """位号 -> 该位号上的元件（编号/型号/描述）。位号是板上物理位置，是对比的主轴。"""
    by_ref: dict[str, dict[str, object]] = {}
    for row in rows:
        for ref in row.get("refs") or []:
            by_ref[ref] = {
                "编号": row.get("part_number", ""),
                "型号": row.get("model", ""),
                "描述": row.get("description", ""),
            }
    return by_ref


def _part_usage(rows: list[dict[str, object]]) -> dict[str, dict[str, object]]:
    """按编号统计用量：有位号者用量=位号数；无位号的整板/父项用量取数量列。"""
    parts: dict[str, dict[str, object]] = {}
    for row in rows:
        part_number = str(row.get("part_number") or "").strip()
        if not part_number:
            continue
        part = parts.setdefault(part_number, {"refs": set(), "model": "", "desc": "", "qty_field": 0, "has_ref": False})
        refs = row.get("refs") or []
        if refs:
            part["refs"].update(refs)
            part["has_ref"] = True
        part["qty_field"] += _to_qty(row.get("quantity"))
        if not part["model"] and row.get("model"):
            part["model"] = row["model"]
        if not part["desc"] and row.get("description"):
            part["desc"] = row["description"]
    return parts


def _usage_count(part: dict[str, object]) -> int:
    return len(part["refs"]) if part["has_ref"] else part["qty_field"]


def _annotate_origin(
    rows: list[dict[str, object]],
    ref_other: dict[str, dict[str, object]],
    parts_other: set[str],
    side: str,
) -> list[dict[str, object]]:
    """把另一份 BOM 的差异标注到“本份原始行”上：状态 + 该行需高亮的位号。"""
    annotated: list[dict[str, object]] = []
    for row in rows:
        part_number = str(row.get("part_number") or "")
        refs = row.get("refs") or []
        # 区分：换料（位号仍在，但换成别的料号）vs 位号在对面缺失
        swapped = [r for r in refs if r in ref_other and str(ref_other[r].get("编号", "")) != part_number]
        gone = [r for r in refs if r not in ref_other]
        changed = swapped + gone
        if not refs:
            status = ("removed" if side == "left" else "added") if part_number and part_number not in parts_other else "same"
        elif swapped:
            status = "swap"
        elif gone and len(gone) == len(refs) and part_number not in parts_other:
            status = "removed" if side == "left" else "added"
        elif gone:
            status = "changed"
        else:
            status = "same"
        annotated.append(
            {
                "cells": [
                    part_number,
                    row.get("model", ""),
                    row.get("description", ""),
                    _jsonable(row.get("quantity")),
                    row.get("reference", ""),
                ],
                "status": status,
                "changed_refs": changed,
            }
        )
    return annotated


_NC_RE = re.compile(r"(^|[,/\s（(])NC([,/\s）)]|$)", re.IGNORECASE)
_MECH_KW = ["螺丝", "螺钉", "螺母", "垫片", "华司", "铜柱", "支柱", "定位孔", "安装孔", "MOUNTINGHOLE", "散热片", "导热垫"]
_TP_PREFIX = ("TP", "JP", "Z_TP", "FID", "MK", "MH")
_TP_KW = ["测试点", "跳线", "FIDUCIAL", "基准", "拼板", "工艺边", "MARK点"]
_VERSION_SENSITIVE_RE = re.compile(r"\b(E?MMC|LP?DDR\d*[A-Z0-9]*)\b|DDR", re.IGNORECASE)


def _looks_like_pcb(row: dict[str, object]) -> bool:
    """裸板特征：印制板/覆铜板/PCB/HDI/任意阶，或“N层”层数描述。"""
    desc = str(row.get("description", ""))
    blob = f"{row.get('part_number','')} {row.get('model','')} {desc}".upper()
    if any(k in blob for k in ["PCB", "HDI"]) or any(k in desc for k in ["印制板", "覆铜板", "任意阶"]):
        return True
    return bool(re.search(r"\d+\s*层", desc))


def _looks_like_shield_bracket(row: dict[str, object]) -> bool:
    refs = [str(ref).upper() for ref in row.get("refs") or []]
    text = f"{row.get('part_number','')} {row.get('model','')} {row.get('description','')} {row.get('name','')}".upper()
    if any(ref.startswith("SH") for ref in refs):
        return True
    return "屏蔽支架" in text or "SHIELD BRACKET" in text


def _version_sensitive_parts(rows: list[dict[str, object]]) -> list[dict[str, object]]:
    out = []
    for row in rows:
        text = f"{row.get('part_number','')} {row.get('model','')} {row.get('description','')} {row.get('name','')}"
        if _VERSION_SENSITIVE_RE.search(text):
            out.append(row)
    return out


# 位号前缀 -> 期望器件类型（占位换料检查：天线匹配网络常见 C/R 位号换成电感等）
_PREFIX_EXPECT = {"C": "电容", "R": "电阻", "L": "电感"}


_CODE_PREFIX_TYPE = {"L": "电感", "C": "电容", "R": "电阻"}


def _actual_type(desc: str, code: str, model: str) -> str | None:
    # 1) 描述中文类型词（最可靠）
    if "电感" in desc:
        return "电感"
    if "电阻" in desc:
        return "电阻"
    if "电容" in desc:
        return "电容"
    # 2) 料号编码规则前缀（L./C./R. = 电感/电容/电阻），描述为空时的关键判据
    match = re.match(r"^([A-Za-z]+)\.", str(code))
    if match:
        by_code = _CODE_PREFIX_TYPE.get(match.group(1).upper())
        if by_code:
            return by_code
    # 3) 型号单位：nH/uH/mH = 电感；pF/nF/uF = 电容
    if re.search(r"\d\s*[munpμµ]?H\b", str(model)):
        return "电感"
    if re.search(r"\d\s*[munpμµ]?F\b", str(model)):
        return "电容"
    return None


def _type_mismatches(rows: list[dict[str, object]]) -> list[dict[str, str]]:
    out: list[dict[str, str]] = []
    for row in rows:
        actual = _actual_type(str(row.get("description", "")), str(row.get("part_number", "")), str(row.get("model", "")))
        if not actual:
            continue
        for ref in row.get("refs") or []:
            match = re.match(r"^([A-Za-z]+)", ref)
            prefix = match.group(1).upper() if match else ""
            expect = _PREFIX_EXPECT.get(prefix)
            if expect and expect != actual:
                out.append({
                    "ref": ref, "code": row.get("part_number", ""), "desc": row.get("description", ""),
                    "note": f"位号 {prefix}（通常为{expect}）实为{actual}",
                })
    return out


def _risk_check(rows: list[dict[str, object]]) -> list[dict[str, str]]:
    """单份 BOM 的导入前风险自检，输出一组检查项（ok/warn/info）。"""
    from collections import Counter

    def blob(row: dict[str, object]) -> str:
        return f"{row.get('part_number','')} {row.get('model','')} {row.get('description','')}"

    findings: list[dict[str, str]] = []

    pcb = sorted({r["part_number"] for r in rows if _looks_like_pcb(r)})
    findings.append({"name": "PCB 裸板", "status": "ok" if pcb else "warn",
                     "message": "找到 " + ", ".join(pcb) if pcb else "未发现 PCB 裸板项（PCBA BOM 通常应包含裸板）"})

    shield_brackets = sorted({r["part_number"] for r in rows if _looks_like_shield_bracket(r)})
    findings.append({"name": "屏蔽支架", "status": "ok" if shield_brackets else "info",
                     "message": "找到 " + ", ".join(shield_brackets) if shield_brackets else "未发现屏蔽支架（如设计需要请确认）"})

    nc_refs = [rf for r in rows if any(k in blob(r) for k in ["未贴", "不贴", "DNP"]) or _NC_RE.search(blob(r)) for rf in (r.get("refs") or [])]
    findings.append({"name": "NC/未贴器件", "status": "warn" if nc_refs else "ok",
                     "message": f"混入 {len(nc_refs)} 个：" + ",".join(nc_refs[:15]) if nc_refs else "无"})

    mech = sorted({r["part_number"] for r in rows if any(k.upper() in blob(r).upper() for k in _MECH_KW)})
    findings.append({"name": "机构件/螺丝/孔", "status": "warn" if mech else "ok",
                     "message": "混入：" + ", ".join(mech) if mech else "无"})

    tp_refs = sorted({rf for r in rows for rf in (r.get("refs") or []) if rf.upper().startswith(_TP_PREFIX)}) + \
        sorted({r["part_number"] for r in rows if any(k.upper() in blob(r).upper() for k in _TP_KW)})
    findings.append({"name": "测试点/跳线/工艺", "status": "info" if tp_refs else "ok",
                     "message": "发现：" + ", ".join(tp_refs[:15]) if tp_refs else "无"})

    counts = Counter(rf for r in rows for rf in (r.get("refs") or []))
    dup = sorted([rf for rf, n in counts.items() if n > 1])
    findings.append({"name": "重复位号", "status": "warn" if dup else "ok",
                     "message": ", ".join(dup[:15]) if dup else "无"})

    empty = [r for r in rows if (r.get("refs") or []) and not str(r.get("part_number") or "").strip()]
    findings.append({"name": "空编号行", "status": "warn" if empty else "ok",
                     "message": f"{len(empty)} 行有位号但料号为空" if empty else "无"})

    mismatch = [r for r in rows if (r.get("refs") or []) and str(r.get("quantity")).strip() not in ("", str(len(r["refs"])))]
    findings.append({"name": "数量=位号数", "status": "warn" if mismatch else "ok",
                     "message": f"{len(mismatch)} 行不符，例 {mismatch[0]['part_number']}：数量{mismatch[0]['quantity']}≠位号{len(mismatch[0]['refs'])}" if mismatch else "全部一致"})

    mism = _type_mismatches(rows)
    findings.append({"name": "位号/器件类型", "status": "warn" if mism else "ok",
                     "message": (f"{len(mism)} 处位号与器件类型不符（疑似占位换料）：" + ", ".join(m["ref"] for m in mism[:10])) if mism else "无"})

    flagged = [r for r in rows if str(r.get("grade") or "").strip() and str(r.get("grade")).strip() not in ("优选", "正常")]
    if flagged:
        kinds = Counter(str(r.get("grade")).strip() for r in flagged)
        detail = ", ".join(f"{g}×{n}" for g, n in kinds.most_common())
        findings.append({"name": "物料优选等级", "status": "warn", "message": f"{len(flagged)} 项非优选/正常：{detail}"})
    else:
        has_grade = any(str(r.get("grade") or "").strip() for r in rows)
        findings.append({"name": "物料优选等级", "status": "ok" if has_grade else "info",
                         "message": "均为优选/正常" if has_grade else "未提供等级列"})

    sensitive = _version_sensitive_parts(rows)
    if sensitive:
        codes = sorted({str(row.get("part_number") or "").strip() for row in sensitive if str(row.get("part_number") or "").strip()})
        findings.append({"name": "硬件版本敏感物料", "status": "info",
                         "message": "发现 eMMC/DDR 相关物料：" + ", ".join(codes[:10]) + "；请注意核对硬件版本号、容量/速率和替代关系"})
    else:
        findings.append({"name": "硬件版本敏感物料", "status": "ok", "message": "未发现 eMMC/DDR 相关物料"})

    return findings


def run_bom_compare(root: Path, params: dict[str, object]) -> dict[str, object]:
    bom1, error = _required_file(params, "bom1", "BOM1 文件")
    if error:
        return _error("bom_compare", error)
    bom2, error = _required_file(params, "bom2", "BOM2 文件")
    if error:
        return _error("bom_compare", error)

    rows1 = _read_bom_rows(bom1, require_refs=False)
    rows2 = _read_bom_rows(bom2, require_refs=False)

    # —— 主轴：按位号对比，自动识别 换料 / 新增 / 删除 / 一致 ——
    ref1 = _index_by_ref(rows1)
    ref2 = _index_by_ref(rows2)
    all_refs = sorted(set(ref1) | set(ref2), key=_natural_key)

    pos_rows: list[list[object]] = []
    pos_items: list[dict[str, object]] = []
    status_counts = {"same": 0, "swap": 0, "added": 0, "removed": 0, "param": 0}
    changed = 0
    for ref in all_refs:
        left = ref1.get(ref)
        right = ref2.get(ref)
        diffs: list[str] = []
        if left and not right:
            status, kind, badge = "删除/未贴", "only_left", "删除"
            status_counts["removed"] += 1
        elif right and not left:
            status, kind, badge = "新增贴装", "only_right", "新增"
            status_counts["added"] += 1
        else:
            for field in ("编号", "型号", "描述"):
                if str(left.get(field, "")) != str(right.get(field, "")):
                    diffs.append(field)
            if "编号" in diffs:
                status, kind, badge = "换料", "diff", "换料"
                status_counts["swap"] += 1
            elif diffs:
                status, kind, badge = "参数差异", "diff", "参数差异"
                status_counts["param"] += 1
            else:
                status, kind, badge = "一致", "same", "一致"
                status_counts["same"] += 1
        if kind != "same":
            changed += 1
        pos_rows.append(
            [
                ref,
                status,
                left["编号"] if left else "",
                right["编号"] if right else "",
                left["型号"] if left else "",
                right["型号"] if right else "",
                left["描述"] if left else "",
                right["描述"] if right else "",
            ]
        )
        pos_items.append(
            {
                "key": ref,
                "status": status,
                "kind": kind,
                "badge": badge,
                "left": left,
                "right": right,
                "diff": diffs,
            }
        )

    # —— 辅助：料号用量变化（采购/ERP 视角，能识别换料替入/替出，而非孤立的新增/移除）——
    usage1 = _part_usage(rows1)
    usage2 = _part_usage(rows2)
    summary_rows: list[list[object]] = []
    for part_number in sorted(set(usage1) | set(usage2), key=_natural_key):
        s1 = usage1[part_number]["refs"] if part_number in usage1 else set()
        s2 = usage2[part_number]["refs"] if part_number in usage2 else set()
        count1 = _usage_count(usage1[part_number]) if part_number in usage1 else 0
        count2 = _usage_count(usage2[part_number]) if part_number in usage2 else 0
        if s1 == s2 and count1 == count2:
            continue
        ref_part = usage1.get(part_number) or usage2.get(part_number)
        lost, gained = s1 - s2, s2 - s1
        out_targets = sorted({ref2[r]["编号"] for r in lost if r in ref2 and ref2[r]["编号"]})
        in_sources = sorted({ref1[r]["编号"] for r in gained if r in ref1 and ref1[r]["编号"]})

        if count1 == 0:  # 本份没有 -> 该料是新出现的
            if gained and all(r in ref1 for r in gained):
                change = "换料替入 ← " + (in_sources[0] if len(in_sources) == 1 else "多项")
            elif in_sources:
                change = "新增(部分替换)"
            else:
                change = "新增"
        elif count2 == 0:  # 本份有、对面没有 -> 该料消失
            if lost and all(r in ref2 for r in lost):
                change = "换料替出 → " + (out_targets[0] if len(out_targets) == 1 else "多项")
            elif out_targets:
                change = "移除(部分被替换)"
            else:
                change = "移除"
        else:  # 同料号、位号增减
            delta = count2 - count1
            change = f"+{delta}" if delta > 0 else str(delta) if delta < 0 else "位号变更"
        summary_rows.append([part_number, ref_part["model"], ref_part["desc"], count1, count2, change])

    pos_headers = ["位号", "状态", "BOM1编号", "BOM2编号", "BOM1型号", "BOM2型号", "BOM1描述", "BOM2描述"]
    sum_headers = ["编号", "型号", "描述", "BOM1数量", "BOM2数量", "变化"]
    output = _output_dir(params, root, "bom") / f"BOM差异报告_{_timestamp()}.xlsx"
    _write_sheets(output, [("位号对照", pos_headers, pos_rows), ("料号用量小结", sum_headers, summary_rows)])

    table = _table(pos_headers, pos_rows, status_col=1, diff_pairs=[[2, 3], [4, 5], [6, 7]])
    compare = _compare("位号", "BOM1", "BOM2", ["编号", "型号", "描述"], pos_items)
    part_summary = _table(sum_headers, summary_rows, status_col=5, diff_pairs=[[3, 4]])
    origin = {
        "left_label": "BOM1",
        "right_label": "BOM2",
        "columns": ["编号", "型号", "描述", "数量", "位号"],
        "ref_col": 4,
        "left_rows": _annotate_origin(rows1, ref2, set(usage2), "left"),
        "right_rows": _annotate_origin(rows2, ref1, set(usage1), "right"),
        "total_left": len(rows1),
        "total_right": len(rows2),
    }

    focus_priority = {"换料": 0, "删除/未贴": 1, "新增贴装": 2, "参数差异": 3}
    focus_items = sorted(
        [item for item in pos_items if item["kind"] != "same"],
        key=lambda item: (focus_priority.get(str(item.get("status")), 9), _natural_key(str(item.get("key", "")))),
    )[:200]
    review_guide = {
        "换料": "重点确认同一位号的物料编码是否按设计变更替换，并回查型号、描述和替代关系。",
        "新增贴装": "确认是否为新增器件、原 NC 改贴或设计补料，必要时同步工艺和采购。",
        "删除/未贴": "确认是否为删除器件、改 NC/DNP 或漏导出，避免误删应贴物料。",
        "参数差异": "料号未变但型号/描述变化，通常用于确认库属性、规格描述和 PLM 元数据是否同步。",
        "料号用量": "从采购/ERP 视角核对每个编码的数量变化，识别替入、替出、新增和移除。",
    }

    result = _result(
        "bom_compare",
        [output],
        {
            "total_positions": len(all_refs),
            "changed_positions": changed,
            "part_changes": len(summary_rows),
            "status_counts": status_counts,
        },
        table,
        compare,
    )
    result["part_summary"] = part_summary
    result["focus_items"] = focus_items
    result["review_guide"] = review_guide
    result["origin"] = origin
    result["risks"] = {
        "left_label": "BOM1",
        "right_label": "BOM2",
        "left": _risk_check(rows1),
        "right": _risk_check(rows2),
    }
    return result


def run_bom_risk_check(root: Path, params: dict[str, object]) -> dict[str, object]:
    bom, error = _required_file(params, "bom", "BOM 文件")
    if error:
        return _error("bom_risk_check", error)

    rows = _read_bom_rows(bom, require_refs=False)
    findings = _risk_check(rows)
    positions = sum(len(row.get("refs") or []) for row in rows)
    parts = len({row["part_number"] for row in rows if row.get("part_number")})

    level_cn = {"ok": "通过", "warn": "警告", "info": "提示"}
    headers = ["检查项", "结果", "说明"]
    report_rows = [[f["name"], level_cn.get(f["status"], f["status"]), f["message"]] for f in findings]
    output = _output_dir(params, root, "risk") / f"BOM风险检查_{_timestamp()}.xlsx"
    _write_table(output, "风险检查", headers, report_rows)

    def _grade(row: dict[str, object]) -> str:
        return str(row.get("grade") or "").strip()

    grade_flags = [
        {"code": r.get("part_number", ""), "desc": r.get("description", ""),
         "refs": r.get("reference", ""), "grade": _grade(r)}
        for r in rows
        if _grade(r) and _grade(r) not in ("优选", "正常")
    ]
    review_headers = ["编号", "型号", "描述", "数量", "位号", "等级"]
    review_rows = [
        [r.get("part_number", ""), r.get("model", ""), r.get("description", ""),
         _jsonable(r.get("quantity")), r.get("reference", ""), _grade(r)]
        for r in rows
    ]

    warnings = sum(1 for f in findings if f["status"] == "warn")
    result = _result(
        "bom_risk_check",
        [output],
        {"rows": len(rows), "positions": positions, "parts": parts, "warnings": warnings, "grade_flags": len(grade_flags)},
    )
    result["risk_report"] = {
        "label": bom.name,
        "findings": findings,
        "stats": {"数据行": len(rows), "位号数": positions, "料号数": parts},
        "grade_flags": grade_flags,
        "type_flags": _type_mismatches(rows),
        "review": {"headers": review_headers, "rows": review_rows, "grade_col": 5},
    }
    return result


def run_generic_bom_import(root: Path, params: dict[str, object]) -> dict[str, object]:
    source, error = _required_file(params, "source_bom", "原始 BOM 文件")
    if error:
        return _error("bom_import", error)
    rows = _read_bom_rows(source)
    main_rows = []
    excluded_rows = []
    for row in rows:
        refs = row["refs"]
        text = f"{row.get('description', '')} {row.get('name', '')}"
        value = str(row.get("value", ""))
        reason = ""
        if not row.get("part_number"):
            reason = "子项编码为空"
        elif value.upper().startswith("NC") or "NC/" in text.upper():
            reason = "NC/未贴"
        elif any(ref.upper().startswith(("TP", "Z_TP", "JP", "SH")) for ref in refs):
            reason = "测试点/跳线/非贴装"
        elif any(token in text for token in ["测试点", "跳线", "Test"]):
            reason = "字段包含测试/跳线"

        if reason:
            excluded_rows.append(
                [
                    row["source_row"],
                    row["reference"],
                    row["part_number"],
                    row["name"],
                    row["package"],
                    row["description"],
                    reason,
                ]
            )
        else:
            main_rows.append(
                [
                    row["part_number"],
                    row["name"],
                    row["package"],
                    row["description"],
                    len(refs),
                    row["reference"],
                    "直接领料",
                    "是",
                    "否",
                ]
            )

    out_dir = _output_dir(params, root, "bom")
    stem = source.stem
    main_output = out_dir / f"{stem}_主BOM.xlsx"
    nc_output = out_dir / f"{stem}_NC未贴器件汇总.xlsx"
    _write_table(main_output, "主BOM", ["子项编码", "名称", "型号/封装", "描述", "数量", "位号", "发料方式", "是否参与MRP运算", "是否跳层"], main_rows)
    _write_table(nc_output, "NC未贴汇总", ["原始行号", "位号", "子项编码", "名称", "型号/封装", "描述", "过滤原因"], excluded_rows)
    return _result(
        "bom_import",
        [main_output, nc_output],
        {"main_rows": len(main_rows), "excluded_rows": len(excluded_rows)},
    )


def _read_text_guess(path: Path) -> str:
    for encoding in ("utf-8", "gb18030", "cp936"):
        try:
            return path.read_text(encoding=encoding)
        except UnicodeDecodeError:
            continue
    return path.read_text(encoding="utf-8", errors="ignore")


def _clean_pst_string(value: str) -> str:
    text = value.strip().rstrip(";").rstrip(":").strip()
    if len(text) >= 2 and text[0] == "'" and text[-1] == "'":
        text = text[1:-1]
    return text.strip()


def _natural_join(values: Iterable[str]) -> str:
    return ",".join(sorted(set(values), key=_natural_key))


def _parse_node_tokens(line: str) -> tuple[str, str] | None:
    tokens = line.strip().split()
    if len(tokens) >= 3 and tokens[0].upper() == "NODE_NAME":
        return tokens[1], tokens[2]
    return None


def _parse_net_file(folder: Path) -> dict[str, dict[str, list[str]]]:
    path = folder / "pstxnet.dat"
    if not path.exists():
        raise ValueError(f"缺少 pstxnet.dat: {folder}")
    nets: dict[str, dict[str, set[str]]] = {}
    current: str | None = None
    pending_name = False
    for raw in _read_text_guess(path).splitlines():
        line = raw.strip()
        if not line:
            continue
        upper = line.upper()
        if upper == "NET_NAME":
            pending_name = True
            current = None
            continue
        if pending_name:
            name = _clean_pst_string(line)
            if name and not name.startswith("@") and "=" not in name:
                current = name
                nets.setdefault(current, {"refs": set(), "pins": set(), "nodes": set()})
                pending_name = False
            continue
        node = _parse_node_tokens(line)
        if node and current:
            ref, pin = node
            nets[current]["refs"].add(ref)
            nets[current]["pins"].add(pin)
            nets[current]["nodes"].add(f"{ref}.{pin}")
            continue

        # Fallback for simple whitespace netlists: NET N1 R1.1 C1.2
        parts = line.split()
        if not parts:
            continue
        if parts[0].upper() == "NET" and len(parts) >= 2:
            name = parts[1]
            tokens = parts[2:]
        else:
            name = parts[0]
            tokens = parts[1:]
        if not tokens or name in {"FILE_TYPE", "C_SIGNAL"} or "=" in name:
            continue
        entry = nets.setdefault(_clean_pst_string(name), {"refs": set(), "pins": set(), "nodes": set()})
        for token in tokens:
            clean = _clean_pst_string(token)
            ref, _, pin = clean.partition(".")
            if not ref:
                continue
            entry["refs"].add(ref)
            if pin:
                entry["pins"].add(pin)
                entry["nodes"].add(f"{ref}.{pin}")
            else:
                entry["nodes"].add(ref)
    return {
        name: {
            "refs": sorted(data["refs"], key=_natural_key),
            "pins": sorted(data["pins"], key=_natural_key),
            "nodes": sorted(data["nodes"], key=_natural_key),
        }
        for name, data in nets.items()
    }


def _parse_part_file(folder: Path) -> dict[str, str]:
    path = folder / "pstxprt.dat"
    if not path.exists():
        raise ValueError(f"缺少 pstxprt.dat: {folder}")
    parts: dict[str, str] = {}
    pending_part = False
    part_re = re.compile(r"^([A-Za-z]+\d+[A-Za-z0-9_-]*)\s+'([^']+)'")
    for raw in _read_text_guess(path).splitlines():
        line = raw.strip()
        if not line:
            continue
        if line.upper() == "PART_NAME":
            pending_part = True
            continue
        match = part_re.match(line)
        if match:
            parts[match.group(1)] = match.group(2).strip()
            pending_part = False
            continue
        if pending_part:
            tokens = line.split(None, 1)
            if len(tokens) >= 2:
                parts[tokens[0]] = _clean_pst_string(tokens[1])
                pending_part = False
    return parts


def _parse_part_file_optional(folder: Path) -> tuple[dict[str, str], str | None]:
    try:
        return _parse_part_file(folder), None
    except ValueError as exc:
        if "pstxprt.dat" in str(exc):
            return {}, f"缺少 pstxprt.dat：{folder}，已跳过器件封装变化检查，仅执行网络节点对比。"
        raise


def _package_tokens(value: str) -> set[str]:
    raw = str(value or "").upper()
    tokens = {token for token in re.split(r"[^A-Z0-9]+", raw) if len(token) >= 2}
    joined = re.sub(r"[^A-Z0-9]", "", raw)
    if joined:
        tokens.add(joined)
    return tokens


_COMMON_PACKAGE_SIZES = {
    "008004", "01005", "0201", "03015", "0402", "0603", "0805", "0806", "1005", "1206", "1210",
    "1608", "2012", "2016", "2520", "3216", "3225", "3528", "4020", "5032", "6032", "7343",
}

_VENDOR_SIZE_PATTERNS = [
    (re.compile(r"\b(CL03|GRM03|GJM03|0201X|RC0201|WR02|RTT01|0201WMF)"), "0201"),
    (re.compile(r"\b(CL05|GRM15[35]|GCM15[35]|GJM15[35]|TDK105|C0402|0402X|RC0402|WR04|RTT02|0402WMF)"), "0402"),
]


def _package_size_codes(value: str) -> set[str]:
    text = str(value or "").upper()
    found = {match.group(1) for match in re.finditer(r"(?<!\d)(\d{4,6})(?!\d)", text) if match.group(1) in _COMMON_PACKAGE_SIZES}
    found.update(match.group(1) for match in re.finditer(r"[CRL](\d{4})(?:[^0-9]|$)", text) if match.group(1) in _COMMON_PACKAGE_SIZES)
    for pattern, size in _VENDOR_SIZE_PATTERNS:
        if pattern.search(text):
            found.add(size)
    return found


def _package_matches(net_package: str, bom_text: str) -> tuple[bool, str]:
    net_tokens = _package_tokens(net_package)
    bom_tokens = _package_tokens(bom_text)
    if not net_tokens or not bom_tokens:
        return False, "缺少可比对封装信息"
    common = net_tokens & bom_tokens
    if common:
        return True, "匹配到封装关键词: " + ",".join(sorted(common, key=_natural_key)[:5])
    long_net = {t for t in net_tokens if len(t) >= 5}
    long_bom = {t for t in bom_tokens if len(t) >= 5}
    for a in long_net:
        for b in long_bom:
            if a in b or b in a:
                return True, f"封装字段近似匹配: {a}/{b}"
    common_sizes = _package_size_codes(net_package) & _package_size_codes(bom_text)
    if common_sizes:
        return True, "匹配到封装尺寸码: " + ",".join(sorted(common_sizes, key=_natural_key)[:5])
    return False, "网表封装未出现在 BOM 描述/名称/封装字段"


_SMT_HIGH_RISK_RE = re.compile(
    r"(^|[^A-Z0-9])(BGA|WLCSP|CSP|QFN|DFN|LGA|QFP|BTB|FPC|FFC|USB|HDMI|TYPEC|EMMC|UFS|DDR|LPDDR|MIPI|CSI|DSI)",
    re.IGNORECASE,
)


def _smt_status_key(status: str) -> str:
    return {
        "通过": "passed",
        "近似通过": "near",
        "需要确认": "manual",
        "BOM 缺位号": "missing_bom",
        "BOM 多余位号": "extra_bom",
        "同料多封装": "multi_package",
        "高风险封装": "high_risk",
        "NC 未贴跳过": "nc_skipped",
        "非贴片对象跳过": "non_smt_skipped",
    }.get(status, "manual")


def _smt_item(
    ref: str,
    status: str,
    net_package: str = "",
    bom_row: dict[str, object] | None = None,
    note: str = "",
    severity: str = "medium",
    part_number: str = "",
) -> dict[str, object]:
    row = bom_row or {}
    return {
        "key": f"{status}:{ref}:{part_number or row.get('part_number', '')}",
        "ref": ref,
        "status": status,
        "kind": _smt_status_key(status),
        "severity": severity,
        "part_number": part_number or str(row.get("part_number", "")),
        "net_package": net_package,
        "bom_package": str(row.get("package", "")),
        "model": str(row.get("model", "")),
        "description": str(row.get("description", "")),
        "name": str(row.get("name", "")),
        "grade": str(row.get("grade", "")),
        "note": note,
    }


def _is_high_risk_package(*values: object) -> bool:
    return bool(_SMT_HIGH_RISK_RE.search(" ".join(str(value or "") for value in values)))


def _is_nc_package(value: object) -> bool:
    text = str(value or "").upper()
    return bool(re.search(r"(^|[^A-Z0-9])NC([_/\\\-\s]|$)", text))


def _is_non_smt_netlist_part(ref: object, package: object) -> bool:
    ref_text = str(ref or "").upper()
    pkg = str(package or "").upper()
    if ref_text.startswith(("TP", "JP", "MTG", "MH")):
        return True
    if ref_text.startswith("H") and re.match(r"^H\d+", ref_text):
        return True
    return any(token in pkg for token in ["SHORT", "TP_", "MARK", "FID", "HOLE", "SCREW", "MTG", "MOUNT"])


def _build_smt_package_review(parts: dict[str, str], bom_rows: list[dict[str, object]]) -> dict[str, object]:
    by_ref = {ref: row for row in bom_rows for ref in row["refs"]}
    items: list[dict[str, object]] = []
    rows: list[list[object]] = []
    status_counts = {
        "passed": 0,
        "near": 0,
        "manual": 0,
        "missing_bom": 0,
        "extra_bom": 0,
        "multi_package": 0,
        "high_risk": 0,
        "nc_skipped": 0,
        "non_smt_skipped": 0,
    }

    for ref, package in sorted(parts.items(), key=lambda item: _natural_key(item[0])):
        bom_row = by_ref.get(ref)
        if not bom_row:
            if _is_nc_package(package):
                status = "NC 未贴跳过"
                note = "网表中为 NC/未贴器件，最终 PCBA BOM 不包含该位号属于正常情况。"
                item = _smt_item(ref, status, package, None, note, "low")
            elif _is_non_smt_netlist_part(ref, package):
                status = "非贴片对象跳过"
                note = "测试点、短接、安装孔或工艺对象通常不进入最终贴片 BOM，未在 BOM 中出现时不按缺失处理。"
                item = _smt_item(ref, status, package, None, note, "low")
            else:
                status = "BOM 缺位号"
                note = "网表存在该位号，但 BOM 中没有找到。确认是否漏导、未贴或不应进入贴片 BOM。"
                item = _smt_item(ref, status, package, None, note, "high")
                rows.append([ref, package, "", "", "", status, note])
        else:
            desc = str(bom_row.get("description", ""))
            name = str(bom_row.get("name", ""))
            bom_package = str(bom_row.get("package", ""))
            model = str(bom_row.get("model", ""))
            ok, note = _package_matches(package, f"{desc} {name} {bom_package} {model}")
            if ok and note.startswith("封装字段近似匹配"):
                status = "近似通过"
                severity = "low"
            elif ok:
                status = "通过"
                severity = "low"
            else:
                status = "需要确认"
                severity = "medium"
            item = _smt_item(ref, status, package, bom_row, note, severity)
            table_status = "机器初筛通过" if status in {"通过", "近似通过"} else "请人工判断"
            rows.append([ref, package, bom_package or model, desc, name, table_status, note])
        status_counts[_smt_status_key(status)] += 1
        items.append(item)

    for ref in sorted(set(by_ref) - set(parts), key=_natural_key):
        bom_row = by_ref[ref]
        note = "BOM 中存在该位号，但 pstxprt.dat 没有找到。确认是否机构/辅料、手工添加、或网表导出不完整。"
        item = _smt_item(ref, "BOM 多余位号", "", bom_row, note, "medium")
        items.append(item)
        status_counts["extra_bom"] += 1

    packages_by_part: dict[str, set[str]] = {}
    refs_by_part: dict[str, list[str]] = {}
    rows_by_part: dict[str, dict[str, object]] = {}
    for ref, package in parts.items():
        row = by_ref.get(ref)
        part_number = str(row.get("part_number", "") if row else "").strip()
        if not part_number:
            continue
        packages_by_part.setdefault(part_number, set()).add(package)
        refs_by_part.setdefault(part_number, []).append(ref)
        rows_by_part.setdefault(part_number, row)
    for part_number, packages in sorted(packages_by_part.items(), key=lambda item: _natural_key(item[0])):
        normalized = {re.sub(r"[^A-Z0-9]", "", package.upper()) for package in packages if package}
        if len(normalized) <= 1:
            continue
        refs = sorted(refs_by_part.get(part_number, []), key=_natural_key)
        note = "同一个物料编码对应多个网表封装：" + " / ".join(sorted(packages, key=_natural_key))
        item = _smt_item(
            ",".join(refs),
            "同料多封装",
            " / ".join(sorted(packages, key=_natural_key)),
            rows_by_part.get(part_number),
            note,
            "high",
            part_number,
        )
        item["refs"] = refs
        items.append(item)
        status_counts["multi_package"] += 1

    high_risk_seen: set[str] = set()
    for item in list(items):
        if item["status"] in {"BOM 缺位号", "同料多封装"}:
            continue
        if not _is_high_risk_package(item.get("net_package"), item.get("bom_package"), item.get("model"), item.get("description"), item.get("name")):
            continue
        ref = str(item.get("ref", ""))
        if ref in high_risk_seen:
            continue
        high_risk_seen.add(ref)
        note = "BGA/QFN/连接器/存储/高速相关封装，建议人工确认 footprint、焊盘、硬件版本和替代料一致性。"
        items.append(_smt_item(ref, "高风险封装", str(item.get("net_package", "")), {
            "part_number": item.get("part_number", ""),
            "package": item.get("bom_package", ""),
            "model": item.get("model", ""),
            "description": item.get("description", ""),
            "name": item.get("name", ""),
            "grade": item.get("grade", ""),
        }, note, "medium"))
        status_counts["high_risk"] += 1

    priority = {"high": 0, "medium": 1, "low": 2}
    focus_items = sorted(
        [item for item in items if item["status"] not in {"通过", "近似通过", "高风险封装", "NC 未贴跳过", "非贴片对象跳过"}],
        key=lambda item: (priority.get(str(item.get("severity")), 9), _natural_key(str(item.get("ref", ""))), str(item.get("status", ""))),
    )[:500]
    return {
        "items": items,
        "focus_items": focus_items,
        "status_counts": status_counts,
        "table_rows": rows,
        "review_guide": {
            "通过": "网表封装和 BOM 封装/型号/描述中存在明确关键词匹配，通常可快速放行。",
            "近似通过": "封装字段存在包含关系，建议抽查命名是否为同一 footprint。",
            "需要确认": "网表封装没有在 BOM 描述、名称、型号或封装字段中匹配到，需要人工核对。",
            "BOM 缺位号": "pstxprt.dat 中有位号但 BOM 没有，重点确认是否漏导或不应入 BOM。",
            "BOM 多余位号": "BOM 中有位号但网表没有，重点确认手工添加、机构辅料或位号错误。",
            "同料多封装": "同一个物料编码出现在多个 footprint 上，通常需要拆料号或确认封装兼容。",
            "高风险封装": "BGA/QFN/连接器/存储/高速相关物料变更成本高，建议人工二次确认。",
            "NC 未贴跳过": "网表中标为 NC/未贴且最终 PCBA BOM 不包含时，不作为缺失异常处理。",
            "非贴片对象跳过": "测试点、短接、安装孔或工艺对象不进入最终贴片 BOM 时，不作为缺失异常处理。",
        },
    }


_CRITICAL_NET_RE = re.compile(r"(GND|VSS|VDD|VCC|POWER|BUCK|SYS_|USB|MIPI|CSI|DSI|DDR|EMMC|CLK|CLOCK|RST|RESET|I2C|SPI|UART)", re.IGNORECASE)


def _net_signature(net: dict[str, list[str]]) -> tuple[str, ...]:
    return tuple(sorted(net.get("nodes", []), key=_natural_key))


def _is_critical_net(name: str) -> bool:
    return bool(_CRITICAL_NET_RE.search(str(name or "")))


def _net_item(
    key: str,
    status: str,
    kind: str,
    left: dict[str, object] | None,
    right: dict[str, object] | None,
    diff: list[str],
    message: str,
    severity: str = "medium",
) -> dict[str, object]:
    if status == "关键网络变化":
        severity = "high"
    return {
        "key": key,
        "status": status,
        "kind": kind,
        "left": left,
        "right": right,
        "diff": diff,
        "message": message,
        "severity": severity,
        "critical": _is_critical_net(key) or (left and _is_critical_net(str(left.get("网络", "")))) or (right and _is_critical_net(str(right.get("网络", "")))),
    }


def _build_netlist_review(nets1: dict[str, dict[str, list[str]]], nets2: dict[str, dict[str, list[str]]], package_rows: list[list[object]]) -> dict[str, object]:
    sig1: dict[tuple[str, ...], list[str]] = {}
    sig2: dict[tuple[str, ...], list[str]] = {}
    for name, data in nets1.items():
        sig1.setdefault(_net_signature(data), []).append(name)
    for name, data in nets2.items():
        sig2.setdefault(_net_signature(data), []).append(name)

    left_node_to_net = {node: name for name, data in nets1.items() for node in data.get("nodes", [])}
    right_node_to_net = {node: name for name, data in nets2.items() for node in data.get("nodes", [])}
    covered_left: set[str] = set()
    covered_right: set[str] = set()
    items: list[dict[str, object]] = []
    status_counts = {"rename": 0, "split": 0, "merge": 0, "node_change": 0, "added": 0, "removed": 0, "package": 0, "same": 0}

    def side(name: str, data: dict[str, list[str]]) -> dict[str, object]:
        return {"网络": name, "位号": _natural_join(data.get("refs", [])), "节点": _natural_join(data.get("nodes", [])), "节点数": len(data.get("nodes", []))}

    for signature, left_names in sig1.items():
        right_names = sig2.get(signature, [])
        if len(left_names) == 1 and len(right_names) == 1 and left_names[0] != right_names[0]:
            left_name, right_name = left_names[0], right_names[0]
            covered_left.add(left_name)
            covered_right.add(right_name)
            status_counts["rename"] += 1
            items.append(_net_item(
                f"{left_name} -> {right_name}",
                "网络改名",
                "rename",
                side(left_name, nets1[left_name]),
                side(right_name, nets2[right_name]),
                ["网络名"],
                "节点集合完全一致，仅网络名变化。",
                "low",
            ))

    for left_name, left_data in nets1.items():
        if left_name in covered_left or left_name in nets2:
            continue
        nodes = set(left_data.get("nodes", []))
        targets = sorted({right_node_to_net[node] for node in nodes if node in right_node_to_net}, key=_natural_key)
        if len(targets) >= 2 and nodes and all(node in right_node_to_net for node in nodes):
            covered_left.add(left_name)
            covered_right.update(targets)
            status_counts["split"] += 1
            items.append(_net_item(
                left_name,
                "疑似拆网",
                "split",
                side(left_name, left_data),
                {"网络": ",".join(targets), "节点": _natural_join(node for name in targets for node in nets2[name].get("nodes", []))},
                ["节点"],
                "旧版一个网络的节点在新版分散到多个网络。",
                "high",
            ))

    for right_name, right_data in nets2.items():
        if right_name in covered_right or right_name in nets1:
            continue
        nodes = set(right_data.get("nodes", []))
        sources = sorted({left_node_to_net[node] for node in nodes if node in left_node_to_net}, key=_natural_key)
        if len(sources) >= 2 and nodes and all(node in left_node_to_net for node in nodes):
            covered_right.add(right_name)
            covered_left.update(sources)
            status_counts["merge"] += 1
            items.append(_net_item(
                right_name,
                "疑似并网",
                "merge",
                {"网络": ",".join(sources), "节点": _natural_join(node for name in sources for node in nets1[name].get("nodes", []))},
                side(right_name, right_data),
                ["节点"],
                "旧版多个网络的节点在新版合并到一个网络。",
                "high",
            ))

    for name in sorted(set(nets1) | set(nets2), key=_natural_key):
        if name in covered_left or name in covered_right:
            continue
        in1, in2 = name in nets1, name in nets2
        if in1 and not in2:
            status_counts["removed"] += 1
            items.append(_net_item(name, "网络删除", "removed", side(name, nets1[name]), None, ["节点"], "该网络只存在于网表1。", "medium"))
        elif in2 and not in1:
            status_counts["added"] += 1
            items.append(_net_item(name, "网络新增", "added", None, side(name, nets2[name]), ["节点"], "该网络只存在于网表2。", "medium"))
        else:
            left_nodes = set(nets1[name].get("nodes", []))
            right_nodes = set(nets2[name].get("nodes", []))
            if left_nodes != right_nodes:
                status_counts["node_change"] += 1
                status = "关键网络变化" if _is_critical_net(name) else "节点变化"
                items.append(_net_item(
                    name,
                    status,
                    "node_change",
                    side(name, nets1[name]),
                    side(name, nets2[name]),
                    ["节点"],
                    "同名网络的连接节点发生变化。",
                    "high" if _is_critical_net(name) else "medium",
                ))
            else:
                status_counts["same"] += 1

    for row in package_rows:
        ref = str(row[0])
        status_counts["package"] += 1
        items.append(_net_item(
            f"器件:{ref}",
            "封装变化",
            "package",
            {"位号": ref, "封装": row[1]},
            {"位号": ref, "封装": row[2]},
            ["封装"],
            "同一位号在两版网表中的封装不同。",
            "medium",
        ))

    priority = {"high": 0, "medium": 1, "low": 2}
    focus_items = sorted([item for item in items if item["status"] != "一致"], key=lambda item: (priority.get(str(item.get("severity")), 9), _natural_key(str(item.get("key", "")))))[:300]
    return {
        "items": items,
        "focus_items": focus_items,
        "status_counts": status_counts,
        "review_guide": {
            "网络改名": "节点集合一致但网络名变化，通常用于确认命名整理或跨页网络名变更是否符合预期。",
            "节点变化": "同名网络的 pin 连接发生变化，需要确认新增/删除节点是否为设计变更。",
            "关键网络变化": "电源、时钟、复位、高速或存储相关网络发生连接变化，优先人工复核。",
            "疑似拆网": "旧版一个网络在新版拆成多个网络，重点确认是否误断开。",
            "疑似并网": "旧版多个网络在新版合成一个网络，重点确认是否误短接。",
            "封装变化": "同位号封装变化，需要确认 PCB footprint 和 BOM 描述同步。",
        },
    }


def run_netlist_compare(root: Path, params: dict[str, object]) -> dict[str, object]:
    net1, error = _required_folder(params, "netlist1", "网表1文件夹")
    if error:
        return _error("netlist_compare", error)
    net2, error = _required_folder(params, "netlist2", "网表2文件夹")
    if error:
        return _error("netlist_compare", error)
    nets1 = _parse_net_file(net1)
    nets2 = _parse_net_file(net2)
    parts1, part_warning1 = _parse_part_file_optional(net1)
    parts2, part_warning2 = _parse_part_file_optional(net2)
    warnings = [warning for warning in [part_warning1, part_warning2] if warning]

    rows = []
    package_rows = []
    compare_items: list[dict[str, object]] = []

    def net_side(refs: str, nodes: str) -> dict[str, object]:
        return {"位号": refs, "节点": nodes}

    for name in sorted(set(nets1) | set(nets2)):
        in1, in2 = name in nets1, name in nets2
        left_refs = _natural_join(nets1.get(name, {}).get("refs", []))
        right_refs = _natural_join(nets2.get(name, {}).get("refs", []))
        left_nodes = _natural_join(nets1.get(name, {}).get("nodes", []))
        right_nodes = _natural_join(nets2.get(name, {}).get("nodes", []))
        if not in1:
            status, kind, diff = "仅网表2存在", "only_right", ["位号", "节点"]
        elif not in2:
            status, kind, diff = "仅网表1存在", "only_left", ["位号", "节点"]
        elif left_nodes != right_nodes:
            status, kind, diff = "网络节点差异", "diff", ["节点"]
            if left_refs != right_refs:
                diff.insert(0, "位号")
        else:
            status, kind, diff = "一致", "same", []
        rows.append([name, left_refs, left_nodes, right_nodes, right_refs, status])
        compare_items.append(
            {
                "key": name,
                "status": status,
                "kind": kind,
                "left": net_side(left_refs, left_nodes) if in1 else None,
                "right": net_side(right_refs, right_nodes) if in2 else None,
                "diff": diff,
            }
        )

    for ref in sorted(set(parts1) | set(parts2)):
        if parts1.get(ref) != parts2.get(ref):
            in1, in2 = ref in parts1, ref in parts2
            if not in1:
                kind = "only_right"
            elif not in2:
                kind = "only_left"
            else:
                kind = "diff"
            package_rows.append([ref, parts1.get(ref, ""), parts2.get(ref, ""), "封装差异"])
            compare_items.append(
                {
                    "key": f"器件:{ref}",
                    "status": "封装差异",
                    "kind": kind,
                    "left": {"封装": parts1[ref]} if in1 else None,
                    "right": {"封装": parts2[ref]} if in2 else None,
                    "diff": ["封装"] if kind == "diff" else [],
                }
            )

    headers = ["网络", "网表1位号", "网表1节点", "网表2节点", "网表2位号", "状态"]
    package_headers = ["位号", "网表1封装", "网表2封装", "状态"]
    output = _output_dir(params, root, "netlist") / f"网表比较差异_{_timestamp()}.xlsx"
    _write_sheets(output, [("网络节点差异", headers, rows), ("器件封装差异", package_headers, package_rows)])
    diff_count = sum(1 for row in rows if row[-1] != "一致")
    table = _table(headers, rows, status_col=5, diff_pairs=[[2, 3]])
    package_table = _table(package_headers, package_rows, status_col=3, diff_pairs=[[1, 2]])
    compare = _compare("对象", "网表1", "网表2", ["节点", "封装"], compare_items)
    review = _build_netlist_review(nets1, nets2, package_rows)
    result = _result(
        "netlist_compare",
        [output],
        {
            "diff_count": diff_count + len(package_rows),
            "node_diffs": diff_count,
            "package_diffs": len(package_rows),
            "package_check": "skipped" if warnings else "ok",
            "critical_changes": sum(1 for item in review["items"] if item.get("critical") and item.get("status") != "一致"),
        },
        table,
        compare,
    )
    result["package_table"] = package_table
    result["netlist_review"] = review
    result["warnings"] = warnings
    return result


def run_smt_package_check(root: Path, params: dict[str, object]) -> dict[str, object]:
    netlist, error = _required_folder(params, "netlist", "网表文件夹")
    if error:
        return _error("smt_package_check", error)
    bom, error = _required_file(params, "bom", "BOM 文件")
    if error:
        return _error("smt_package_check", error)
    parts = _parse_part_file(netlist)
    bom_rows = _read_bom_rows(bom)
    review = _build_smt_package_review(parts, bom_rows)
    rows = review["table_rows"]
    headers = ["位号", "网表封装", "BOM封装/型号", "描述", "名称", "状态", "说明"]
    output = _output_dir(params, root, "smt") / f"贴片封装检查结果_{_timestamp()}.xlsx"
    _write_table(output, "封装检查", headers, rows)
    counts = review["status_counts"]
    manual_count = counts["manual"] + counts["missing_bom"] + counts["extra_bom"] + counts["multi_package"]
    passed_count = counts["passed"] + counts["near"]
    table = _table(headers, rows, status_col=5)
    result = _result(
        "smt_package_check",
        [output],
        {
            "total": len(parts),
            "passed_count": passed_count,
            "near_count": counts["near"],
            "manual_count": manual_count,
            "missing_bom": counts["missing_bom"],
            "extra_bom": counts["extra_bom"],
            "multi_package": counts["multi_package"],
            "high_risk": counts["high_risk"],
        },
        table,
    )
    result["smt_package_review"] = {key: value for key, value in review.items() if key != "table_rows"}
    return result


def run_single_network_check(root: Path, params: dict[str, object]) -> dict[str, object]:
    netlist, error = _required_folder(params, "netlist", "网表文件夹")
    if error:
        return _error("single_network_check", error)
    nets = _parse_net_file(netlist)
    rows = []
    for name, data in sorted(nets.items()):
        refs = data.get("refs", [])
        nodes = data.get("nodes", [])
        is_nc = name.upper().startswith("NC")
        is_single_ref = len(refs) == 1
        if is_nc or is_single_ref:
            if is_nc and is_single_ref:
                kind = "NC网络/单一位号网络"
            elif is_nc:
                kind = "NC网络"
            else:
                kind = "单一位号网络"
            rows.append([name, kind, _natural_join(refs), _natural_join(nodes), len(refs), len(nodes)])
    headers = ["网络", "类型", "位号", "节点/Pin", "位号数", "节点数"]
    output = _output_dir(params, root, "netlist") / f"单网络检查结果_{_timestamp()}.xlsx"
    _write_table(output, "单网络检查", headers, rows)
    table = _table(headers, rows, status_col=1)
    nc_count = sum(1 for row in rows if "NC网络" in row[1])
    single_ref_count = sum(1 for row in rows if "单一位号网络" in row[1])
    return _result("single_network_check", [output], {"matched_count": len(rows), "nc_count": nc_count, "single_ref_count": single_ref_count}, table)


def run_bom_process(root: Path, params: dict[str, object]) -> dict[str, object]:
    from app.backend.tools import bom_process

    source, error = _required_file(params, "source_bom", "原始 BOM 文件")
    if error:
        return _error("bom_process", error)
    formats = params.get("formats") or ["plm", "oa"]
    if isinstance(formats, str):
        formats = [formats]
    formats = [f for f in formats if f in ("plm", "oa")] or ["plm"]
    extras = params.get("extras") if isinstance(params.get("extras"), list) else []
    out_dir = _output_dir(params, root, "bom")
    template = next(root.rglob("203010100819_ERP_BOM导入模板.xlsx"), None)
    source_rows_for_checks, _ = bom_process.load_source(source, include_shields=True)
    shield_candidates = bom_process.detect_shield_candidates(source_rows_for_checks)
    if shield_candidates and "confirm_shields" not in params:
        return {
            "status": "needs_confirmation",
            "tool": "bom_process",
            "reason": "shield_bracket_candidates",
            "message": "发现 SH 位号物料，疑似屏蔽支架。请确认是否作为结构件进入最终 BOM。",
            "shield_count": len(shield_candidates),
            "shield_candidates": shield_candidates,
            "summary": {"shield_candidates": len(shield_candidates)},
        }
    confirm_shields = bool(params.get("confirm_shields"))
    source_rows, _ = bom_process.load_source(source, include_shields=confirm_shields)
    conflicts = bom_process.conflict_summary(source_rows)
    if conflicts and "merge_conflicts" not in params:
        return {
            "status": "needs_confirmation",
            "tool": "bom_process",
            "reason": "part_property_conflicts",
            "message": "发现相同物料编码存在不同型号/描述/名称/等级，请确认是否按编码合并。",
            "conflict_count": len(conflicts),
            "conflicts": conflicts,
            "summary": {"conflicts": len(conflicts)},
        }
    merge_conflicts = bool(params.get("merge_conflicts"))
    conflict_choices = params.get("conflict_choices") if isinstance(params.get("conflict_choices"), dict) else {}
    result = bom_process.process(
        source,
        formats,
        str(params.get("parent_code") or ""),
        str(params.get("parent_desc") or ""),
        str(params.get("name") or ""),
        extras,
        out_dir,
        _timestamp(),
        template,
        merge_conflicts,
        conflict_choices,
        confirm_shields,
    )
    outputs = [str(p) for p in result["outputs"]] + [str(result["nc_summary"])]
    preview_rows = [
        [r["code"], r["model"], r["desc"], r["qty"], ",".join(r["refs"]), r["grade"]]
        for r in result["records"]
    ]
    return {
        "status": "ok",
        "tool": "bom_process",
        "outputs": outputs,
        "summary": result["summary"],
        "preview": {"headers": ["编号", "型号", "描述", "数量", "位号", "等级"], "rows": preview_rows},
        "process_file": str(result["outputs"][0]) if result["outputs"] else "",
        "next_step": "成品 BOM 已生成，请核对后走 OA YF25 备料 / PLM 导入。",
    }


def create_analysis_tools(root: Path) -> list[Tool]:
    return [
        Tool("bom_process", "BOM 处理", "Capture 原始 BOM → 可导入 PLM/OA 成品（过滤、合并、可加 PCB/屏蔽支架等附加物料）。", "available", "BOM", lambda params=None: run_bom_process(root, params or {})),
        Tool("bom_compare", "BOM 差异比较", "比较两个 BOM Excel，输出位号、编号、描述、数量差异报告。", "available", "BOM", lambda params=None: run_bom_compare(root, params or {})),
        Tool("bom_risk_check", "BOM 风险检查", "单份 BOM 导入前体检：PCB/屏蔽支架/NC 未贴/机构件/测试点/重复位号/数量一致性/eMMC-DDR 版本提醒。", "available", "BOM", lambda params=None: run_bom_risk_check(root, params or {})),
        Tool("netlist_compare", "网表差异比较", "比较两个网表文件夹中的网络节点和器件信息。", "available", "Netlist", lambda params=None: run_netlist_compare(root, params or {})),
        Tool("smt_package_check", "贴片封装检查", "选择 Allegro 目录和已处理后的 PLM/OA 成品 BOM，检查网表封装与 BOM 型号/描述的一致性。", "available", "SMT", lambda params=None: run_smt_package_check(root, params or {})),
        Tool("single_network_check", "单网络检查", "提取 NC 网络和只有单一位号的网络，辅助硬件检查。", "available", "Netlist", lambda params=None: run_single_network_check(root, params or {})),
    ]

