from __future__ import annotations

from contextlib import contextmanager
from pathlib import Path
from typing import Iterator

from openpyxl import load_workbook


def build_merged_cell_lookup(worksheet: object) -> dict[tuple[int, int], object]:
    """Map non-anchor merged-cell coordinates to their shared anchor value."""
    lookup: dict[tuple[int, int], object] = {}
    for merged_range in worksheet.merged_cells.ranges:
        anchor_value = worksheet.cell(merged_range.min_row, merged_range.min_col).value
        for row in range(merged_range.min_row, merged_range.max_row + 1):
            for col in range(merged_range.min_col, merged_range.max_col + 1):
                if (row, col) != (merged_range.min_row, merged_range.min_col):
                    lookup[(row, col)] = anchor_value
    return lookup


@contextmanager
def open_bom_workbook(path: Path, **kwargs: object) -> Iterator[object]:
    workbook = load_workbook(path, **kwargs)
    try:
        yield workbook
    finally:
        workbook.close()
