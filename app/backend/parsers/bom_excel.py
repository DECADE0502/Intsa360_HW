from __future__ import annotations

import re
from pathlib import Path
from typing import Iterable


REF_SPLIT_RE = re.compile(r"[,;\s]+")

FIELD_ALIASES = {
    "reference": ["位号", "位置号", "Reference", "Ref", "RefDes", "Designator", "BOM1位号", "BOM2位号", "浣嶅彿", "浣嶇疆鍙?"],
    "part_number": ["编号", "料号", "物料编码", "子项编码", "编码子", "编码（子）", "编码(子)", "Part Number", "PN", "缂栧彿", "鏂欏彿", "鐗╂枡缂栫爜", "瀛愰」缂栫爜"],
    "description": ["描述", "物料描述", "子项描述", "描述子", "描述（子）", "描述(子)", "器件描述", "器件描述（新整理）", "Description", "鎻忚堪", "鐗╂枡鎻忚堪", "瀛愰」鎻忚堪", "鍣ㄤ欢鎻忚堪", "鍣ㄤ欢鎻忚堪锛堟柊鏁寸悊锛?"],
    "quantity": ["数量", "用量", "Qty", "Quantity", "鏁伴噺", "鐢ㄩ噺"],
    "name": ["名称", "物料名称", "Part Type", "鍚嶇О", "鐗╂枡鍚嶇О"],
    "package": ["封装名", "PCB封装", "PCB Footprint", "Package", "Footprint", "灏佽鍚?", "PCB灏佽"],
    "value": ["Value", "值", "鍊?"],
    "model": ["型号", "规格型号", "规格", "Model", "MPN", "鍨嬪彿", "瑙勬牸鍨嬪彿", "瑙勬牸"],
    "grade": ["物料优选等级", "优选等级", "物料等级", "等级", "鐗╂枡浼橀€夌瓑绾?", "浼橀€夌瓑绾?", "鐗╂枡绛夌骇", "绛夌骇"],
}


def normalize_header(value: object) -> str:
    text = str(value or "").strip()
    text = re.sub(r"##.*$", "", text)
    text = text.replace("*", "")
    text = text.replace(" ", "")
    return text


def find_header(ws, required: Iterable[str], scan_rows: int = 30) -> tuple[int, dict[str, int]]:
    required = list(required)
    best_row = 1
    best_map: dict[str, int] = {}
    for row in range(1, min(ws.max_row, scan_rows) + 1):
        values = [str(ws.cell(row, col).value or "").strip() for col in range(1, ws.max_column + 1)]
        normalized_values = [normalize_header(value) for value in values]
        mapping: dict[str, int] = {}
        for canonical, aliases in FIELD_ALIASES.items():
            normalized_aliases = {normalize_header(alias) for alias in aliases}
            for idx, value in enumerate(normalized_values, start=1):
                if value in normalized_aliases:
                    mapping[canonical] = idx
                    break
        score = sum(1 for key in required if key in mapping)
        if score > sum(1 for key in required if key in best_map):
            best_row = row
            best_map = mapping
    missing = [key for key in required if key not in best_map]
    if missing:
        raise ValueError(f"表头识别失败，缺少字段: {', '.join(missing)}")
    return best_row, best_map


def _choose_best_column(
    normalized_values: list[str],
    aliases: Iterable[str],
    anchor_col: int | None = None,
    prefer_after_anchor: bool = False,
) -> int | None:
    normalized_aliases = {normalize_header(alias) for alias in aliases}
    matches = [idx for idx, value in enumerate(normalized_values, start=1) if value in normalized_aliases]
    if not matches:
        return None
    if anchor_col is not None and prefer_after_anchor:
        after_anchor = [idx for idx in matches if idx > anchor_col]
        if after_anchor:
            return after_anchor[0]
    return matches[0]


def refine_bom_mapping(ws, header_row: int, mapping: dict[str, int]) -> dict[str, int]:
    values = [str(ws.cell(header_row, col).value or "").strip() for col in range(1, ws.max_column + 1)]
    normalized_values = [normalize_header(value) for value in values]
    refined = dict(mapping)

    part_col = _choose_best_column(normalized_values, FIELD_ALIASES["part_number"])
    if part_col:
        refined["part_number"] = part_col

    for key in ["description", "quantity", "name", "package", "value", "model", "grade", "reference"]:
        col = _choose_best_column(
            normalized_values,
            FIELD_ALIASES[key],
            anchor_col=refined.get("part_number"),
            prefer_after_anchor=key in {"description", "quantity", "name", "package", "value", "reference"},
        )
        if col:
            refined[key] = col
    return refined


def split_refs(value: object) -> list[str]:
    return [part for part in REF_SPLIT_RE.split(str(value or "").strip()) if part]


def read_bom_rows(path: Path, require_refs: bool = True) -> list[dict[str, object]]:
    from openpyxl import load_workbook

    wb = load_workbook(path, data_only=True)
    ws = wb.active
    header_row, mapping = find_header(ws, ["reference", "part_number", "description", "quantity"])
    mapping = refine_bom_mapping(ws, header_row, mapping)
    rows: list[dict[str, object]] = []
    for row in range(header_row + 1, ws.max_row + 1):
        refs = split_refs(ws.cell(row, mapping["reference"]).value)
        part_number = str(ws.cell(row, mapping["part_number"]).value or "").strip()
        if require_refs and not refs:
            continue
        if not refs and not part_number:
            continue
        rows.append(
            {
                "source_row": row,
                "reference": ",".join(refs),
                "refs": refs,
                "part_number": part_number,
                "model": str(ws.cell(row, mapping["model"]).value or "").strip() if "model" in mapping else "",
                "grade": str(ws.cell(row, mapping["grade"]).value or "").strip() if "grade" in mapping else "",
                "description": str(ws.cell(row, mapping["description"]).value or "").strip(),
                "quantity": ws.cell(row, mapping["quantity"]).value,
                "name": str(ws.cell(row, mapping.get("name", mapping["description"])).value or "").strip(),
                "package": str(ws.cell(row, mapping.get("package", mapping["description"])).value or "").strip(),
                "value": str(ws.cell(row, mapping.get("value", mapping["description"])).value or "").strip(),
            }
        )
    return rows
