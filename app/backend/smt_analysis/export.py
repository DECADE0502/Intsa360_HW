from __future__ import annotations

import hashlib
import json
import shutil
import zipfile
from datetime import datetime, timezone
from pathlib import Path
from typing import Callable, Iterable, Sequence
from uuid import uuid4

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.backend.contracts.smt_analysis import (
    SmtAnalysisRunResponse,
    SmtPlacement,
)


_STATE_LABELS = {
    "installed": "已装机",
    "confirmed_nc": "确认 NC",
    "candidate_nc": "候选 NC",
    "non_smt": "非 SMT",
    "bom_only": "BOM 有 / 坐标无",
    "coordinate_only": "坐标有 / BOM 无",
    "conflicting": "数据冲突",
    "unresolved": "待确认",
}
_ROLE_LABELS = {
    "smt_component": "SMT 器件",
    "tht_component": "插件器件",
    "manual_assembly": "手工装配",
    "fiducial": "光学定位点",
    "tooling_hole": "工艺孔",
    "mounting_hole": "安装孔",
    "test_point": "测试点",
    "mechanical": "机械件",
    "panel_object": "拼板对象",
    "unknown": "未知",
}
_SIDE_LABELS = {"top": "正面", "bottom": "背面", "unknown": "未知"}
_STATE_COLORS = {
    "installed": "#2f855a",
    "confirmed_nc": "#d4380d",
    "candidate_nc": "#d48806",
    "non_smt": "#595959",
    "bom_only": "#722ed1",
    "coordinate_only": "#1677ff",
    "conflicting": "#cf1322",
    "unresolved": "#ad6800",
}


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        while chunk := handle.read(1024 * 1024):
            digest.update(chunk)
    return digest.hexdigest()


def _snapshot_fingerprint(snapshot: SmtAnalysisRunResponse) -> str:
    payload = json.dumps(
        snapshot.model_dump(mode="json"),
        ensure_ascii=False,
        sort_keys=True,
        separators=(",", ":"),
        allow_nan=False,
    )
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()


def _text(value: object) -> str:
    return "" if value is None else str(value)


def _write_text_cell(cell, value: object) -> None:
    cell.value = _text(value)
    cell.data_type = "s"


def _append_row(sheet, values: Sequence[object]) -> None:
    row_index = (
        1
        if sheet.max_row == 1 and sheet.max_column == 1 and sheet["A1"].value is None
        else sheet.max_row + 1
    )
    for column_index, value in enumerate(values, start=1):
        cell = sheet.cell(row=row_index, column=column_index)
        if isinstance(value, (int, float)) and not isinstance(value, bool):
            cell.value = value
        else:
            _write_text_cell(cell, value)


def _style_table(sheet, *, freeze: str = "A2") -> None:
    sheet.freeze_panes = freeze
    sheet.auto_filter.ref = sheet.dimensions
    header_fill = PatternFill("solid", fgColor="E8F3F1")
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="163A35")
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center")
    for column in range(1, sheet.max_column + 1):
        values = [
            len(_text(sheet.cell(row=row, column=column).value))
            for row in range(1, min(sheet.max_row, 300) + 1)
        ]
        width = min(42, max([10, *values]) + 2)
        sheet.column_dimensions[get_column_letter(column)].width = width


def _primary_material(placement: SmtPlacement):
    if placement.bom_requirement is None:
        return None
    return next(
        (
            item
            for item in placement.bom_requirement.materials
            if item.is_primary
        ),
        placement.bom_requirement.materials[0]
        if placement.bom_requirement.materials
        else None,
    )


def _evidence_text(placement: SmtPlacement) -> str:
    return "；".join(
        f"{item.weight}:{item.message}"
        for item in placement.evidence_chain
    )


def _placement_row(placement: SmtPlacement) -> list[object]:
    material = _primary_material(placement)
    decision = placement.decision
    return [
        placement.ref,
        _SIDE_LABELS[placement.side],
        _ROLE_LABELS[placement.role],
        _STATE_LABELS[placement.assembly_state],
        material.part_number if material else "",
        material.model if material else "",
        material.description if material else "",
        material.grade if material else "",
        placement.bom_requirement.quantity
        if placement.bom_requirement
        else None,
        placement.image_x,
        placement.image_y,
        "是" if placement.netlist_present is True else (
            "否" if placement.netlist_present is False else "未知"
        ),
        "是" if placement.drawing_present is True else (
            "否" if placement.drawing_present is False else "未知"
        ),
        decision.source if decision else "rule",
        decision.reason if decision else "",
        "；".join(placement.blocking_reasons),
        _evidence_text(placement),
    ]


def _build_workbook(snapshot: SmtAnalysisRunResponse, destination: Path) -> None:
    workbook = Workbook()
    summary = workbook.active
    summary.title = "审查摘要"
    _append_row(summary, ["项目", "结果"])
    for label, value in (
        ("运行 ID", snapshot.run_id),
        ("规则版本", snapshot.rule_version),
        ("解析版本", snapshot.parser_version),
        ("来源指纹", snapshot.source_fingerprint),
        ("物理位号总数", snapshot.summary.placement_count),
        ("确认装机", snapshot.summary.installed_count),
        ("确认 NC", snapshot.summary.confirmed_nc_count),
        ("候选 NC", snapshot.summary.candidate_nc_count),
        ("待解决", snapshot.summary.unresolved_count),
        ("阻断项", snapshot.summary.blocking_count),
        ("完成时间", snapshot.updated_at.isoformat()),
    ):
        _append_row(summary, [label, value])
    _style_table(summary)

    detail = workbook.create_sheet("位号明细")
    headers = [
        "位号",
        "面别",
        "角色",
        "装机状态",
        "主料编码",
        "型号",
        "描述",
        "等级",
        "BOM 数量",
        "图像 X",
        "图像 Y",
        "网表存在",
        "位号图存在",
        "判定来源",
        "人工说明",
        "阻断原因",
        "证据链",
    ]
    _append_row(detail, headers)
    for placement in sorted(snapshot.placements, key=lambda item: item.ref):
        _append_row(detail, _placement_row(placement))
    _style_table(detail)

    fai = workbook.create_sheet("首件核对表")
    _append_row(
        fai,
        [
            "位号",
            "面别",
            "主料编码",
            "型号",
            "描述",
            "封装/坐标证据",
            "核对结果",
            "备注",
        ],
    )
    for placement in sorted(snapshot.placements, key=lambda item: item.ref):
        if placement.assembly_state != "installed":
            continue
        material = _primary_material(placement)
        _append_row(
            fai,
            [
                placement.ref,
                _SIDE_LABELS[placement.side],
                material.part_number if material else "",
                material.model if material else "",
                material.description if material else "",
                _evidence_text(placement),
                "",
                placement.decision.reason if placement.decision else "",
            ],
        )
    _style_table(fai)

    nc_sheet = workbook.create_sheet("NC 与非 SMT")
    _append_row(
        nc_sheet,
        ["位号", "面别", "状态", "角色", "主料编码", "判定来源", "说明"],
    )
    for placement in sorted(snapshot.placements, key=lambda item: item.ref):
        if placement.assembly_state not in {"confirmed_nc", "non_smt"}:
            continue
        material = _primary_material(placement)
        _append_row(
            nc_sheet,
            [
                placement.ref,
                _SIDE_LABELS[placement.side],
                _STATE_LABELS[placement.assembly_state],
                _ROLE_LABELS[placement.role],
                material.part_number if material else "",
                placement.decision.source if placement.decision else "rule",
                placement.decision.reason if placement.decision else "",
            ],
        )
    _style_table(nc_sheet)

    unresolved = workbook.create_sheet("未解决项")
    _append_row(unresolved, headers)
    for placement in sorted(snapshot.placements, key=lambda item: item.ref):
        if placement.assembly_state in {
            "candidate_nc",
            "bom_only",
            "coordinate_only",
            "conflicting",
            "unresolved",
        }:
            _append_row(unresolved, _placement_row(placement))
    _style_table(unresolved)

    workbook.save(destination)
    workbook.close()


def _annotate_page(
    *,
    source: Path,
    destination: Path,
    placements: Iterable[SmtPlacement],
) -> None:
    try:
        from PIL import Image, ImageColor, ImageDraw
    except ImportError as exc:
        raise RuntimeError("当前运行时缺少本地图片处理组件") from exc

    with Image.open(source) as original:
        image = original.convert("RGB")
        draw = ImageDraw.Draw(image)
        radius = max(4, min(12, round(max(image.width, image.height) / 180)))
        for placement in placements:
            if placement.image_x is None or placement.image_y is None:
                continue
            x = float(placement.image_x)
            y = float(placement.image_y)
            color = ImageColor.getrgb(
                _STATE_COLORS.get(placement.assembly_state, "#595959")
            )
            draw.ellipse(
                (x - radius, y - radius, x + radius, y + radius),
                outline=color,
                width=max(2, radius // 3),
            )
            draw.text(
                (x + radius + 2, y - radius),
                placement.ref,
                fill=color,
                stroke_width=1,
                stroke_fill="white",
            )
        image.save(destination, format="PNG", optimize=True)
        image.close()


def _verify_outputs(
    *,
    snapshot: SmtAnalysisRunResponse,
    folder: Path,
    workbook_name: str,
    snapshot_name: str,
    image_names: Sequence[str],
    archive_name: str,
) -> None:
    expected_refs = sorted(item.ref for item in snapshot.placements)

    restored = SmtAnalysisRunResponse.model_validate_json(
        (folder / snapshot_name).read_text(encoding="utf-8")
    )
    if sorted(item.ref for item in restored.placements) != expected_refs:
        raise ValueError("SMT 快照回读后的物理位号集合不一致")
    if restored.summary != snapshot.summary:
        raise ValueError("SMT 快照回读后的汇总不一致")

    workbook = load_workbook(folder / workbook_name, read_only=True, data_only=True)
    try:
        detail = workbook["位号明细"]
        workbook_refs = sorted(
            str(value[0]).strip()
            for value in detail.iter_rows(min_row=2, values_only=True)
            if value and value[0]
        )
        if workbook_refs != expected_refs:
            raise ValueError("SMT Excel 回读后的物理位号集合不一致")
        unresolved = workbook["未解决项"]
        unresolved_count = sum(
            1
            for row in unresolved.iter_rows(min_row=2, values_only=True)
            if row and row[0]
        )
        if unresolved_count != snapshot.summary.unresolved_count:
            raise ValueError("SMT Excel 回读后的未解决项数量不一致")
    finally:
        workbook.close()

    try:
        from PIL import Image
    except ImportError as exc:
        raise RuntimeError("当前运行时缺少本地图片处理组件") from exc
    for image_name in image_names:
        with Image.open(folder / image_name) as image:
            image.verify()

    expected_members = {
        workbook_name,
        snapshot_name,
        *image_names,
    }
    with zipfile.ZipFile(folder / archive_name) as archive:
        if set(archive.namelist()) != expected_members:
            raise ValueError("SMT 交付压缩包内容不完整")
        bad_member = archive.testzip()
        if bad_member is not None:
            raise ValueError(f"SMT 交付压缩包文件损坏：{bad_member}")


def export_analysis(
    *,
    snapshot: SmtAnalysisRunResponse,
    outputs_root: Path,
    page_resolver: Callable[[str], tuple[Path, str]],
) -> dict[str, object]:
    if snapshot.summary.unresolved_count:
        raise ValueError("仍有待解决位号，不能生成 SMT 交付包")
    if snapshot.state != "deliver":
        raise ValueError("请先完成 SMT 装配复核")

    outputs = Path(outputs_root).resolve()
    outputs.mkdir(parents=True, exist_ok=True)
    fingerprint = _snapshot_fingerprint(snapshot)
    folder_name = f"SMT装配审查_{snapshot.run_id[:8]}_{fingerprint[:10]}"
    target = (outputs / folder_name).resolve()
    try:
        target.relative_to(outputs)
    except ValueError as exc:
        raise ValueError("SMT 导出目录越界") from exc

    workbook_name = "SMT装配审查报告.xlsx"
    snapshot_name = "SMT装配审查快照.json"
    archive_name = "SMT装配审查交付包.zip"
    temporary = outputs / f".{folder_name}.{uuid4().hex}.tmp"
    temporary.mkdir(parents=True)
    image_names: list[str] = []
    try:
        snapshot_path = temporary / snapshot_name
        snapshot_path.write_text(
            json.dumps(
                snapshot.model_dump(mode="json"),
                ensure_ascii=False,
                sort_keys=True,
                indent=2,
                allow_nan=False,
            ),
            encoding="utf-8",
        )
        _build_workbook(snapshot, temporary / workbook_name)

        for side, label in (("top", "正面"), ("bottom", "背面")):
            registration = next(
                (
                    item
                    for item in snapshot.registrations
                    if item.side == side
                    and item.confidence_state == "verified"
                ),
                None,
            )
            if registration is None:
                continue
            source, _ = page_resolver(registration.page_id)
            image_name = f"{label}位号标注图.png"
            _annotate_page(
                source=source,
                destination=temporary / image_name,
                placements=(
                    item
                    for item in snapshot.placements
                    if item.side in {side, "unknown"}
                ),
            )
            image_names.append(image_name)

        with zipfile.ZipFile(
            temporary / archive_name,
            "w",
            compression=zipfile.ZIP_DEFLATED,
        ) as archive:
            for name in (workbook_name, snapshot_name, *image_names):
                archive.write(temporary / name, arcname=name)

        _verify_outputs(
            snapshot=snapshot,
            folder=temporary,
            workbook_name=workbook_name,
            snapshot_name=snapshot_name,
            image_names=image_names,
            archive_name=archive_name,
        )
        if target.exists():
            _verify_outputs(
                snapshot=snapshot,
                folder=target,
                workbook_name=workbook_name,
                snapshot_name=snapshot_name,
                image_names=image_names,
                archive_name=archive_name,
            )
            shutil.rmtree(temporary)
        else:
            temporary.replace(target)
    except Exception:
        shutil.rmtree(temporary, ignore_errors=True)
        raise

    artifact_names = [workbook_name, snapshot_name, *image_names]
    artifacts = []
    for name in artifact_names:
        path = target / name
        artifacts.append(
            {
                "label": Path(name).stem,
                "path": path.relative_to(outputs).as_posix(),
                "media_type": (
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    if path.suffix.lower() == ".xlsx"
                    else "application/json"
                    if path.suffix.lower() == ".json"
                    else "image/png"
                ),
                "size": path.stat().st_size,
                "sha256": _sha256(path),
            }
        )
    archive_path = target / archive_name
    return {
        "status": "ok",
        "run_id": snapshot.run_id,
        "snapshot_fingerprint": fingerprint,
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "package_path": archive_path.relative_to(outputs).as_posix(),
        "package_sha256": _sha256(archive_path),
        "artifacts": artifacts,
    }
