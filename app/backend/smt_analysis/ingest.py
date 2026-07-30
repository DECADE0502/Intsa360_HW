from __future__ import annotations

import hashlib
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook

from app.backend.contracts.smt_analysis import SmtSourceAsset
from app.backend.smt_analysis.classifiers import classify_source, sniff_media_type


_HASH_CHUNK = 1024 * 1024
_SAMPLE_LIMIT = 256 * 1024
_PDF_TEXT_PAGE_LIMIT = 4


def _sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        while chunk := handle.read(_HASH_CHUNK):
            digest.update(chunk)
    return digest.hexdigest()


def source_fingerprint(assets: Iterable[SmtSourceAsset]) -> str:
    digest = hashlib.sha256()
    for asset in sorted(assets, key=lambda item: item.relative_path.casefold()):
        digest.update(asset.relative_path.replace("\\", "/").encode("utf-8"))
        digest.update(b"\0")
        digest.update(asset.sha256.encode("ascii"))
        digest.update(b"\0")
    return digest.hexdigest()


def _stable_asset_id(relative_path: str, sha256: str) -> str:
    path_hash = hashlib.sha256(relative_path.casefold().encode("utf-8")).hexdigest()[:12]
    return f"src-{sha256[:16]}-{path_hash}"


def _pdf_probe(path: Path) -> tuple[int | None, str]:
    try:
        import pypdfium2 as pdfium  # type: ignore[import-not-found]
    except ImportError:
        return None, ""

    document = None
    text_parts: list[str] = []
    try:
        document = pdfium.PdfDocument(path)
        page_count = len(document)
        for page_number in range(min(page_count, _PDF_TEXT_PAGE_LIMIT)):
            page = document[page_number]
            text_page = None
            try:
                text_page = page.get_textpage()
                text_parts.append(text_page.get_text_range())
            finally:
                if text_page is not None:
                    text_page.close()
                page.close()
        return page_count, "\n".join(text_parts)
    except Exception:  # noqa: BLE001 - a broken PDF is a quality issue, not a scan failure
        return None, ""
    finally:
        if document is not None:
            document.close()


def _xlsx_probe(path: Path) -> tuple[list[str], list[str]]:
    workbook = None
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
        sheet_names = list(workbook.sheetnames)
        headers: list[str] = []
        for sheet in workbook.worksheets:
            for row in sheet.iter_rows(min_row=1, max_row=12, values_only=True):
                populated = [str(value).strip() for value in row if value not in (None, "")]
                if populated:
                    headers.extend(populated[:80])
                    break
        return sheet_names, headers
    except Exception:  # noqa: BLE001 - keep damaged/unsupported workbooks visible
        return [], []
    finally:
        if workbook is not None:
            workbook.close()


def scan_source_directory(folder: Path) -> list[SmtSourceAsset]:
    root = Path(folder).resolve()
    if not root.is_dir():
        raise ValueError(f"SMT 资料目录不存在：{folder}")

    assets: list[SmtSourceAsset] = []
    for path in sorted(root.rglob("*"), key=lambda item: item.as_posix().casefold()):
        if path.is_symlink() or not path.is_file():
            continue
        relative = path.relative_to(root).as_posix()
        with path.open("rb") as handle:
            prefix = handle.read(_SAMPLE_LIMIT)
        media_type = sniff_media_type(path, prefix)
        page_count: int | None = None
        extracted_text = ""
        sheet_names: list[str] = []
        sheet_headers: list[str] = []
        if media_type == "application/pdf":
            page_count, extracted_text = _pdf_probe(path)
        elif path.suffix.casefold() == ".xlsx":
            sheet_names, sheet_headers = _xlsx_probe(path)
        classification = classify_source(
            path,
            media_type=media_type,
            prefix=prefix,
            extracted_text=extracted_text,
            page_count=page_count,
            sheet_headers=sheet_headers,
        )
        sha256 = _sha256_file(path)
        assets.append(
            SmtSourceAsset(
                asset_id=_stable_asset_id(relative, sha256),
                relative_path=relative,
                sha256=sha256,
                media_type=classification.media_type,
                file_size=path.stat().st_size,
                roles=list(classification.roles),
                classification_state=classification.state,
                evidence=list(classification.evidence),
                page_count=page_count,
                sheet_names=sheet_names,
            )
        )
    return assets
