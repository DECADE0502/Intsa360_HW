from __future__ import annotations

import csv
import hashlib
import io
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable, Sequence

from openpyxl import load_workbook

from app.backend.contracts.smt_analysis import (
    BoardSide,
    SmtCoordinateOccurrence,
    SmtCoordinateQuality,
    SmtCoordinateSet,
    SmtQualityIssue,
)
from app.backend.smt_analysis.coordinates.base import CoordinateProbe


_HEADER_ALIASES = {
    "ref": {"ref", "reference", "refdes", "designator", "part reference", "位号"},
    "x": {"x", "x coordinate", "x-coordinate", "x坐标", "center-x", "center x"},
    "y": {"y", "y coordinate", "y-coordinate", "y坐标", "center-y", "center y"},
    "side": {"side", "layer", "board side", "面", "面别", "板面"},
    "rotation": {"rotation", "angle", "rot", "旋转", "角度"},
    "footprint": {"footprint", "package", "pattern", "封装", "pcb footprint"},
}
_UNIT_RE = re.compile(r"(?:\(|\[|\b)(mm|mil|mils|inch|in)(?:\)|\]|\b)", re.IGNORECASE)


@dataclass(frozen=True)
class _Table:
    section: str
    rows: tuple[tuple[object, ...], ...]


def _normalize_header(value: object) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip().casefold())


def _field_for_header(value: object) -> str | None:
    normalized = _normalize_header(value)
    without_unit = _UNIT_RE.sub("", normalized).strip(" ()[]")
    for field, aliases in _HEADER_ALIASES.items():
        if normalized in aliases or without_unit in aliases:
            return field
    return None


def _decode_text(path: Path) -> str:
    raw = path.read_bytes()
    for encoding in ("utf-8-sig", "utf-16", "gb18030"):
        try:
            text = raw.decode(encoding)
        except UnicodeError:
            continue
        if "\x00" not in text or encoding == "utf-16":
            return text
    raise ValueError("坐标文本编码无法识别")


def _text_table(path: Path) -> list[_Table]:
    text = _decode_text(path)
    sample = text[:65536]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
        delimiter = dialect.delimiter
    except csv.Error:
        candidates = {delimiter: sample.count(delimiter) for delimiter in (",", ";", "\t", "|")}
        delimiter = max(candidates, key=candidates.get)
        if candidates[delimiter] == 0:
            return []
    rows = tuple(tuple(cell.strip() for cell in row) for row in csv.reader(io.StringIO(text), delimiter=delimiter))
    return [_Table(section="", rows=rows)]


def _excel_tables(path: Path) -> list[_Table]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        return [
            _Table(
                section=sheet.title,
                rows=tuple(
                    tuple(value for value in row)
                    for row in sheet.iter_rows(values_only=True)
                    if any(value not in (None, "") for value in row)
                ),
            )
            for sheet in workbook.worksheets
        ]
    finally:
        workbook.close()


def _tables(path: Path) -> list[_Table]:
    if path.suffix.casefold() == ".xlsx":
        return _excel_tables(path)
    if path.suffix.casefold() in {".csv", ".txt", ".dat", ".xy", ""}:
        return _text_table(path)
    return []


def _header_candidates(table: _Table) -> list[tuple[int, dict[str, int]]]:
    candidates: list[tuple[int, dict[str, int]]] = []
    for row_index, row in enumerate(table.rows[:40]):
        mapping: dict[str, int] = {}
        duplicate = False
        for column_index, value in enumerate(row):
            field = _field_for_header(value)
            if field is None:
                continue
            if field in mapping:
                duplicate = True
                continue
            mapping[field] = column_index
        if {"ref", "x", "y"} <= set(mapping) and not duplicate:
            candidates.append((row_index, mapping))
    return candidates


def _unit_from_headers(row: Sequence[object], mapping: dict[str, int]) -> tuple[str | None, str | None]:
    units: set[str] = set()
    for field in ("x", "y"):
        match = _UNIT_RE.search(str(row[mapping[field]] or ""))
        if match:
            raw = match.group(1).casefold()
            units.add("mil" if raw in {"mil", "mils"} else "inch" if raw in {"inch", "in"} else "mm")
    if len(units) == 1:
        unit = next(iter(units))
        return unit, unit
    if len(units) > 1:
        return "conflicting", None
    return None, None


def _side(raw: object) -> BoardSide:
    value = str(raw or "").strip().casefold()
    if value in {"top", "t", "front", "component", "正面", "顶层"}:
        return "top"
    if value in {"bottom", "bot", "b", "back", "solder", "反面", "底层"}:
        return "bottom"
    return "unknown"


def _set_id(path: Path, section: str) -> str:
    digest = hashlib.sha256(f"{path.resolve()}|{section}".casefold().encode("utf-8")).hexdigest()
    return f"coords-tabular-{digest[:20]}"


def _source_id(path: Path) -> str:
    digest = hashlib.sha256(str(path.resolve()).casefold().encode("utf-8")).hexdigest()
    return f"source-local-{digest[:20]}"


class TabularCoordinateAdapter:
    adapter_id = "tabular_coordinates_v1"

    def probe(self, path: Path) -> list[CoordinateProbe]:
        try:
            tables = _tables(path)
        except (OSError, ValueError):
            return []
        probes: list[CoordinateProbe] = []
        for table in tables:
            candidates = _header_candidates(table)
            for row_index, mapping in candidates:
                confidence = 86 if len(candidates) == 1 else 72
                optional = sorted(set(mapping) - {"ref", "x", "y"})
                probes.append(
                    CoordinateProbe(
                        adapter_id=self.adapter_id,
                        confidence=confidence,
                        reasons=(
                            f"第 {row_index + 1} 行包含唯一位号/X/Y 字段",
                            f"可选字段：{', '.join(optional) if optional else '无'}",
                        ),
                        sheet_or_section=table.section,
                    )
                )
        return probes

    def parse(self, path: Path, probe: CoordinateProbe) -> SmtCoordinateSet:
        table = next(
            (item for item in _tables(path) if item.section == probe.sheet_or_section),
            None,
        )
        if table is None:
            raise ValueError("坐标工作表或数据区不存在")
        candidates = _header_candidates(table)
        if len(candidates) != 1:
            raise ValueError("坐标文件存在多个表头候选，请先确认数据区")
        header_index, mapping = candidates[0]
        declared_unit, normalized_unit = _unit_from_headers(table.rows[header_index], mapping)
        unit_state = "conflicting" if declared_unit == "conflicting" else "declared" if normalized_unit else "unknown"
        scale = {"mm": 1.0, "mil": 0.0254, "inch": 25.4}.get(normalized_unit or "")
        issues: list[SmtQualityIssue] = []
        if unit_state == "unknown":
            issues.append(
                SmtQualityIssue(
                    code="coordinate_unit_unknown",
                    severity="blocking",
                    message="坐标单位未声明，系统不会把数值静默当作毫米。",
                )
            )
        elif unit_state == "conflicting":
            issues.append(
                SmtQualityIssue(
                    code="coordinate_unit_conflicting",
                    severity="blocking",
                    message="X/Y 坐标表头声明了不同单位。",
                )
            )

        occurrences: list[SmtCoordinateOccurrence] = []
        rejected = 0
        unnamed = 0
        ref_counts: dict[str, int] = {}
        set_id = _set_id(path, probe.sheet_or_section)
        for source_index, row in enumerate(table.rows[header_index + 1 :], start=header_index + 2):
            if len(row) <= max(mapping.values()):
                rejected += 1
                continue
            raw_ref = str(row[mapping["ref"]] or "").strip()
            if not raw_ref:
                unnamed += 1
                continue
            raw_x = str(row[mapping["x"]] or "").strip()
            raw_y = str(row[mapping["y"]] or "").strip()
            try:
                x_value = float(raw_x)
                y_value = float(raw_y)
            except ValueError:
                rejected += 1
                issues.append(
                    SmtQualityIssue(
                        code="invalid_coordinate_row",
                        severity="warning",
                        message=f"第 {source_index} 行坐标不是有效数值。",
                        source_location=f"{probe.sheet_or_section or path.name}!{source_index}",
                    )
                )
                continue
            normalized_ref = raw_ref.upper()
            ref_counts[normalized_ref] = ref_counts.get(normalized_ref, 0) + 1
            raw_side = str(row[mapping["side"]] or "").strip() if "side" in mapping else ""
            raw_rotation = (
                str(row[mapping["rotation"]] or "").strip()
                if "rotation" in mapping
                else ""
            )
            rotation: float | None
            try:
                rotation = float(raw_rotation) % 360 if raw_rotation else None
            except ValueError:
                rotation = None
            occurrences.append(
                SmtCoordinateOccurrence(
                    occurrence_id=f"{set_id}-{source_index}",
                    raw_ref=raw_ref,
                    ref=normalized_ref,
                    raw_x=raw_x,
                    raw_y=raw_y,
                    normalized_x=x_value * scale if scale is not None else None,
                    normalized_y=y_value * scale if scale is not None else None,
                    raw_side=raw_side,
                    side=_side(raw_side),
                    raw_rotation=raw_rotation,
                    normalized_rotation=rotation,
                    footprint=(
                        str(row[mapping["footprint"]] or "").strip()
                        if "footprint" in mapping
                        else ""
                    ),
                    source_line=source_index,
                    warnings=[],
                )
            )
        duplicates = sorted(ref for ref, count in ref_counts.items() if count > 1)
        issues.extend(
            SmtQualityIssue(
                code="duplicate_coordinate_ref",
                severity="blocking",
                message=f"位号 {ref} 在坐标数据中出现多次。",
            )
            for ref in duplicates
        )
        return SmtCoordinateSet(
            coordinate_set_id=set_id,
            source_asset_id=_source_id(path),
            adapter_id=self.adapter_id,
            sheet_or_section=probe.sheet_or_section,
            declared_unit=None if declared_unit in {None, "conflicting"} else declared_unit,
            normalized_unit=normalized_unit,
            unit_state=unit_state,
            scope_semantics="unknown",
            side_mapping={},
            rotation_semantics="unknown",
            quality_report=SmtCoordinateQuality(
                valid_rows=len(occurrences),
                rejected_rows=rejected,
                unnamed_rows=unnamed,
                duplicate_refs=duplicates,
                issues=issues,
            ),
            occurrences=occurrences,
        )
