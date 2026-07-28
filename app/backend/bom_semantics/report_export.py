from __future__ import annotations

import json
from pathlib import Path
from typing import Iterable, Mapping

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.backend.bom_semantics.contracts import CompareResult


REPORT_SHEETS = (
    "对比摘要",
    "业务事件",
    "实际贴装差异",
    "替代关系差异",
    "板级元数据差异",
    "普通字段差异",
    "原始行差异",
    "风险与阻断项",
)


def _cell_value(value: object) -> object:
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    return json.dumps(value, ensure_ascii=False, sort_keys=True, default=str)


def _write_table(
    worksheet: object,
    headers: tuple[str, ...],
    rows: Iterable[Mapping[str, object]],
) -> None:
    worksheet.append(list(headers))
    for row in rows:
        worksheet.append([_cell_value(row.get(header)) for header in headers])
    for cell in worksheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(fill_type="solid", fgColor="1F4E78")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    for index, header in enumerate(headers, start=1):
        lengths = [len(str(header))]
        lengths.extend(
            min(len(str(worksheet.cell(row, index).value or "")), 60)
            for row in range(2, worksheet.max_row + 1)
        )
        width = max(lengths)
        worksheet.column_dimensions[get_column_letter(index)].width = min(max(width + 2, 10), 62)
        for row in range(2, worksheet.max_row + 1):
            cell = worksheet.cell(row, index)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if "编码" in header:
                cell.number_format = "@"


def _summary_rows(result: CompareResult) -> list[Mapping[str, object]]:
    summary = result.summary.payload()
    labels = {
        "parent_count_old": "旧版父项数",
        "parent_count_new": "新版父项数",
        "material_count_old": "旧版物料种类数",
        "material_count_new": "新版物料种类数",
        "actual_reference_count_old": "旧版实际位号数",
        "actual_reference_count_new": "新版实际位号数",
        "placement_change_group_count": "贴装变化组数",
        "placement_changed_reference_count": "贴装变化位号数",
        "substitute_group_count_old": "旧版替代组数",
        "substitute_group_count_new": "新版替代组数",
        "changed_event_count": "全部事件数",
        "review_event_count": "需复核事件数",
        "metadata_event_count": "元数据事件数",
        "metadata_change_count": "非装配变更记录数",
        "metadata_field_count": "非装配字段变化数",
        "blocker_count": "阻断问题数",
        "blocking_record_count": "受影响阻断记录数",
        "event_counts": "差异分类统计",
    }
    return [
        {"指标": labels[key], "值": value}
        for key, value in summary.items()
    ]


def _event_rows(result: CompareResult) -> list[Mapping[str, object]]:
    return [
        {
            "事件ID": event.event_id,
            "父项编码": event.parent_code,
            "分类": event.kind.value,
            "影响": event.impact.value,
            "标题": event.title,
            "位号": list(event.references),
            "替代组编码": list(event.group_codes),
            "OA变更类型": event.oa_change_type,
            "旧值": dict(event.old_snapshot),
            "新值": dict(event.new_snapshot),
            "阻断原因": list(event.blocker_reasons),
        }
        for event in result.events
    ]


def _finding_rows(result: CompareResult) -> list[Mapping[str, object]]:
    return [
        {
            "级别": finding.severity.value,
            "代码": finding.code,
            "父项编码": finding.parent_code,
            "位号": list(finding.references),
            "消息": finding.message,
            "详情": dict(finding.details),
            "来源": list(finding.source_ids),
        }
        for finding in (*result.blockers, *result.warnings)
    ]


def export_compare_report(result: CompareResult, path: Path) -> Path:
    destination = Path(path)
    destination.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    workbook.remove(workbook.active)

    summary = workbook.create_sheet(REPORT_SHEETS[0])
    _write_table(summary, ("指标", "值"), _summary_rows(result))

    events = workbook.create_sheet(REPORT_SHEETS[1])
    _write_table(
        events,
        ("事件ID", "父项编码", "分类", "影响", "标题", "位号", "替代组编码", "OA变更类型", "旧值", "新值", "阻断原因"),
        _event_rows(result),
    )

    placements = workbook.create_sheet(REPORT_SHEETS[2])
    _write_table(
        placements,
        (
            "group_id",
            "parent_code",
            "references",
            "reference_count",
            "status",
            "old_material_code",
            "new_material_code",
        ),
        result.placement_groups,
    )

    substitutes = workbook.create_sheet(REPORT_SHEETS[3])
    _write_table(substitutes, ("status", "old", "new"), result.substitute_diff)

    board_metadata = workbook.create_sheet(REPORT_SHEETS[4])
    _write_table(
        board_metadata,
        ("comparison_parent_code", "changed_fields", "old", "new"),
        result.board_metadata_diff,
    )

    metadata = workbook.create_sheet(REPORT_SHEETS[5])
    _write_table(
        metadata,
        (
            "parent_code",
            "material_code",
            "changed_fields",
            "old_variants",
            "new_variants",
            "old_metadata",
            "new_metadata",
        ),
        result.metadata_diff,
    )

    raw_rows = workbook.create_sheet(REPORT_SHEETS[6])
    _write_table(
        raw_rows,
        ("parent_code", "material_code", "status", "old_rows", "new_rows", "old_source_ids", "new_source_ids"),
        result.raw_row_diff,
    )

    findings = workbook.create_sheet(REPORT_SHEETS[7])
    _write_table(
        findings,
        ("级别", "代码", "父项编码", "位号", "消息", "详情", "来源"),
        _finding_rows(result),
    )
    workbook.save(destination)
    workbook.close()
    verify_compare_report(destination, result)
    return destination


def verify_compare_report(path: Path, result: CompareResult) -> None:
    workbook = load_workbook(path, data_only=True, read_only=False)
    try:
        if tuple(workbook.sheetnames) != REPORT_SHEETS:
            raise ValueError("BOM 对比报告工作表结构不完整。")
        event_rows = max(workbook["业务事件"].max_row - 1, 0)
        if event_rows != len(result.events):
            raise ValueError("BOM 对比报告事件数量与结果不一致。")
        placement_rows = max(workbook["实际贴装差异"].max_row - 1, 0)
        if placement_rows != len(result.placement_groups):
            raise ValueError("BOM 对比报告实际贴装差异数量不一致。")
        summary_values = {
            str(row[0].value): row[1].value
            for row in workbook["对比摘要"].iter_rows(min_row=2, max_col=2)
        }
        if summary_values.get("全部事件数") != result.summary.changed_event_count:
            raise ValueError("BOM 对比报告摘要统计不一致。")
    finally:
        workbook.close()
