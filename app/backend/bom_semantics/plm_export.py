from __future__ import annotations

from copy import copy
from dataclasses import dataclass
from decimal import Decimal
import json
from pathlib import Path
from typing import Iterable, Mapping, Sequence

from openpyxl import load_workbook
from openpyxl.worksheet.cell_range import CellRange

from app.backend.bom_semantics.models import (
    BoardBOM,
    CanonicalRow,
    FindingSeverity,
    SubstituteGroup,
    ValidationFinding,
    WorkbookProfile,
)
from app.backend.bom_semantics.normalization import normalize_workbook
from app.backend.bom_semantics.substitutes import build_board_boms
from app.backend.bom_semantics.validation import (
    validate_material_membership,
    validate_substitute_members,
    validate_unique_placements,
)
from app.backend.bom_semantics.workbook_reader import read_workbook_envelope


class PLMExportError(RuntimeError):
    """Raised when a semantic BOM cannot be rendered into a PLM workbook."""


class PLMExportBlockedError(PLMExportError):
    """Raised before writing when BoardBOM validation contains blockers."""

    def __init__(self, findings: Sequence[ValidationFinding]) -> None:
        self.findings = tuple(findings)
        summary = "; ".join(
            f"{finding.code}({finding.parent_code or 'unknown'})"
            for finding in self.findings
        )
        super().__init__(f"PLM export is blocked: {summary}")


class PLMExportVerificationError(PLMExportError):
    """Raised when reopening an exported workbook changes its BOM semantics."""


@dataclass(frozen=True)
class PLMExportResult:
    output_path: Path
    template_path: Path
    data_sheet: str
    rows_written: int
    parent_codes: tuple[str, ...]
    output_fingerprint: str


@dataclass(frozen=True)
class _RowStyle:
    styles: tuple[object, ...]
    height: float | None
    hidden: bool
    outline_level: int
    collapsed: bool


_TEXT_IDENTIFIER_FIELDS = {
    "parent_code",
    "material_code",
    "substitute_group_code",
}

_ROW_FIELDS = {
    "item": "item",
    "parent_code": "parent_code",
    "parent_description": "parent_description",
    "hardware_version": "hardware_version",
    "material_code": "material_code",
    "name": "name",
    "model": "model",
    "description": "description",
    "unit": "unit",
    "quantity": "quantity",
    "remark": "remark",
    "grade": "grade",
    "grade_remark": "grade_remark",
    "substitute_group_code": "substitute_group_code",
    "substitute_strategy": "substitute_strategy",
    "substitute_mode": "substitute_mode",
    "substitute_priority": "substitute_priority",
    "issue_method": "issue_method",
    "mrp": "mrp",
    "jump_level": "jump_level",
}
_EXTRA_FIELDS = {
    "manufacturer",
    "pcb_footprint",
    "pcb_package",
    "change_type",
    "change_status",
    "affected_bom",
    "highest_bom",
    "project",
    "level",
}


def _coerce_boards(boards: BoardBOM | Sequence[BoardBOM]) -> tuple[BoardBOM, ...]:
    if isinstance(boards, BoardBOM):
        resolved = (boards,)
    else:
        resolved = tuple(boards)
    if not resolved:
        raise PLMExportError("PLM export requires at least one BoardBOM.")
    return resolved


def _blocker(
    code: str,
    message: str,
    *,
    parent_code: str = "",
    references: Iterable[str] = (),
    details: Mapping[str, object] | None = None,
) -> ValidationFinding:
    return ValidationFinding(
        code=code,
        severity=FindingSeverity.BLOCKER,
        message=message,
        parent_code=parent_code,
        references=tuple(references),
        details=details or {},
    )


def collect_plm_export_blockers(
    boards: BoardBOM | Sequence[BoardBOM],
) -> tuple[ValidationFinding, ...]:
    """Revalidate the model at the output boundary instead of trusting UI state."""

    resolved = _coerce_boards(boards)
    findings: list[ValidationFinding] = []
    parent_codes: set[str] = set()
    all_groups: list[SubstituteGroup] = []

    for board in resolved:
        if not board.parent_code:
            findings.append(_blocker("parent_code_missing", "BoardBOM has no parent code."))
        elif board.parent_code in parent_codes:
            findings.append(
                _blocker(
                    "parent_code_duplicate",
                    "Each BoardBOM must have a unique parent code.",
                    parent_code=board.parent_code,
                )
            )
        parent_codes.add(board.parent_code)
        findings.extend(
            finding
            for finding in board.findings
            if finding.severity == FindingSeverity.BLOCKER
        )

        all_groups.extend(board.substitute_groups)
        placements = [
            (placement.parent_code, placement.reference, placement.material_code, placement.source_ids)
            for placement in board.placements
        ]
        findings.extend(validate_unique_placements(placements))

        placement_index = {
            (placement.reference, placement.material_code, placement.substitute_group_code)
            for placement in board.placements
        }
        for group in board.substitute_groups:
            findings.extend(
                finding
                for finding in group.validation_findings
                if finding.severity == FindingSeverity.BLOCKER
            )
            findings.extend(
                validate_substitute_members(group.parent_code, group.group_code, group.members)
            )
            if group.main_item is None:
                continue
            main_references = set(group.main_item.references)
            if set(group.physical_references) != main_references:
                findings.append(
                    _blocker(
                        "substitute_group_physical_references_mismatch",
                        "Substitute group physical references must equal its main item references.",
                        parent_code=board.parent_code,
                        references=group.physical_references,
                        details={"group_code": group.group_code},
                    )
                )
            missing = [
                reference
                for reference in group.main_item.references
                if (reference, group.main_item.material_code, group.group_code) not in placement_index
            ]
            if missing:
                findings.append(
                    _blocker(
                        "substitute_main_placement_missing",
                        "Every substitute main reference must be represented by a physical placement.",
                        parent_code=board.parent_code,
                        references=missing,
                        details={
                            "group_code": group.group_code,
                            "material_code": group.main_item.material_code,
                        },
                    )
                )

    findings.extend(validate_material_membership(all_groups))
    seen: set[tuple[object, ...]] = set()
    unique: list[ValidationFinding] = []
    for finding in findings:
        key = (
            finding.code,
            finding.parent_code,
            finding.references,
            json.dumps(finding.details, ensure_ascii=False, default=str, sort_keys=True),
        )
        if key not in seen:
            seen.add(key)
            unique.append(finding)
    return tuple(unique)


def _sorted_rows(boards: Sequence[BoardBOM]) -> tuple[CanonicalRow, ...]:
    return tuple(
        row
        for board in boards
        for row in sorted(board.rows, key=lambda value: (value.sheet_name, value.row_number, value.source_id))
    )


def _field_value(row: CanonicalRow, field: str) -> object:
    if field == "reference":
        if row.references:
            return ",".join(row.references)
        return row.raw_reference if row.is_nc else ""
    attribute = _ROW_FIELDS.get(field)
    if attribute is not None:
        return getattr(row, attribute)
    if field in _EXTRA_FIELDS:
        return row.extra_fields.get(field) or row.raw_fields.get(field, "")
    return row.raw_fields.get(field, "")


def _row_has_value(row: CanonicalRow, field: str) -> bool:
    value = _field_value(row, field)
    return value not in (None, "")


def _validate_template_coverage(
    rows: Sequence[CanonicalRow],
    field_columns: Mapping[str, int],
) -> None:
    semantic_fields = tuple(_ROW_FIELDS) + tuple(_EXTRA_FIELDS) + ("reference",)
    missing = sorted(
        field
        for field in semantic_fields
        if field not in field_columns and any(_row_has_value(row, field) for row in rows)
    )
    if missing:
        raise PLMExportError(
            "The PLM template cannot preserve populated semantic fields: " + ", ".join(missing)
        )


def _capture_row_style(worksheet: object, row_number: int) -> _RowStyle:
    dimension = worksheet.row_dimensions[row_number]
    return _RowStyle(
        styles=tuple(copy(worksheet.cell(row_number, column)._style) for column in range(1, worksheet.max_column + 1)),
        height=dimension.height,
        hidden=bool(dimension.hidden),
        outline_level=int(dimension.outlineLevel or 0),
        collapsed=bool(dimension.collapsed),
    )


def _apply_row_style(worksheet: object, row_number: int, style: _RowStyle) -> None:
    for column, source_style in enumerate(style.styles, start=1):
        worksheet.cell(row_number, column)._style = copy(source_style)
    dimension = worksheet.row_dimensions[row_number]
    dimension.height = style.height
    dimension.hidden = style.hidden
    dimension.outlineLevel = style.outline_level
    dimension.collapsed = style.collapsed


def _intersects_data_area(cell_range: CellRange, start_row: int, end_row: int) -> bool:
    return cell_range.min_row <= end_row and cell_range.max_row >= start_row


def _shift_merged_ranges_for_insert(worksheet: object, insert_at: int, amount: int) -> None:
    original = tuple(CellRange(str(cell_range)) for cell_range in worksheet.merged_cells.ranges)
    crossing = [
        cell_range
        for cell_range in original
        if cell_range.min_row < insert_at <= cell_range.max_row
    ]
    if crossing:
        raise PLMExportError(
            "The PLM template has a merged range crossing the expandable data boundary: "
            + ", ".join(str(cell_range) for cell_range in crossing)
        )
    if not any(cell_range.min_row >= insert_at for cell_range in original):
        return
    for cell_range in tuple(str(item) for item in worksheet.merged_cells.ranges):
        worksheet.unmerge_cells(cell_range)
    for cell_range in original:
        if cell_range.min_row >= insert_at:
            cell_range.shift(row_shift=amount)
        worksheet.merge_cells(str(cell_range))


def _prepare_data_area(
    worksheet: object,
    *,
    data_start_row: int,
    data_end_row: int,
    row_count: int,
) -> None:
    existing_count = max(0, data_end_row - data_start_row + 1)
    if any(
        _intersects_data_area(CellRange(str(cell_range)), data_start_row, max(data_end_row, data_start_row))
        for cell_range in worksheet.merged_cells.ranges
    ):
        raise PLMExportError("The PLM template merges cells inside its semantic data area.")
    style_row = data_start_row if existing_count else max(1, data_start_row)
    style = _capture_row_style(worksheet, style_row)
    if row_count > existing_count:
        amount = row_count - existing_count
        insert_at = data_start_row + existing_count
        _shift_merged_ranges_for_insert(worksheet, insert_at, amount)
        worksheet.insert_rows(insert_at, amount)
        for row_number in range(insert_at, insert_at + amount):
            _apply_row_style(worksheet, row_number, style)

    clear_through = max(existing_count, row_count)
    for row_number in range(data_start_row, data_start_row + clear_through):
        for column in range(1, worksheet.max_column + 1):
            cell = worksheet.cell(row_number, column)
            cell.value = None
            cell.hyperlink = None
            cell.comment = None


def _excel_quantity(value: Decimal | None) -> int | float | None:
    if value is None:
        return None
    if value == value.to_integral_value():
        return int(value)
    return float(value)


def _write_rows(
    worksheet: object,
    rows: Sequence[CanonicalRow],
    *,
    start_row: int,
    field_columns: Mapping[str, int],
) -> None:
    for row_offset, row in enumerate(rows):
        excel_row = start_row + row_offset
        for field, column in field_columns.items():
            cell = worksheet.cell(excel_row, column)
            value = _field_value(row, field)
            if field == "quantity":
                cell.value = _excel_quantity(value if isinstance(value, Decimal) else None)
            elif field in _TEXT_IDENTIFIER_FIELDS:
                cell.number_format = "@"
                cell.value = "" if value is None else str(value)
            else:
                cell.value = value


def _variant_snapshot(item: object) -> tuple[tuple[str, ...], ...]:
    return tuple(sorted(variant.signature for variant in item.variants))


def _board_snapshot(board: BoardBOM) -> dict[str, object]:
    return {
        "parent_code": board.parent_code,
        "parent_description": board.parent_description,
        "hardware_version": board.hardware_version,
        "rows": sorted(
            (
                (
                    row.parent_code,
                    row.parent_description,
                    row.hardware_version,
                    row.material_code,
                    row.name,
                    row.model,
                    row.description,
                    row.unit,
                    str(row.quantity) if row.quantity is not None else None,
                    row.references,
                    row.is_nc,
                    row.remark,
                    row.grade,
                    row.grade_remark,
                    row.substitute_group_code,
                    row.substitute_strategy,
                    row.substitute_mode,
                    row.substitute_priority,
                    row.issue_method,
                    row.mrp,
                    row.jump_level,
                    tuple(sorted(row.extra_fields.items())),
                )
                for row in board.rows
            ),
            key=repr,
        ),
        "items": sorted(
            (
                (
                    item.material_code,
                    str(item.quantity) if item.quantity is not None else None,
                    item.references,
                    item.substitute_group_code,
                    item.substitute_priority,
                    _variant_snapshot(item),
                )
                for item in board.items
            ),
            key=repr,
        ),
        "groups": sorted(
            (
                (
                    group.group_code,
                    group.main_item.material_code if group.main_item else "",
                    tuple(item.material_code for item in group.alternative_items),
                    group.physical_references,
                    str(group.quantity) if group.quantity is not None else None,
                )
                for group in board.substitute_groups
            ),
            key=repr,
        ),
        "placements": sorted(
            (
                (
                    placement.reference,
                    placement.material_code,
                    placement.substitute_group_code,
                    placement.is_nc,
                )
                for placement in board.placements
            ),
            key=repr,
        ),
        "non_placement": sorted(
            (
                (
                    item.material_code,
                    str(item.quantity) if item.quantity is not None else None,
                    item.reason,
                )
                for item in board.non_placement_items
            ),
            key=repr,
        ),
    }


def verify_plm_export(
    output_path: Path,
    expected_boards: BoardBOM | Sequence[BoardBOM],
) -> tuple[BoardBOM, ...]:
    """Reopen the output and prove it still represents the same BOM semantics."""

    expected = _coerce_boards(expected_boards)
    normalized = normalize_workbook(Path(output_path))
    if normalized.has_blockers:
        raise PLMExportVerificationError("Reopening the PLM export produced source blockers.")
    actual = build_board_boms(normalized)
    blockers = collect_plm_export_blockers(actual)
    if blockers:
        raise PLMExportVerificationError("Reopening the PLM export produced semantic blockers.")
    expected_snapshots = sorted(
        (_board_snapshot(board) for board in expected),
        key=lambda value: value["parent_code"],
    )
    actual_snapshots = sorted(
        (_board_snapshot(board) for board in actual),
        key=lambda value: value["parent_code"],
    )
    if expected_snapshots != actual_snapshots:
        raise PLMExportVerificationError("PLM export round-trip changed the BoardBOM semantic model.")
    return actual


def export_plm_template(
    boards: BoardBOM | Sequence[BoardBOM],
    template_path: Path,
    output_path: Path,
) -> PLMExportResult:
    """Render semantic BOM rows into a user PLM template or source workbook.

    ``template_path`` is intentionally explicit: it can be either a blank PLM
    template or the source PLM workbook, while ``output_path`` always remains a
    separate file so the user-provided workbook is never overwritten.
    """

    resolved = _coerce_boards(boards)
    blockers = collect_plm_export_blockers(resolved)
    if blockers:
        raise PLMExportBlockedError(blockers)

    template = Path(template_path).resolve()
    output = Path(output_path).resolve()
    if template == output:
        raise PLMExportError("Output path must differ from the user template path.")
    envelope = read_workbook_envelope(template)
    if envelope.profile not in {WorkbookProfile.PLM_SINGLE_BOARD, WorkbookProfile.PLM_MULTI_BOARD}:
        raise PLMExportError("The selected workbook is not a recognized PLM template.")
    source_sheet = next((sheet for sheet in envelope.sheets if sheet.name == envelope.data_sheet), None)
    if source_sheet is None or not source_sheet.field_columns:
        raise PLMExportError("The PLM template does not have a writable BOM data sheet.")
    if not {"parent_code", "material_code", "quantity"}.issubset(source_sheet.field_columns):
        raise PLMExportError("The PLM template is missing required parent, material, or quantity columns.")

    rows = _sorted_rows(resolved)
    _validate_template_coverage(rows, source_sheet.field_columns)
    normalized_template = normalize_workbook(template)
    semantic_data_end = max(
        (
            row.row_number
            for row in normalized_template.rows
            if row.sheet_name == source_sheet.name
            and (row.quantity is not None or bool(row.references))
        ),
        default=source_sheet.data_start_row - 1,
    )
    output.parent.mkdir(parents=True, exist_ok=True)
    workbook = load_workbook(template)
    try:
        worksheet = workbook[source_sheet.name]
        _prepare_data_area(
            worksheet,
            data_start_row=source_sheet.data_start_row,
            data_end_row=semantic_data_end,
            row_count=len(rows),
        )
        _write_rows(
            worksheet,
            rows,
            start_row=source_sheet.data_start_row,
            field_columns=source_sheet.field_columns,
        )
        workbook.save(output)
    finally:
        workbook.close()

    verify_plm_export(output, resolved)
    exported_envelope = read_workbook_envelope(output)
    return PLMExportResult(
        output_path=output,
        template_path=template,
        data_sheet=source_sheet.name,
        rows_written=len(rows),
        parent_codes=tuple(board.parent_code for board in resolved),
        output_fingerprint=exported_envelope.source_fingerprint,
    )
